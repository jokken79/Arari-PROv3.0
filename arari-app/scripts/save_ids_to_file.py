#!/usr/bin/env python
"""Find ALL employee ID positions and save to file"""
import openpyxl
from io import BytesIO

file_path = r"D:\給料明細\Kyuryo\給与明細(派遣社員)2025.1(0217支給).xlsm"
output_file = r"d:\Arari-PROv1.0\Arari-PRO\arari-app\scripts\employee_ids_found.txt"

with open(output_file, 'w', encoding='utf-8') as out:
    out.write(f"Opening: {file_path}\n")
    out.write("=" * 80 + "\n")
    
    with open(file_path, 'rb') as f:
        content = f.read()
    
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    
    skip_sheets = ['集計', 'Summary', '目次', 'Index', '請負']
    data_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name not in skip_sheets:
            data_sheet = sheet_name
            break
    
    out.write(f"\nAnalyzing sheet: '{data_sheet}'\n")
    out.write("=" * 80 + "\n")
    
    ws = wb[data_sheet]
    
    # Scan row 6 for all 6-digit employee IDs
    out.write(f"\nScanning ROW 6 for employee IDs (6-digit numbers):\n")
    out.write("-" * 80 + "\n")
    
    found_ids = []
    for col in range(1, 100):
        cell_value = ws.cell(row=6, column=col).value
        if cell_value is None:
            continue
        
        try:
            val_str = str(cell_value).strip()
            if val_str.isdigit() and len(val_str) == 6:
                found_ids.append((col, val_str))
                out.write(f"  Col {col:3d}: {val_str}\n")
        except:
            pass
    
    out.write(f"\n{'='*80}\n")
    out.write(f"TOTAL: Found {len(found_ids)} employee IDs in row 6\n")
    out.write(f"{'='*80}\n")
    
    if found_ids:
        if len(found_ids) >= 2:
            spacing = found_ids[1][0] - found_ids[0][0]
            out.write(f"\nColumn spacing: {spacing} columns between employees\n")
            out.write(f"First employee at column: {found_ids[0][0]}\n")
            
        out.write(f"\nParser's base_col calculation (offset=9):\n")
        for col, emp_id in found_ids[:5]:
            base_col = col - 9
            out.write(f"  Employee {emp_id}: col {col} - 9 = base_col {base_col}")
            if base_col <= 0:
                out.write(f" ⚠️ INVALID!")
            out.write("\n")
    else:
        out.write("\n⚠️ NO 6-DIGIT IDs FOUND\n")

print(f"Results saved to: {output_file}")
print("Done!")
