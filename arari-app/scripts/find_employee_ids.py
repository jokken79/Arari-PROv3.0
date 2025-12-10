#!/usr/bin/env python
"""Find ALL employee ID positions in the Excel file"""
import openpyxl
from io import BytesIO
import sys

file_path = r"D:\給料明細\Kyuryo\給与明細(派遣社員)2025.1(0217支給).xlsm"

print(f"Opening: {file_path}")
sys.stdout.flush()

with open(file_path, 'rb') as f:
    content = f.read()

wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

skip_sheets = ['集計', 'Summary', '目次', 'Index', '請負']
data_sheet = None
for sheet_name in wb.sheetnames:
    if sheet_name not in skip_sheets:
        data_sheet = sheet_name
        break

print(f"\nAnalyzing sheet: '{data_sheet}'")
print("=" * 80)
sys.stdout.flush()

ws = wb[data_sheet]

# Scan row 6 for all 6-digit employee IDs
print(f"\nScanning ROW 6 for employee IDs (6-digit numbers):")
print("-" * 80)
sys.stdout.flush()

found_ids = []
for col in range(1, 100):  # Check first 100 columns
    cell_value = ws.cell(row=6, column=col).value
    if cell_value is None:
        continue
    
    try:
        val_str = str(cell_value).strip()
        if val_str.isdigit() and len(val_str) == 6:
            found_ids.append((col, val_str))
            print(f"  Col {col:3d}: {val_str}")
            sys.stdout.flush()
    except:
        pass

print(f"\n{'='*80}")
print(f"TOTAL: Found {len(found_ids)} employee IDs in row 6")
print(f"{'='*80}")
sys.stdout.flush()

if found_ids:
    # Calculate column spacing
    if len(found_ids) >= 2:
        spacing = found_ids[1][0] - found_ids[0][0]
        print(f"\nColumn spacing: {spacing} columns between employees")
        print(f"First employee at column: {found_ids[0][0]}")
        sys.stdout.flush()
        
    # Show what the parser would extract (base_col calculation)
    print(f"\nParser's base_col calculation:")
    print(f"  employee_id offset = 9 (default)")
    for col, emp_id in found_ids[:3]:  # Show first 3
        base_col = col - 9
        print(f"  Employee {emp_id}: col {col} - 9 = base_col {base_col}")
        if base_col <= 0:
            print(f"    ⚠️  WARNING: base_col {base_col} is invalid (<=0)!")
        sys.stdout.flush()
else:
    print("\n⚠️  NO 6-DIGIT IDs FOUND")
    print("\nShowing first 10 columns of row 6 to debug:")
    for col in range(1, 11):
        cell_value = ws.cell(row=6, column=col).value
        print(f"  Col {col}: {repr(cell_value)}")
        sys.stdout.flush()

print("\n" + "=" * 80)
sys.stdout.flush()
