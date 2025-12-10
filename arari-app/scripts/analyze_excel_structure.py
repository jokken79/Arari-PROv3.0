#!/usr/bin/env python
"""Analyze Excel structure to find employee ID positions"""
import openpyxl
from io import BytesIO

file_path = r"D:\給料明細\Kyuryo\給与明細(派遣社員)2025.1(0217支給).xlsm"

print(f"Analyzing: {file_path}")
print("=" * 100)

with open(file_path, 'rb') as f:
    content = f.read()

wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

# Check first actual data sheet (skip summary sheets)
skip_sheets = ['集計', 'Summary', '目次', 'Index', '請負']
data_sheet = None
for sheet_name in wb.sheetnames:
    if sheet_name not in skip_sheets:
        data_sheet = sheet_name
        break

if not data_sheet:
    print("No data sheet found!")
    exit(1)

print(f"\nAnalyzing sheet: '{data_sheet}'")
print("=" * 100)

ws = wb[data_sheet]

# Show first 15 rows, first 30 columns
print("\nFIRST 15 ROWS x 30 COLUMNS:")
print("-" * 100)

for row in range(1, 16):
    row_data = []
    for col in range(1, 31):
        cell_value = ws.cell(row=row, column=col).value
        if cell_value is None:
            row_data.append("")
        else:
            # Truncate long values
            val_str = str(cell_value)
            if len(val_str) > 12:
                val_str = val_str[:9] + "..."
            row_data.append(val_str)
    
    # Print row number and values
    print(f"Row {row:2d}: {' | '.join(f'{v:12s}' for v in row_data[:15])}")

print("\n" + "=" * 100)
print("\nLOOKING FOR 6-DIGIT EMPLOYEE IDs:")
print("-" * 100)

found_ids = []
for row in range(1, 16):
    for col in range(1, 31):
        cell_value = ws.cell(row=row, column=col).value
        if cell_value is None:
            continue
        
        try:
            val_str = str(cell_value).strip()
            # Check for 6-digit number
            if val_str.isdigit() and len(val_str) == 6:
                found_ids.append({
                    'row': row,
                    'col': col,
                    'value': val_str
                })
                print(f"  ✓ Found at Row {row}, Col {col}: {val_str}")
        except:
            pass

if not found_ids:
    print("  ✗ NO 6-DIGIT IDs FOUND in rows 1-15, columns 1-30")
    print("\n  Checking for OTHER patterns (5-digit, 7-digit, alphanumeric)...")
    
    for row in range(1, 16):
        for col in range(1, 31):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is None:
                continue
            
            try:
                val_str = str(cell_value).strip()
                # Check for any digit sequence 4-8 chars
                if val_str.replace('.', '').replace('-', '').isdigit():
                    if 4 <= len(val_str) <= 8:
                        print(f"  ? Candidate at Row {row}, Col {col}: '{val_str}' (length {len(val_str)})")
            except:
                pass

print("\n" + "=" * 100)
