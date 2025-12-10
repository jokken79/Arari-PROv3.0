#!/usr/bin/env python
"""List all sheet names in the payroll Excel file"""
import openpyxl
from io import BytesIO

file_path = r"D:\給料明細\Kyuryo\給与明細(派遣社員)2025.1(0217支給).xlsm"

print(f"Opening: {file_path}")

with open(file_path, 'rb') as f:
    content = f.read()

wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

print(f"\nTotal sheets: {len(wb.sheetnames)}")
print("=" * 80)

for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"{i}. '{sheet_name}'")

print("=" * 80)

# Check which would be filtered
skip_list = ['集計', 'Summary', '目次', 'Index', '請負']
print(f"\nSheets that would be SKIPPED by current filter: {skip_list}")
for sheet_name in wb.sheetnames:
    if sheet_name in skip_list:
        print(f"  - '{sheet_name}' (SKIPPED)")

print(f"\nSheets that would be PROCESSED:")
for sheet_name in wb.sheetnames:
    if sheet_name not in skip_list:
        print(f"  + '{sheet_name}' (PROCESS)")
