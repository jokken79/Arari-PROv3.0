#!/usr/bin/env python
"""Replicate exact parser logic for employee column detection"""
import openpyxl
from io import BytesIO
import sys
sys.path.insert(0, r"d:\Arari-PROv1.0\Arari-PRO\arari-app\api")

from salary_parser import SalaryStatementParser

file_path = r"D:\給料明細\Kyuryo\給与明細(派遣社員)2025.1(0217支給).xlsm"
output_file = r"d:\Arari-PROv1.0\Arari-PRO\arari-app\scripts\parser_debug.txt"

with open(output_file, 'w', encoding='utf-8') as out:
    out.write(f"Opening: {file_path}\n")
    out.write("=" * 80 + "\n")
    
    with open(file_path, 'rb') as f:
        content = f.read()
    
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    
    skip_sheets = ['集計', 'Summary', '目次', 'Index', '請負']
    data_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name not in skip_sheets:
            data_sheet = sheet_name
            break
    
    out.write(f"\nAnalyzing sheet: '{data_sheet}'\n")
    out.write("=" * 80 + "\n")
    
    ws = wb[data_sheet]
    
    # Create parser instance
    parser = SalaryStatementParser(use_intelligent_mode=True)
    
    # Run field detection
    parser._detect_field_positions(ws)
    
    out.write(f"\nDETECTED FIELDS ({len(parser.detected_fields)}):\n")
    for field, row in sorted(parser.detected_fields.items()):
        out.write(f"  {field}: row {row}\n")
    
    out.write(f"\nDETECTED ALLOWANCES ({len(parser.detected_allowances)}):\n")
    for allow, row in sorted(parser.detected_allowances.items()):
        out.write(f"  {allow}: row {row}\n")
    
    # Get column offsets
    offsets = parser.current_column_offsets or parser.COLUMN_OFFSETS
    out.write(f"\nCOLUMN OFFSETS:\n")
    for key, val in offsets.items():
        out.write(f"  {key}: {val}\n")
    
    # Get emp_id_row
    emp_id_row = parser.detected_fields.get('employee_id') or parser.FALLBACK_ROW_POSITIONS.get('employee_id', 6)
    out.write(f"\nEMPLOYEE ID ROW: {emp_id_row}\n")
    out.write(f"EMPLOYEE ID OFFSET: {offsets.get('employee_id', 9)}\n")
    
    # Now call the actual _detect_employee_columns method
    out.write(f"\n{'='*80}\n")
    out.write(f"CALLING parser._detect_employee_columns()...\n")
    out.write(f"{'='*80}\n")
    
    employee_cols = parser._detect_employee_columns(ws)
    
    out.write(f"\nRESULT: Found {len(employee_cols)} employee columns\n")
    if employee_cols:
        out.write(f"Columns: {employee_cols}\n")
    else:
        out.write("⚠️ NO EMPLOYEE COLUMNS FOUND!\n")
        out.write("\nManual scan of row 6 for debugging:\n")
        for col in range(1, 100):
            cell_value = ws.cell(row=emp_id_row, column=col).value
            if cell_value:
                val_str = str(cell_value).strip()
                if val_str.isdigit() and len(val_str) == 6:
                    base_col = col - offsets.get('employee_id', 9)
                    out.write(f"  Col {col}: {val_str} → base_col={base_col}\n")

print(f"Results saved to: {output_file}")
