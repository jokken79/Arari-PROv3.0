import os

import openpyxl

template_path = r"d:\Arari-PROv1.0.25.12.11\Arari-PROv1.0\Arari-PRO\arari-app\api\templates\template_format_b.xlsx"

if not os.path.exists(template_path):
    print("Template not found")
    exit()

wb = openpyxl.load_workbook(template_path)
ws = wb.active

print("=== TEMPLATE ANALYSIS ===")
# Scan specific areas for keywords
keywords = [
    "氏名",
    "No",
    "部門",
    "入社",
    "1月",
    "4月",
    "基本",
    "残業",
    "総支給",
    "控除",
    "差引",
    "支給",
    "合計",
]

print("\n[KEYWORD SCAN]")
for row in range(1, 40):
    for col in range(1, 20):
        val = ws.cell(row=row, column=col).value
        if val:
            s_val = str(val).lower()
            cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
            # Print if it contains a keyword OR looks like a header (month)
            found = False
            for k in keywords:
                if k in s_val:
                    print(f"{cell_ref}: {s_val}")
                    found = True
                    break

            # Print row headers that might be important
            if not found and col <= 3 and len(s_val) < 10:
                print(f"{cell_ref} (Header?): {s_val}")

print("\n[MONTH ROW CHECK]")
# Assume headers are somewhere in rows 1-10
for row in range(1, 15):
    row_values = []
    for col in range(1, 20):
        v = ws.cell(row=row, column=col).value
        row_values.append(str(v) if v else ".")
    print(f"Row {row}: {', '.join(row_values)}")
