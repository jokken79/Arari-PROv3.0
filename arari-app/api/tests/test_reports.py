"""
Tests for Reports System
Tests report data aggregation and Excel generation.

Report Types:
- monthly: 月次粗利レポート
- all-employees: 従業員別詳細レポート
- all-companies: 派遣先別分析レポート
- cost-breakdown: コスト内訳レポート
- summary: 経営サマリーレポート
"""

import pytest
from io import BytesIO
from datetime import datetime


class TestMonthlyReportData:
    """Tests for ReportService.get_monthly_report_data()"""

    @pytest.fixture
    def report_service_with_data(self, db_session):
        """Initialize service with test data"""
        from reports import ReportService

        cursor = db_session.cursor()

        # Create test employees
        employees = [
            ("EMP001", "田中太郎", "タナカタロウ", "加藤木材", 1000, 1500, "active"),
            ("EMP002", "佐藤花子", "サトウハナコ", "加藤木材", 1100, 1600, "active"),
            ("EMP003", "鈴木一郎", "スズキイチロウ", "他社", 1200, 1700, "active"),
        ]

        for emp_id, name, kana, company, hourly, billing, status in employees:
            cursor.execute("""
                INSERT INTO employees
                (employee_id, name, name_kana, dispatch_company, hourly_rate, billing_rate, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (emp_id, name, kana, company, hourly, billing, status))

        # Create payroll records with proper margin calculations
        payroll_records = [
            # EMP001: High margin (12%+)
            ("EMP001", "2025年1月", 160, 20, 0, 10, 0, 300000, 340000, 295000, 45000, 13.24),
            # EMP002: Medium margin (10-12%)
            ("EMP002", "2025年1月", 160, 30, 0, 15, 0, 320000, 358000, 318000, 40000, 11.17),
            # EMP003: Low margin (<7%)
            ("EMP003", "2025年1月", 160, 10, 0, 5, 0, 280000, 300000, 285000, 15000, 5.00),
        ]

        for (emp_id, period, work_hours, overtime, overtime_60h, night, holiday,
             gross_salary, billing, total_cost, gross_profit, margin) in payroll_records:
            cursor.execute("""
                INSERT INTO payroll_records
                (employee_id, period, work_hours, overtime_hours, overtime_over_60h,
                 night_hours, holiday_hours, gross_salary, billing_amount,
                 total_company_cost, gross_profit, profit_margin)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (emp_id, period, work_hours, overtime, overtime_60h, night, holiday,
                  gross_salary, billing, total_cost, gross_profit, margin))

        db_session.commit()

        return ReportService(db_session)

    def test_monthly_report_structure(self, report_service_with_data):
        """Should return correct structure"""
        result = report_service_with_data.get_monthly_report_data("2025年1月")

        assert "period" in result
        assert "summary" in result
        assert "by_company" in result
        assert "top_performers" in result
        assert "bottom_performers" in result
        assert "generated_at" in result

    def test_monthly_report_summary_totals(self, report_service_with_data):
        """Should calculate correct totals"""
        result = report_service_with_data.get_monthly_report_data("2025年1月")

        summary = result["summary"]

        assert summary["employee_count"] == 3
        # 340000 + 358000 + 300000 = 998000
        assert summary["total_revenue"] == 998000
        # 295000 + 318000 + 285000 = 898000
        assert summary["total_cost"] == 898000
        # 45000 + 40000 + 15000 = 100000
        assert summary["total_profit"] == 100000

    def test_monthly_report_by_company(self, report_service_with_data):
        """Should group by company"""
        result = report_service_with_data.get_monthly_report_data("2025年1月")

        by_company = result["by_company"]

        assert len(by_company) == 2

        company_names = [c["company"] for c in by_company]
        assert "加藤木材" in company_names
        assert "他社" in company_names

    def test_monthly_report_top_performers(self, report_service_with_data):
        """Should identify top performers by margin"""
        result = report_service_with_data.get_monthly_report_data("2025年1月")

        top = result["top_performers"]

        # Should be ordered by margin DESC
        assert len(top) <= 5
        if len(top) > 0:
            # EMP001 has highest margin (13.24%)
            assert top[0]["employee_id"] == "EMP001"

    def test_monthly_report_bottom_performers(self, report_service_with_data):
        """Should identify bottom performers by margin"""
        result = report_service_with_data.get_monthly_report_data("2025年1月")

        bottom = result["bottom_performers"]

        # Should be ordered by margin ASC
        assert len(bottom) <= 5
        if len(bottom) > 0:
            # EMP003 has lowest margin (5%)
            assert bottom[0]["employee_id"] == "EMP003"

    def test_monthly_report_empty_period(self, db_session):
        """Should handle empty period gracefully"""
        from reports import ReportService

        service = ReportService(db_session)
        result = service.get_monthly_report_data("2030年1月")

        summary = result["summary"]
        assert summary["employee_count"] == 0
        # Empty results may return None or 0
        assert summary["total_revenue"] in (0, None)
        assert summary["total_profit"] in (0, None)


class TestAllEmployeesReportData:
    """Tests for ReportService.get_all_employees_report_data()"""

    @pytest.fixture
    def report_service_with_data(self, db_session):
        """Initialize service with test data"""
        from reports import ReportService

        cursor = db_session.cursor()

        # Create employees and payroll
        cursor.execute("""
            INSERT INTO employees
            (employee_id, name, name_kana, dispatch_company, hourly_rate, billing_rate, status)
            VALUES ('EMP001', '田中太郎', 'タナカタロウ', '加藤木材', 1000, 1500, 'active')
        """)

        cursor.execute("""
            INSERT INTO payroll_records
            (employee_id, period, work_hours, overtime_hours, gross_salary,
             billing_amount, total_company_cost, gross_profit, profit_margin)
            VALUES ('EMP001', '2025年1月', 160, 20, 300000, 340000, 295000, 45000, 13.24)
        """)

        db_session.commit()

        return ReportService(db_session)

    def test_all_employees_structure(self, report_service_with_data):
        """Should return correct structure"""
        result = report_service_with_data.get_all_employees_report_data("2025年1月")

        assert "period" in result
        assert "employees" in result
        assert "totals" in result
        assert "generated_at" in result

    def test_all_employees_includes_all_fields(self, report_service_with_data):
        """Should include all required fields for each employee"""
        result = report_service_with_data.get_all_employees_report_data("2025年1月")

        employees = result["employees"]
        assert len(employees) >= 1

        emp = employees[0]
        required_fields = [
            "employee_id", "name", "company", "work_hours",
            "billing_amount", "total_cost", "gross_profit", "profit_margin"
        ]

        for field in required_fields:
            assert field in emp


class TestAllCompaniesReportData:
    """Tests for ReportService.get_all_companies_report_data()"""

    @pytest.fixture
    def report_service_with_companies(self, db_session):
        """Initialize service with multiple companies"""
        from reports import ReportService

        cursor = db_session.cursor()

        # Create employees in different companies
        companies_data = [
            ("EMP001", "田中太郎", "加藤木材", 340000, 295000, 45000, 13.24),
            ("EMP002", "佐藤花子", "加藤木材", 358000, 318000, 40000, 11.17),
            ("EMP003", "鈴木一郎", "他社", 300000, 285000, 15000, 5.00),
        ]

        for emp_id, name, company, billing, cost, profit, margin in companies_data:
            cursor.execute("""
                INSERT INTO employees
                (employee_id, name, dispatch_company, hourly_rate, billing_rate, status)
                VALUES (?, ?, ?, 1000, 1500, 'active')
            """, (emp_id, name, company))

            cursor.execute("""
                INSERT INTO payroll_records
                (employee_id, period, work_hours, billing_amount,
                 total_company_cost, gross_profit, profit_margin)
                VALUES (?, '2025年1月', 160, ?, ?, ?, ?)
            """, (emp_id, billing, cost, profit, margin))

        db_session.commit()

        return ReportService(db_session)

    def test_all_companies_structure(self, report_service_with_companies):
        """Should return correct structure"""
        result = report_service_with_companies.get_all_companies_report_data("2025年1月")

        assert "period" in result
        assert "companies" in result
        assert "totals" in result

    def test_all_companies_aggregation(self, report_service_with_companies):
        """Should aggregate by company"""
        result = report_service_with_companies.get_all_companies_report_data("2025年1月")

        companies = result["companies"]

        # Should have 2 companies
        assert len(companies) == 2

        # Find 加藤木材 (has 2 employees)
        kato = next((c for c in companies if c["company"] == "加藤木材"), None)
        assert kato is not None
        assert kato["employee_count"] == 2
        # 340000 + 358000 = 698000
        assert kato["revenue"] == 698000

    def test_all_companies_ordered_by_profit(self, report_service_with_companies):
        """Should order by profit DESC"""
        result = report_service_with_companies.get_all_companies_report_data("2025年1月")

        companies = result["companies"]

        # 加藤木材 has higher profit (85000) than 他社 (15000)
        assert companies[0]["company"] == "加藤木材"


class TestCostBreakdownReportData:
    """Tests for ReportService.get_cost_breakdown_report_data()"""

    @pytest.fixture
    def report_service_with_insurance(self, db_session):
        """Initialize service with insurance data"""
        from reports import ReportService

        cursor = db_session.cursor()

        cursor.execute("""
            INSERT INTO employees
            (employee_id, name, dispatch_company, hourly_rate, billing_rate, status)
            VALUES ('EMP001', '田中太郎', '加藤木材', 1000, 1500, 'active')
        """)

        cursor.execute("""
            INSERT INTO payroll_records
            (employee_id, period, work_hours, gross_salary, billing_amount,
             social_insurance, welfare_pension, employment_insurance,
             company_social_insurance, company_employment_insurance, company_workers_comp,
             transport_allowance, total_company_cost, gross_profit, profit_margin)
            VALUES ('EMP001', '2025年1月', 160, 300000, 340000,
                    15000, 27000, 3000, 15000, 2700, 900, 10000,
                    295000, 45000, 13.24)
        """)

        db_session.commit()

        return ReportService(db_session)

    def test_cost_breakdown_structure(self, report_service_with_insurance):
        """Should return correct structure"""
        result = report_service_with_insurance.get_cost_breakdown_report_data("2025年1月")

        assert "period" in result
        assert "employees" in result
        assert "totals" in result

    def test_cost_breakdown_includes_insurance(self, report_service_with_insurance):
        """Should include all insurance fields"""
        result = report_service_with_insurance.get_cost_breakdown_report_data("2025年1月")

        employees = result["employees"]
        assert len(employees) >= 1

        emp = employees[0]
        insurance_fields = [
            "social_insurance", "welfare_pension", "employment_insurance",
            "company_social_insurance", "company_employment_insurance", "company_workers_comp"
        ]

        for field in insurance_fields:
            assert field in emp


class TestSummaryReportData:
    """Tests for ReportService.get_summary_report_data()"""

    @pytest.fixture
    def report_service_with_data(self, db_session):
        """Initialize service with test data"""
        from reports import ReportService

        cursor = db_session.cursor()

        # Create multiple employees
        for i in range(5):
            emp_id = f"EMP{i:03d}"
            cursor.execute("""
                INSERT INTO employees
                (employee_id, name, dispatch_company, hourly_rate, billing_rate, status)
                VALUES (?, ?, '加藤木材', 1000, 1500, 'active')
            """, (emp_id, f"従業員{i}"))

            cursor.execute("""
                INSERT INTO payroll_records
                (employee_id, period, work_hours, billing_amount,
                 total_company_cost, gross_profit, profit_margin)
                VALUES (?, '2025年1月', 160, 340000, 295000, 45000, 13.24)
            """, (emp_id,))

        db_session.commit()

        return ReportService(db_session)

    def test_summary_report_structure(self, report_service_with_data):
        """Should return correct structure"""
        result = report_service_with_data.get_summary_report_data("2025年1月")

        assert "period" in result
        assert "summary" in result
        assert "by_company" in result
        assert "top_performers" in result
        assert "bottom_performers" in result

    def test_summary_report_limits_by_company(self, report_service_with_data):
        """Should limit by_company to first 10"""
        result = report_service_with_data.get_summary_report_data("2025年1月")

        # by_company should be limited to 10
        assert len(result["by_company"]) <= 10


class TestExcelGeneration:
    """Tests for Excel report generation"""

    @pytest.fixture
    def report_service_with_data(self, db_session):
        """Initialize service with test data"""
        from reports import ReportService

        cursor = db_session.cursor()

        cursor.execute("""
            INSERT INTO employees
            (employee_id, name, name_kana, dispatch_company, hourly_rate, billing_rate, status)
            VALUES ('EMP001', '田中太郎', 'タナカタロウ', '加藤木材', 1000, 1500, 'active')
        """)

        cursor.execute("""
            INSERT INTO payroll_records
            (employee_id, period, work_hours, overtime_hours, gross_salary,
             billing_amount, total_company_cost, gross_profit, profit_margin)
            VALUES ('EMP001', '2025年1月', 160, 20, 300000, 340000, 295000, 45000, 13.24)
        """)

        db_session.commit()

        return ReportService(db_session)

    def test_generate_monthly_excel(self, report_service_with_data):
        """Should generate valid Excel for monthly report"""
        data = report_service_with_data.get_monthly_report_data("2025年1月")
        excel_bytes = report_service_with_data.generate_excel_report("monthly", data)

        assert excel_bytes is not None
        assert len(excel_bytes) > 0

        # Verify it's a valid XLSX by checking magic bytes
        assert excel_bytes[:4] == b'PK\x03\x04'  # ZIP file signature (XLSX is ZIP)

    def test_generate_all_employees_excel(self, report_service_with_data):
        """Should generate valid Excel for all-employees report"""
        data = report_service_with_data.get_all_employees_report_data("2025年1月")
        excel_bytes = report_service_with_data.generate_excel_report("all-employees", data)

        assert excel_bytes is not None
        assert excel_bytes[:4] == b'PK\x03\x04'

    def test_generate_all_companies_excel(self, report_service_with_data):
        """Should generate valid Excel for all-companies report"""
        data = report_service_with_data.get_all_companies_report_data("2025年1月")
        excel_bytes = report_service_with_data.generate_excel_report("all-companies", data)

        assert excel_bytes is not None
        assert excel_bytes[:4] == b'PK\x03\x04'

    def test_generate_cost_breakdown_excel(self, report_service_with_data):
        """Should generate valid Excel for cost-breakdown report"""
        data = report_service_with_data.get_cost_breakdown_report_data("2025年1月")
        excel_bytes = report_service_with_data.generate_excel_report("cost-breakdown", data)

        assert excel_bytes is not None
        assert excel_bytes[:4] == b'PK\x03\x04'

    def test_generate_summary_excel(self, report_service_with_data):
        """Should generate valid Excel for summary report"""
        data = report_service_with_data.get_summary_report_data("2025年1月")
        excel_bytes = report_service_with_data.generate_excel_report("summary", data)

        assert excel_bytes is not None
        assert excel_bytes[:4] == b'PK\x03\x04'

    def test_excel_contains_japanese_text(self, report_service_with_data):
        """Excel should preserve Japanese characters"""
        from openpyxl import load_workbook

        data = report_service_with_data.get_monthly_report_data("2025年1月")
        excel_bytes = report_service_with_data.generate_excel_report("monthly", data)

        # Load workbook from bytes
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        # Check for Japanese text in cells
        found_japanese = False
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and any(ord(c) > 127 for c in cell):
                    found_japanese = True
                    break

        assert found_japanese, "Excel should contain Japanese characters"


class TestSafeSheetTitle:
    """Tests for sheet title sanitization"""

    def test_sanitize_japanese_period(self, db_session):
        """Should sanitize Japanese year/month format"""
        from reports import ReportService

        service = ReportService(db_session)

        result = service._safe_sheet_title("2025年1月")

        # Should remove 年 and 月
        assert "年" not in result
        assert "月" not in result
        # Should have numeric parts
        assert "2025" in result or "1" in result

    def test_truncate_long_title(self, db_session):
        """Should truncate to 31 characters"""
        from reports import ReportService

        service = ReportService(db_session)

        long_title = "A" * 50
        result = service._safe_sheet_title(long_title)

        assert len(result) <= 31

    def test_remove_invalid_characters(self, db_session):
        """Should remove Excel invalid characters"""
        from reports import ReportService

        service = ReportService(db_session)

        invalid_title = "Test/Report:Name"
        result = service._safe_sheet_title(invalid_title)

        assert "/" not in result
        assert ":" not in result


class TestEmptyDataHandling:
    """Tests for handling empty data scenarios"""

    def test_empty_period_monthly_report(self, db_session):
        """Should handle empty period without errors"""
        from reports import ReportService

        service = ReportService(db_session)
        result = service.get_monthly_report_data("2030年1月")

        assert result["summary"]["employee_count"] == 0
        assert result["by_company"] == []
        assert result["top_performers"] == []
        assert result["bottom_performers"] == []

    def test_empty_period_all_employees(self, db_session):
        """Should return empty employees list"""
        from reports import ReportService

        service = ReportService(db_session)
        result = service.get_all_employees_report_data("2030年1月")

        assert result["employees"] == []
        assert result["totals"]["employee_count"] == 0

    def test_empty_period_excel_generation(self, db_session):
        """Should generate Excel even with empty data"""
        from reports import ReportService

        service = ReportService(db_session)
        data = service.get_monthly_report_data("2030年1月")
        excel_bytes = service.generate_excel_report("monthly", data)

        # Should still generate valid Excel
        assert excel_bytes is not None
        assert excel_bytes[:4] == b'PK\x03\x04'
