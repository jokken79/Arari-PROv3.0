"""
Specialized parser for .xlsm salary statement files (給与明細)

VERSIÓN 4.0 - Parser Híbrido (Fijo + Dinámico por Empleado)
============================================================
- Filas 1-19: POSICIÓN FIJA (datos básicos siempre en misma fila)
- Filas 20-29: ESCANEO DINÁMICO POR EMPLEADO (allowances varían)
- Filas 30+: POSICIÓN FIJA (totales y deducciones)

Problema resuelto:
- Empleado A tiene ガソリン代 en fila 20
- Empleado B tiene 60H過残業 en fila 20 y ガソリン代 en fila 21
- Empleado C tiene 皆勤手当 en fila 20, 業務手当 en fila 21, ガソリン代 en fila 22

Solución:
- Para cada empleado, escanear filas 20-29 buscando labels
- Detectar qué tiene cada uno individualmente
- Leer valores según lo que se encuentre

Features:
- Detecta automáticamente cualquier 手当 (allowance) POR EMPLEADO
- Distingue entre billable y non-billable allowances
- Valida consistencia: suma componentes vs 総支給額
- Compatible con sistema de templates (para filas fijas)

Handles complex multi-sheet, multi-employee layout where:
- Each file has multiple sheets (1 summary + company sheets)
- Each company sheet has multiple employees side-by-side
- Each employee occupies ~14 columns

Multipliers for billing calculation:
- 基本時間: 単価 × hours
- 残業 (≤60h): 単価 × 1.25
- 残業 (>60h): 単価 × 1.5
- 深夜 (factory): 単価 × 0.25 (extra on top of regular)
- 休日: 単価 × 1.35
"""

import re
import unicodedata
from io import BytesIO
from typing import Any, Dict, List, Optional

import openpyxl

from models import PayrollRecordCreate
from template_manager import TemplateGenerator, TemplateManager


class SalaryStatementParser:
    """Parser inteligente para archivos .xlsm de 給与明細"""

    # Each employee occupies this many columns width
    EMPLOYEE_COLUMN_WIDTH = 14

    # ================================================================
    # FIELD MAPPINGS - Busca estos nombres en las celdas del Excel
    # IMPORTANTE: Los patterns deben ser ESPECÍFICOS para evitar confusiones
    # - Para HORAS: usar patterns que NO contengan 手当/代/割増
    # - Para YEN (pagos): usar patterns que SÍ contengan 手当/代/割増 o 給
    #
    # ACTUALIZADO 2025-12-10: Integrado labels comprehensivos de ChinginGenerator
    # ================================================================
    FIELD_PATTERNS = {
        # ================================================================
        # IDENTIFICACIÓN (Identification) - AMPLIADO con variantes de v8.0
        # ================================================================
        "employee_id": ["従業員番号", "社員番号", "社員No", "ID", "No."],
        "name_jp": ["氏名", "名前"],
        "name_roman": ["氏名ローマ字", "ローマ字"],
        "factory": ["派遣先", "勤務先", "工場"],
        "address": ["住所"],  # NUEVO de v8.0
        "postal_code": ["郵便番号"],  # NUEVO de v8.0
        "phone": ["電話番号", "電話"],  # NUEVO de v8.0
        "emergency_contact": ["緊急連絡先"],  # NUEVO de v8.0

        # ================================================================
        # PERÍODO (Period)
        # ================================================================
        "period": ["支給分", "給与月", "期間", "対象月"],  # NUEVO de v8.0
        "period_start": ["賃金計算期間S"],  # NUEVO de v8.0
        "period_end": ["賃金計算期間F"],  # NUEVO de v8.0

        # ================================================================
        # DÍAS Y HORAS (Time data) - SOLO patterns de HORAS, sin 手当/代
        # ================================================================
        "work_days": ["出勤日数", "労働日数", "日数", "勤務日数"],
        "paid_leave_days": ["有給日数", "有給", "有休日数"],
        "absence_days": ["欠勤日数", "欠勤"],
        "work_hours": ["労働時間", "実働時", "実働時間", "勤務時間", "所定時間"],
        "overtime_hours": ["残業時間", "所定時間外", "時間外労働", "残業時間数"],
        "overtime_minutes": ["残業時間数分"],  # NUEVO de v8.0 - minutos separados
        "night_hours": ["深夜時間", "深夜労働時間", "深夜労働"],
        "holiday_hours": ["休日時間", "休日労働", "休日出勤"],
        "overtime_over_60h_hours": ["60H過残業", "60時間超", "60H超"],  # NUEVO - horas >60h

        # ================================================================
        # INGRESOS - SALARIO BASE (Base salary)
        # ================================================================
        "base_salary": [
            "基本給", "基本賃金", "本給", "基　本　給", "給与",
            "基本給(時給)",  # NUEVO de v8.0
        ],

        # ================================================================
        # INGRESOS - HORAS EXTRA (Overtime pay)
        # ================================================================
        "overtime_pay": [
            "残業手当", "時間外手当", "残業代", "普通残業", "普通残業手当",
        ],
        "overtime_over_60h_pay": [
            "60H過手当", "60時間超手当", "60H超手当",
            "60H過", "60時間超",  # NUEVO de v8.0 - variantes sin 手当
        ],

        # ================================================================
        # INGRESOS - NOCTURNOS (Night pay)
        # ================================================================
        "night_pay": [
            "深夜手当", "深夜割増", "深夜代",
        ],
        "night_overtime_pay": [
            "深夜残業", "深夜残業手当",  # SEPARADO de night_pay - es diferente
        ],

        # ================================================================
        # INGRESOS - FESTIVOS Y VACACIONES (Holiday & Leave pay)
        # ================================================================
        "holiday_pay": [
            "休日手当", "休日割増", "休出手当", "休日勤務", "休日勤務手当",
        ],
        "paid_leave_amount": [
            "有給金額", "有休金額", "有給手当", "有給支給", "有給休暇", "有休手当",
        ],

        # ================================================================
        # INGRESOS - TRANSPORTE (Transport allowance)
        # ================================================================
        "transport_allowance": [
            "交通費", "通勤手当", "通勤費",
            "ガソリン", "ガソリン代",  # Gasolina
        ],
        "transport_non_taxable": [
            "通勤手当(非)", "通勤手当（非）",  # Non-taxable transport
        ],

        # ================================================================
        # INGRESOS - ALLOWANCES/BONOS (Various allowances) - AMPLIADO de v8.0
        # ================================================================
        "attendance_bonus": ["皆勤手当", "皆勤賞"],  # Perfect attendance
        "job_allowance": ["職務手当"],  # Job allowance
        "position_allowance": ["役職手当"],  # Position allowance
        "qualification_allowance": ["資格手当"],  # Qualification allowance
        "special_allowance": ["特別手当"],  # Special allowance
        "adjustment_allowance": ["調整手当"],  # Adjustment allowance
        "busy_season_allowance": ["繁忙期手当"],  # Busy season allowance
        "setup_allowance": ["段取手当"],  # Setup allowance
        "shift_allowance": ["交代手当"],  # Shift allowance
        "hourly_rate_diff": ["時給差額"],  # NUEVO de v8.0 - Hourly rate difference
        "leave_compensation": ["休業補償"],  # NUEVO de v8.0 - Leave compensation
        "dept_bonus": ["部会賞金"],  # Department bonus

        # ================================================================
        # DEDUCCIONES - SEGUROS (Insurance deductions) - AMPLIADO de v8.0
        # ================================================================
        "social_insurance": ["社会保険", "社保", "健康保険", "健康保険料"],
        "welfare_pension": ["厚生年金", "厚生年金保険"],
        "employment_insurance": ["雇用保険", "雇用保険料"],
        "care_insurance": ["介護保険", "介護"],  # ❗NUEVO CRÍTICO de v8.0
        "social_insurance_total": ["社保計", "社会保険計", "社会保険料計"],

        # ================================================================
        # DEDUCCIONES - IMPUESTOS (Tax deductions) - AMPLIADO de v8.0
        # ================================================================
        "income_tax": ["所得税", "源泉税", "源泉所得税"],
        "resident_tax": ["住民税", "市民税"],
        "year_end_adjustment": ["年調過不足", "年末調整"],
        "fixed_tax_reduction": ["定額減税"],  # NUEVO de v8.0 - Tax reduction
        "income_tax_after_reduction": ["減税徴収"],  # NUEVO de v8.0

        # ================================================================
        # DEDUCCIONES - VIVIENDA Y SERVICIOS (Housing & Utilities)
        # ================================================================
        "rent_deduction": ["家賃", "寮費"],
        "utilities": ["水道光熱", "光熱費", "電気代", "ガス代", "wifi代"],  # AMPLIADO

        # ================================================================
        # DEDUCCIONES - OTROS (Other deductions) - AMPLIADO de v8.0
        # ================================================================
        "advance_payment": ["前貸", "前借", "プール金"],  # AMPLIADO
        "meal_cost": ["弁当", "弁当代", "食事代", "食事"],
        "uniform_cost": ["制服代"],  # NUEVO de v8.0
        "gloves_cost": ["軍手・ビニ手代"],  # NUEVO de v8.0
        "visa_renewal": ["ビザ更新料"],  # NUEVO de v8.0
        "badge_cost": ["名札代"],  # NUEVO de v8.0
        "union_fee": ["組合費"],  # NUEVO de v8.0
        "life_insurance": ["生命保険"],  # NUEVO de v8.0
        "deduction_generic": ["控除1", "控除2", "控除3"],  # NUEVO de v8.0

        # ================================================================
        # TOTALES (Totals) - AMPLIADO
        # ================================================================
        "gross_salary": [
            "総支給額", "支給合計", "総支給", "給与総額", "合　　計", "合計",
        ],
        "net_salary": [
            "差引支給額", "手取り", "振込額", "差引額", "差引支給",
        ],
        "deduction_total": ["控除合計", "控除計", "控除　合計"],
    }

    # Patterns to detect ANY 手当 (allowance)
    ALLOWANCE_PATTERNS = [
        r".*手当.*",  # Cualquier cosa con 手当
        r".*割増.*",  # Cualquier割増 (premium)
        r".*加算.*",  # Cualquier加算 (addition)
    ]

    # Known allowances to EXCLUDE from "other_allowances" (ya tienen campo propio)
    KNOWN_ALLOWANCES = [
        "残業手当",
        "時間外手当",
        "深夜手当",
        "休日手当",
        "休出手当",
        "60H過手当",
        "60時間超手当",
        "通勤手当",
        "有給手当",
        "残業代",
        "深夜代",
        "深夜割増",
        "休日割増",
        "通勤費",  # Added - now handled as non_billable in dynamic zone
    ]

    # ================================================================
    # NON-BILLABLE ALLOWANCES (会社負担のみ、派遣先に請求しない)
    # These are paid to employee but NOT billed to client
    # ================================================================
    NON_BILLABLE_ALLOWANCES = [
        "通勤手当",  # Transport allowance
        "通勤手当（非）",  # Transport allowance (non-taxable) - 全角 brackets
        "通勤手当(非)",  # Transport allowance (non-taxable) - 半角 brackets
        "通勤費",  # Transport cost
        "業務手当",  # Work allowance
    ]

    # Fallback row positions (used if intelligent detection fails)
    # ACTUALIZADO 2025-12-09 - Basado en análisis REAL del Excel completo
    FALLBACK_ROW_POSITIONS = {
        "period": 10,  # Período (datetime en row 10)
        "employee_id": 6,  # ID empleado (200901)
        "name": 7,  # Nombre
        "work_days": 11,  # Días trabajados (16) - en columna Days
        "paid_leave_days": 12,  # Días de vacaciones pagadas
        "work_hours": 13,  # Horas de trabajo (128)
        "overtime_hours": 14,  # Horas extra (13)
        "night_hours": 15,  # Horas nocturnas (70)
        "base_salary": 16,  # Salario base (¥172,800)
        "overtime_pay": 17,  # Pago horas extra (¥23,210)
        "night_pay": 18,  # Pago nocturno (¥23,829)
        "gross_salary": 30,  # Salario bruto (¥219,839)
        "social_insurance": 31,  # 健康保険 (¥15,030)
        "welfare_pension": 32,  # 厚生年金 - AGREGADO para cálculo correcto
        "employment_insurance": 33,  # 雇用保険 (¥1,319)
        "income_tax": 34,  # 所得税
        "resident_tax": 35,  # 住民税
        "net_salary": 47,  # 差引支給額 (¥182,677)
        # Campos en zona dinámica (20-29):
        # holiday_hours, overtime_over_60h_pay, transport_allowance, paid_leave_amount
    }

    # Column offsets within an employee block
    # CRITICAL: Este Excel tiene datos en MÚLTIPLES columnas
    # Para employee 1 (base_col=1): C1=marker, C2=category, C3=label, C4=value
    COLUMN_OFFSETS = {
        "period": 8,  # Period está en col 9 (base_col=1, offset=8)
        "employee_id": 9,  # Employee ID está en col 10 (base_col=1, offset=9)
        "name": 9,  # Name también en col 10
        "label": 2,  # Labels en col 3 (base_col=1, offset=2) ← CORREGIDO
        "value": 3,  # VALUES (salarios, horas) en col 4 (base_col=1, offset=3)
        "days": 5,  # DAYS (work_days) en col 6 (base_col=1, offset=5)
        "minutes": 9,  # MINUTES para horas (col 10) - usado para convertir HH:MM a decimal
    }

    # ================================================================
    # ZONA DINÁMICA - Filas donde los allowances varían por empleado
    # ================================================================
    DYNAMIC_ZONE_START = 20  # Primera fila de la zona dinámica
    DYNAMIC_ZONE_END = 29  # Última fila de la zona dinámica

    # Labels que pueden aparecer en la zona dinámica (20-29)
    # ACTUALIZADO 2026-01-30: Integrado ~200 labels de Chingin v8.0
    DYNAMIC_ZONE_LABELS = {
        # ================================================================
        # OVERTIME OVER 60H (Horas extra > 60h)
        # ================================================================
        "60H過残業": "overtime_over_60h_pay",
        "60H過": "overtime_over_60h_pay",
        "60時間超": "overtime_over_60h_pay",
        "60h超残業": "overtime_over_60h_pay",
        "60H過手当": "overtime_over_60h_pay",
        "60時間超手当": "overtime_over_60h_pay",

        # ================================================================
        # PAID LEAVE (有給)
        # ================================================================
        "有給休暇": "paid_leave_amount",
        "有給": "paid_leave_amount",
        "有休": "paid_leave_amount",
        "有休手当": "paid_leave_amount",
        "有給手当": "paid_leave_amount",

        # ================================================================
        # NON-BILLABLE (company cost only - 会社負担のみ)
        # ================================================================
        "通勤手当(非)": "non_billable",
        "通勤手当（非）": "non_billable",
        "業務手当": "non_billable",
        "通勤費": "non_billable",
        "通勤手当": "non_billable",
        "ガソリン": "non_billable",
        "ガソリン代": "non_billable",
        "交通費": "non_billable",

        # ================================================================
        # BILLABLE ALLOWANCES (手当 - facturables)
        # ================================================================
        "休業補償": "other_allowance",
        "皆勤手当": "other_allowance",
        "皆勤賞": "other_allowance",
        "変則手当": "other_allowance",
        "土日手当": "other_allowance",
        "繁忙期手当": "other_allowance",
        "職務手当": "other_allowance",
        "役職手当": "other_allowance",
        "資格手当": "other_allowance",
        "特別手当": "other_allowance",
        "調整手当": "other_allowance",
        "段取手当": "other_allowance",
        "交代手当": "other_allowance",
        "部会賞金": "other_allowance",
        "半日有給": "other_allowance",
        "深夜残業": "other_allowance",
        "前月給与": "other_allowance",
        # NUEVOS de v8.0
        "時給差額": "other_allowance",  # Hourly rate difference
        "住宅手当": "other_allowance",  # Housing allowance
        "家族手当": "other_allowance",  # Family allowance
        "精勤手当": "other_allowance",  # Diligence allowance
        "能率手当": "other_allowance",  # Efficiency allowance
        "技能手当": "other_allowance",  # Skill allowance
        "危険手当": "other_allowance",  # Hazard allowance
        "寒冷地手当": "other_allowance",  # Cold region allowance
        "単身赴任手当": "other_allowance",  # Single assignment allowance

        # ================================================================
        # DEDUCTIONS - HOUSING & UTILITIES (Deducciones de vivienda)
        # ================================================================
        "家賃": "rent_deduction",
        "寮費": "rent_deduction",
        "住居費": "rent_deduction",  # NUEVO
        "水道光熱": "utilities",
        "光熱費": "utilities",
        "電気代": "utilities",
        "ガス代": "utilities",  # NUEVO
        "水道代": "utilities",  # NUEVO
        "wifi代": "utilities",  # NUEVO de v8.0

        # ================================================================
        # DEDUCTIONS - ADVANCES & LOANS (Adelantos y préstamos)
        # ================================================================
        "前貸": "advance_payment",
        "前借": "advance_payment",
        "プール金": "advance_payment",  # NUEVO de v8.0
        "立替金": "advance_payment",  # NUEVO - Reimbursement

        # ================================================================
        # DEDUCTIONS - MEALS (Comidas)
        # ================================================================
        "弁当": "meal_deduction",
        "弁当代": "meal_deduction",
        "食事代": "meal_deduction",
        "食事": "meal_deduction",
        "食費": "meal_deduction",  # NUEVO

        # ================================================================
        # DEDUCTIONS - INSURANCE & TAX ADJUSTMENTS
        # ================================================================
        "年調過不足": "year_end_adjustment",
        "年末調整": "year_end_adjustment",
        "定額減税": "tax_reduction",  # NUEVO de v8.0
        "介護保険": "care_insurance",  # NUEVO de v8.0 - Care insurance
        "生命保険": "life_insurance",  # NUEVO de v8.0

        # ================================================================
        # DEDUCTIONS - OTHER (Otras deducciones)
        # ================================================================
        "制服代": "uniform_cost",  # NUEVO de v8.0
        "作業服代": "uniform_cost",  # NUEVO - Work clothes
        "軍手・ビニ手代": "gloves_cost",  # NUEVO de v8.0
        "ビザ更新料": "visa_renewal",  # NUEVO de v8.0
        "在留資格": "visa_renewal",  # NUEVO - Residence status
        "名札代": "badge_cost",  # NUEVO de v8.0
        "組合費": "union_fee",  # NUEVO de v8.0
        "親睦会費": "union_fee",  # NUEVO - Social club fee
        "積立金": "savings_deduction",  # NUEVO - Savings
        "財形貯蓄": "savings_deduction",  # NUEVO - Asset formation savings
        "駐車場代": "parking_deduction",  # NUEVO - Parking
        "社宅費": "company_housing",  # NUEVO - Company housing
        "備品代": "equipment_cost",  # NUEVO - Equipment
        "保証金": "deposit",  # NUEVO - Deposit
        "その他控除": "other_deduction",  # NUEVO - Other deductions
    }

    def __init__(
        self,
        use_intelligent_mode: bool = True,
        template_manager: Optional[TemplateManager] = None,
        employee_name_map: Dict[str, str] = None,
    ):
        """
        Initialize parser with template support.

        Args:
            use_intelligent_mode: If True, scan Excel for field names dynamically
                                 DEFAULT TRUE - now uses template system
            template_manager: Optional TemplateManager instance for template storage
                            If None, creates one automatically
            employee_name_map: Optional map of employee names to IDs for lookup in vertical sheets.
        """
        self.use_intelligent_mode = use_intelligent_mode
        self.template_manager = template_manager or TemplateManager()
        self.employee_name_map = employee_name_map or {}  # map[name] -> employee_id
        self.template_generator = TemplateGenerator()

        # Current parsing state
        self.detected_fields: Dict[str, int] = {}  # field_name -> row_number
        self.detected_allowances: Dict[str, int] = {}  # allowance_name -> row_number
        self.detected_non_billable: Dict[str, int] = {}  # non-billable allowances
        self.current_column_offsets: Dict[str, int] = {}  # Current column offsets
        self.validation_warnings: List[str] = []

        # Template tracking
        self.templates_used: List[str] = []  # Factory names where template was used
        self.templates_generated: List[str] = (
            []
        )  # Factory names where template was generated
        self.using_template: bool = False  # Whether current sheet uses a template

    def parse(self, content: bytes) -> List[PayrollRecordCreate]:
        """
        Parse .xlsm file and extract all employee payroll records

        Args:
            content: Binary content of the Excel file

        Returns:
            List of PayrollRecordCreate objects
        """
        try:
            # Keep BytesIO open throughout parsing
            file_buffer = BytesIO(content)
            wb = openpyxl.load_workbook(file_buffer, data_only=True, read_only=False)
        except Exception as e:
            print(f"[ERROR] Error loading Excel file: {e}")
            return []

        records = []

        print(f"[DEBUG] Starting SalaryStatementParser. Sheets: {wb.sheetnames}")

        # Process all sheets except the summary sheet (集計) and Contract (請負)
        for sheet_name in wb.sheetnames:
            # Skip only summary and index sheets. '請負' (Ukeoi) is now ALLOWED.
            if sheet_name in [
                "集計",
                "Summary",
                "目次",
                "Index",
                "請負",
                "DBUkeoiX",
                "請負社員",
            ]:
                print(f"[DEBUG] Skipping sheet: {sheet_name}")
                continue

            try:
                print(f"[DEBUG] Processing sheet: {sheet_name}")
                ws = wb[sheet_name]
                sheet_records = self._parse_sheet(ws, sheet_name)
                print(
                    f"[DEBUG] Sheet '{sheet_name}' yielded {len(sheet_records)} records"
                )
                records.extend(sheet_records)
            except Exception as e:
                print(f"[WARNING] Error parsing sheet '{sheet_name}': {e}")
                import traceback

                traceback.print_exc()
                continue

        print(f"[OK] Parsed {len(records)} employee records from Excel")

        # Show template usage summary
        if self.templates_used or self.templates_generated:
            print("\n[TEMPLATES] Summary:")
            if self.templates_used:
                print(f"   Used existing templates: {', '.join(self.templates_used)}")
            if self.templates_generated:
                print(
                    f"   Generated new templates: {', '.join(self.templates_generated)}"
                )

        # Show validation warnings
        if self.validation_warnings:
            print(f"\n[WARNING] VALIDATION WARNINGS ({len(self.validation_warnings)}):")
            for warning in self.validation_warnings[:10]:  # Show first 10
                print(f"   {warning}")

        return records

    def get_parsing_stats(self) -> Dict[str, Any]:
        """
        Return statistics about the parsing process.
        """
        return {
            "templates_used": self.templates_used,
            "templates_generated": self.templates_generated,
            "validation_warnings": self.validation_warnings,
        }

    def _detect_layout_type(self, ws) -> str:
        """
        Detect whether the sheet uses Standard (horizontal), Vertical, or Kintaihyo layout.

        Standard:
        - Employee IDs in a single row (e.g., row 6)
        - Repeats every 14 columns
        - Headers are within the 14-col block

        Vertical (Format B):
        - Headers in Column A (Row 1-20)
        - Month columns (e.g. 4月, 5月) in a header row
        - Values are in the intersection of Header Row (e.g. Basic Salary) and Month Column

        Kintaihyo (8-row blocks):
        - Daily attendance sheets (勤怠表)
        - Each employee occupies 8 rows
        - Row+0: English name + daily data [8] [8.5] [0] [2]
        - Row+1: Employee ID (5-7 digits, typically > 10000)
        - Row+2: Japanese name
        - Row+3: Base rate (> 500)
        - Row+4-6: Payments and taxes
        - Row+7: Detailed daily attendance
        """
        # Check for Kintaihyo format first (highest priority for this specific format)
        if self._detect_kintaihyo_format(ws):
            return "kintaihyo"

        # Check for vertical layout indicators (Labels in Column A)
        vertical_indicators = [
            "基本給",
            "残業手当",
            "総支給額",
            "差引支給額",
            "労働日数",
            "労働時間",
            "時間外労働",
        ]
        col_a_values = []
        for r in range(1, 60):
            val = ws.cell(row=r, column=1).value
            if val:
                col_a_values.append(str(val).strip().replace(" ", "").replace("　", ""))

        matches = sum(
            1 for ind in vertical_indicators if any(ind in val for val in col_a_values)
        )

        # Check for standard layout indicators (Employee IDs in row 6)
        # We'll assume if it looks like Vertical, it is Vertical.
        if matches >= 3:
            return "vertical"

        return "standard"

    def _detect_kintaihyo_format(self, ws) -> bool:
        """
        Detect if the sheet uses Kintaihyo 8-row format.

        Pattern detection:
        - Row+0: Text (English name or header)
        - Row+1: Number > 10000 (Employee ID, typically 5-7 digits)
        - Row+2: Text (Japanese name with kanji/hiragana/katakana)
        - Row+3: Number > 500 (Base rate/tariff)

        Returns:
            True if Kintaihyo format is detected, False otherwise
        """
        # Scan first 50 rows looking for the characteristic pattern
        blocks_found = 0

        for start_row in range(1, min(50, ws.max_row - 3)):
            # Check multiple columns (employees might be side by side)
            for col in range(1, min(30, ws.max_column + 1)):
                if self._is_kintaihyo_block_start(ws, start_row, col):
                    blocks_found += 1
                    if blocks_found >= 2:
                        # Found at least 2 blocks - confident it's Kintaihyo format
                        return True

        # If only 1 block found, still accept if it looks valid
        return blocks_found >= 1

    def _is_kintaihyo_block_start(self, ws, row: int, col: int) -> bool:
        """
        Check if the given position is the start of a Kintaihyo 8-row block.

        Pattern:
        - Row+0: Text (name in English or label)
        - Row+1: Number > 10000 (Employee ID)
        - Row+2: Text (Japanese name)
        - Row+3: Number > 500 (Base rate)

        Args:
            ws: Worksheet
            row: Starting row to check
            col: Column to check

        Returns:
            True if this looks like a Kintaihyo block start
        """
        try:
            # Row+0: Should be text (English name or header)
            val_row0 = ws.cell(row=row, column=col).value
            if val_row0 is None:
                return False

            # Must be text, not a number
            if isinstance(val_row0, (int, float)):
                return False

            val_row0_str = str(val_row0).strip()
            if not val_row0_str or val_row0_str.isdigit():
                return False

            # Row+1: Should be Employee ID (number > 10000, typically 5-7 digits)
            val_row1 = ws.cell(row=row + 1, column=col).value
            if val_row1 is None:
                return False

            try:
                emp_id = int(float(str(val_row1).strip().replace(",", "")))
                if emp_id < 10000 or emp_id > 9999999:
                    return False
            except (ValueError, TypeError):
                return False

            # Row+2: Should be Japanese name (text with Japanese characters)
            val_row2 = ws.cell(row=row + 2, column=col).value
            if val_row2 is None:
                return False

            val_row2_str = str(val_row2).strip()
            if not val_row2_str:
                return False

            # Check for Japanese characters (hiragana, katakana, kanji)
            has_japanese = any(
                "\u3040" <= c <= "\u309f"  # Hiragana
                or "\u30a0" <= c <= "\u30ff"  # Katakana
                or "\u4e00" <= c <= "\u9fff"  # CJK Unified Ideographs (Kanji)
                for c in val_row2_str
            )
            if not has_japanese:
                return False

            # Row+3: Should be base rate (number > 500)
            val_row3 = ws.cell(row=row + 3, column=col).value
            if val_row3 is None:
                return False

            try:
                base_rate = float(str(val_row3).strip().replace(",", ""))
                if base_rate < 500:
                    return False
            except (ValueError, TypeError):
                return False

            # All checks passed - this is a valid Kintaihyo block start
            return True

        except Exception:
            return False

    def _parse_sheet(self, ws, sheet_name: str) -> List[PayrollRecordCreate]:
        """
        Parse a single company sheet, dispatching to correct parser.
        """
        # 1. Detect layout type
        layout_type = self._detect_layout_type(ws)
        print(f"[DEBUG] Sheet '{sheet_name}' detected layout: {layout_type}")

        if layout_type == "kintaihyo":
            return self._parse_kintaihyo_sheet(ws, sheet_name)
        elif layout_type == "vertical":
            return self._parse_vertical_sheet(ws, sheet_name)
        else:
            return self._parse_standard_sheet(ws, sheet_name)

    def _parse_standard_sheet(self, ws, sheet_name: str) -> List[PayrollRecordCreate]:
        """
        Legacy/Standard parser logic (Moved from _parse_sheet)
        """
        records = []
        self.using_template = False

        # ... (rest of the original _parse_sheet logic mostly unchanged)
        # ================================================================
        # STEP 1: Try to load existing template
        # ================================================================
        template = self.template_manager.find_matching_template(sheet_name)

        if template and template.get("detection_confidence", 0) >= 0.5:
            # Use template
            self._apply_template(template)
            self.using_template = True
            self.templates_used.append(sheet_name)
            print(
                f"  [Sheet '{sheet_name}'] Using saved template "
                f"(confidence={template.get('detection_confidence', 0):.2f})"
            )
        else:
            # ================================================================
            # STEP 2: No template - use intelligent detection
            # ================================================================
            if self.use_intelligent_mode:
                self._detect_field_positions(ws)

                # Check if detection was successful
                required_fields = ["gross_salary", "base_salary", "work_hours"]
                found_required = sum(
                    1 for f in required_fields if f in self.detected_fields
                )
                detection_confidence = found_required / len(required_fields)

                if detection_confidence >= 0.6:
                    self._save_detected_template(ws, sheet_name, detection_confidence)
                    self.templates_generated.append(sheet_name)
                    print(
                        f"  [Sheet '{sheet_name}'] Generated new template "
                        f"(confidence={detection_confidence:.2f})"
                    )
                else:
                    print(
                        f"  [Sheet '{sheet_name}'] Low detection confidence "
                        f"({detection_confidence:.2f}), using fallback"
                    )

            if not self.current_column_offsets:
                self.current_column_offsets = self.COLUMN_OFFSETS.copy()

        # STEP 4: Detect employee column positions
        employee_cols = self._detect_employee_columns(ws)

        if not employee_cols:
            print(f"  [WARNING] No employee IDs found in sheet '{sheet_name}'")
            return records

        mode = "template" if self.using_template else "detection"
        print(
            f"  [Sheet '{sheet_name}'] {len(employee_cols)} employees, "
            f"{len(self.detected_fields)} fields, "
            f"{len(self.detected_allowances)} allowances, "
            f"mode={mode}"
        )

        # STEP 5: Extract data for each employee
        for col_idx in employee_cols:
            record = self._extract_employee_data(ws, col_idx, sheet_name)
            if record:
                records.append(record)

        return records

    def _parse_vertical_sheet(self, ws, sheet_name: str) -> List[PayrollRecordCreate]:
        import re

        """
        Parse Vertical Layout (Format B) sheet.
        """
        print(f"[VerticalParser] Parsing sheet '{sheet_name}' using Vertical Logic")
        records = []

        # 1. Map Rows (Labels in Column A/B)
        row_map = {}
        header_vals = (
            {}
        )  # To store global sheet headers like ID, Name if they are fixed

        # Scan first 60 rows for row labels
        for r in range(1, 60):
            # Combine Col 1 and 2 for label matching (sometimes split)
            val1 = str(ws.cell(row=r, column=1).value or "").strip()
            val2 = str(ws.cell(row=r, column=2).value or "").strip()
            label = (val1 + val2).replace(" ", "").replace("　", "")

            # Map standard fields
            for field, patterns in self.FIELD_PATTERNS.items():
                if field in row_map:
                    continue
                # Match patterns
                for p in patterns:
                    p_norm = p.replace(" ", "").replace("　", "")
                    if p_norm in label or label == p_norm:
                        row_map[field] = r
                        break

        # 2. Find Month Columns
        # Look for cells like "4月", "5月"
        month_cols = {}
        month_row = None

        # Scan first 10 rows for header row
        for r in range(1, 15):
            months_found = 0
            temp_month_cols = {}
            for c in range(1, 30):
                val = str(ws.cell(row=r, column=c).value or "").strip()
                # row_values.append(val)

                # Regex to search for month (handles full-width numbers １-１２ and optional suffixes like 分)
                match = re.search(r"([0-9０-９]+)月", val)
                if match:
                    try:
                        # Normalize full-width to half-width
                        m_str = match.group(1).translate(
                            str.maketrans("０１２３４５６７８９", "0123456789")
                        )
                        m = int(m_str)
                        if 1 <= m <= 12:
                            temp_month_cols[m] = c
                            months_found += 1
                    except:
                        pass

            # print(f"[DEBUG] Row {r} scan... Found {months_found} months")

            if months_found >= 3:  # Found the header row
                month_row = r
                month_cols = temp_month_cols
                break

        if not month_cols:
            print(f"[VerticalParser] Could not find Month header row in '{sheet_name}'")
            return []

        # 3. Find Employee Info (Name, ID)
        emp_name = "Unknown"
        emp_id = "000000"
        name_found = False
        id_found = False

        for r in range(1, 15):
            for c in range(1, 20):
                val = str(ws.cell(row=r, column=c).value or "").strip()
                if not val:
                    continue

                # Name Detection
                if val in ["氏名", "氏　名", "名前"]:
                    # Name is likely in next column or 2 columns over
                    possible_name = str(
                        ws.cell(row=r, column=c + 1).value or ""
                    ).strip()
                    if not possible_name and c + 2 <= 20:
                        possible_name = str(
                            ws.cell(row=r, column=c + 2).value or ""
                        ).strip()

                    if possible_name:
                        emp_name = possible_name
                        name_found = True

                # ID Detection
                if val in ["社員No", "社員No.", "社員番号", "NO.", "No."]:
                    possible_id = str(ws.cell(row=r, column=c + 1).value or "").strip()
                    # If C+1 is empty, try C+2
                    if not possible_id:
                        possible_id = str(
                            ws.cell(row=r, column=c + 2).value or ""
                        ).strip()

                    if possible_id:
                        emp_id = possible_id
                        id_found = True

        # 3.2 Try to find ID via Name Lookup if missing
        if (
            (emp_id == "000000" or not id_found)
            and name_found
            and self.employee_name_map
        ):
            # Try exact match
            if emp_name in self.employee_name_map:
                emp_id = self.employee_name_map[emp_name]
                print(
                    f"[VerticalParser] Resolved ID {emp_id} for name '{emp_name}' via lookup"
                )
            else:
                # Try fuzzy/normalized match (remove spaces, lower case)
                norm_name = emp_name.lower().replace(" ", "").replace("　", "")
                for name_key, id_val in self.employee_name_map.items():
                    if name_key.lower().replace(" ", "").replace("　", "") == norm_name:
                        emp_id = id_val
                        print(
                            f"[VerticalParser] Resolved ID {emp_id} for name '{emp_name}' via fuzzy lookup"
                        )
                        break

        print(
            f"[VerticalParser] Found Employee: {emp_name} ({emp_id}) - detected {len(month_cols)} months"
        )

        # 4. Extract data for each month
        for month, col in month_cols.items():
            try:
                # Basic check: if "Basic Salary" cell in this column is empty/zero, skip
                if "base_salary" in row_map:
                    check_val = self._get_numeric(ws, row_map["base_salary"], col)
                    # if check_val == 0: continue

                # Construct Record
                # Year resolution: Need to know year. Use file mod time or assume current?
                # For now assume 2024 or extract from sheet?
                # Let's try to find Year in header
                year = 2025  # Default updated to current/next year

                # Check for year in first few rows
                for r in range(1, 15):
                    for c in range(1, 40):
                        val = str(ws.cell(row=r, column=c).value or "")
                        if "年度" in val or ("年" in val and "月" not in val and c < 5):
                            # import re
                            ym = re.search(r"(\d{4})", val)
                            if ym:
                                year = int(ym.group(1))

                # Adjust year for Jan-Mar if it's Fiscal Year
                # Usually Jan-Mar is Year+1 of Fiscal Year start.
                # But let's keep it simple for now using extracted year.

                period_str = f"{year}年{month}月"

                # Extract values using row_map
                def get_val(field):
                    if field in row_map:
                        return self._get_numeric(ws, row_map[field], col)
                    return 0.0

                def get_hours(field):
                    if field in row_map:
                        # Vertical layout typically uses decimal hours in the same cell
                        return self._get_numeric(ws, row_map[field], col)
                    return 0.0

                record = PayrollRecordCreate(
                    employee_id=emp_id,
                    employee_name=emp_name,
                    period=period_str,
                    work_days=get_val("work_days"),
                    paid_leave_days=get_val("paid_leave_days"),
                    absence_days=get_val("absence_days"),
                    work_hours=get_hours("work_hours"),
                    overtime_hours=get_hours("overtime_hours"),
                    night_hours=get_hours("night_hours"),
                    holiday_hours=get_hours("holiday_hours"),
                    base_salary=get_val("base_salary"),
                    overtime_pay=get_val("overtime_pay"),
                    night_pay=get_val("night_pay"),
                    holiday_pay=get_val("holiday_pay"),
                    overtime_over_60h_pay=get_val("overtime_over_60h_pay"),
                    transport_allowance=get_val("transport_allowance"),
                    paid_leave_amount=get_val("paid_leave_amount"),
                    social_insurance=get_val("social_insurance"),
                    welfare_pension=get_val("welfare_pension"),
                    employment_insurance=get_val("employment_insurance"),
                    income_tax=get_val("income_tax"),
                    resident_tax=get_val("resident_tax"),
                    gross_salary=get_val("gross_salary"),
                    net_salary=get_val("net_salary"),
                    # Extra fields (need to calculate totals)
                    non_billable_total=0.0,  # Implement if needed
                    other_allowances_total=0.0,
                    deductions_total=get_val("deduction_total")
                    or (get_val("social_insurance") + get_val("income_tax")),  # Approx
                    source_sheet=sheet_name,
                )

                # Filter empty records (no work days, no gross salary)
                if record.gross_salary > 0 or record.work_days > 0:
                    records.append(record)

            except Exception as e:
                print(f"[VerticalParser] Error parsing month {month}: {e}")
                continue

        return records

    # ================================================================
    # KINTAIHYO FORMAT PARSER (勤怠表 - 8-row blocks)
    # ================================================================

    def _parse_kintaihyo_sheet(self, ws, sheet_name: str) -> List[PayrollRecordCreate]:
        """
        Parse a Kintaihyo format sheet (勤怠表 - daily attendance).

        This format uses 8-row blocks per employee:
        - Row+0: English name + daily attendance data [8] [8.5] [0] [2]
        - Row+1: Employee ID (5-7 digits, typically > 10000)
        - Row+2: Japanese name
        - Row+3: Base rate (> 500)
        - Row+4-6: Payments, taxes, and deductions
        - Row+7: Detailed daily attendance breakdown

        Args:
            ws: Worksheet to parse
            sheet_name: Name of the sheet (used as dispatch_company)

        Returns:
            List of PayrollRecordCreate objects
        """
        print(f"[KintaihyoParser] Parsing sheet '{sheet_name}' using Kintaihyo Logic")
        records = []

        # Detect period from sheet (usually in first few rows)
        period = self._detect_kintaihyo_period(ws)
        if not period:
            print(f"[KintaihyoParser] Could not detect period in sheet '{sheet_name}'")
            # Try to use a default period
            from datetime import datetime
            now = datetime.now()
            period = f"{now.year}年{now.month}月"

        print(f"[KintaihyoParser] Detected period: {period}")

        # Find all worker blocks
        blocks = self.detect_kintaihyo_blocks(ws)
        print(f"[KintaihyoParser] Found {len(blocks)} worker blocks")

        # Extract data from each block
        for start_row, col in blocks:
            try:
                record = self.extract_kintaihyo_worker(ws, start_row, col, period, sheet_name)
                if record:
                    records.append(record)
            except Exception as e:
                print(f"[KintaihyoParser] Error extracting worker at row {start_row}, col {col}: {e}")
                import traceback
                traceback.print_exc()
                continue

        print(f"[KintaihyoParser] Extracted {len(records)} records from sheet '{sheet_name}'")
        return records

    def _detect_kintaihyo_period(self, ws) -> Optional[str]:
        """
        Detect the period (年月) from a Kintaihyo sheet.

        Looks for patterns like:
        - 2025年1月 or 2025年01月
        - 令和7年1月
        - Dates in various formats

        Args:
            ws: Worksheet

        Returns:
            Period string in format "YYYY年M月" or None if not found
        """
        from datetime import datetime

        # Scan first 15 rows and first 40 columns for period indicators
        for row in range(1, 16):
            for col in range(1, 41):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is None:
                    continue

                # Handle datetime objects
                if isinstance(cell_value, datetime):
                    return f"{cell_value.year}年{cell_value.month}月"

                value_str = str(cell_value).strip()

                # Look for 年月 pattern
                match = re.search(r"(\d{4})年(\d{1,2})月", value_str)
                if match:
                    year = int(match.group(1))
                    month = int(match.group(2))
                    return f"{year}年{month}月"

                # Look for 令和 (Reiwa era) pattern
                reiwa_match = re.search(r"令和\s*(\d{1,2})\s*年\s*(\d{1,2})\s*月", value_str)
                if reiwa_match:
                    reiwa_year = int(reiwa_match.group(1))
                    month = int(reiwa_match.group(2))
                    # Convert Reiwa to Western: Reiwa 1 = 2019
                    year = 2018 + reiwa_year
                    return f"{year}年{month}月"

        return None

    def detect_kintaihyo_blocks(self, ws) -> List[tuple]:
        """
        Find all 8-row worker blocks in a Kintaihyo sheet.

        Scans the worksheet looking for the characteristic pattern:
        - Row+0: Text (name)
        - Row+1: Number > 10000 (Employee ID)
        - Row+2: Text with Japanese (name in kanji)
        - Row+3: Number > 500 (base rate)

        Args:
            ws: Worksheet to scan

        Returns:
            List of tuples (start_row, column) for each block found
        """
        blocks = []
        found_positions = set()  # Track found positions to avoid duplicates

        # Scan through the worksheet
        max_row = min(500, ws.max_row)  # Limit scan to 500 rows
        max_col = min(50, ws.max_column)  # Limit scan to 50 columns

        for row in range(1, max_row - 7):  # Need at least 8 rows below
            for col in range(1, max_col + 1):
                # Skip if we've already found a block near this position
                pos_key = (row // 8, col)  # Group by approximate 8-row blocks
                if pos_key in found_positions:
                    continue

                if self._is_kintaihyo_block_start(ws, row, col):
                    blocks.append((row, col))
                    found_positions.add(pos_key)

        return blocks

    def extract_kintaihyo_worker(
        self,
        ws,
        start_row: int,
        col: int,
        period: str,
        sheet_name: str
    ) -> Optional[PayrollRecordCreate]:
        """
        Extract worker data from an 8-row Kintaihyo block.

        Block structure (8 rows):
        - Row+0: English name / header + daily attendance [8] [8.5] [0] [2]
        - Row+1: Employee ID (5-7 digits)
        - Row+2: Japanese name
        - Row+3: Base rate (hourly/daily rate)
        - Row+4: Payment row 1 (may contain work_hours, overtime, etc.)
        - Row+5: Payment row 2 (deductions, taxes)
        - Row+6: Payment row 3 (totals)
        - Row+7: Detailed daily breakdown

        Args:
            ws: Worksheet
            start_row: Starting row of the block
            col: Column of the block
            period: Period string (e.g., "2025年1月")
            sheet_name: Sheet name (dispatch company)

        Returns:
            PayrollRecordCreate object or None if extraction fails
        """
        try:
            # Row+0: English name or header
            name_en = str(ws.cell(row=start_row, column=col).value or "").strip()

            # Row+1: Employee ID
            emp_id_raw = ws.cell(row=start_row + 1, column=col).value
            try:
                employee_id = str(int(float(str(emp_id_raw).strip().replace(",", ""))))
            except (ValueError, TypeError):
                print(f"[KintaihyoParser] Invalid employee ID at row {start_row + 1}: {emp_id_raw}")
                return None

            # Row+2: Japanese name
            name_jp = str(ws.cell(row=start_row + 2, column=col).value or "").strip()
            employee_name = name_jp if name_jp else name_en

            # Row+3: Base rate (tariff)
            rate_raw = ws.cell(row=start_row + 3, column=col).value
            try:
                base_rate = float(str(rate_raw).strip().replace(",", ""))
            except (ValueError, TypeError):
                base_rate = 0.0

            print(f"[KintaihyoParser] Extracting: {employee_name} (ID: {employee_id}, Rate: {base_rate})")

            # Extract daily attendance data from Row+0 (columns to the right)
            # Format: [8] [8.5] [0] [2] representing daily hours
            daily_hours = self._extract_kintaihyo_daily_hours(ws, start_row, col)

            # Calculate totals from daily hours
            work_hours = daily_hours.get("total_hours", 0.0)
            work_days = daily_hours.get("work_days", 0)
            overtime_hours = daily_hours.get("overtime_hours", 0.0)
            night_hours = daily_hours.get("night_hours", 0.0)
            holiday_hours = daily_hours.get("holiday_hours", 0.0)

            # Split overtime at 60h threshold
            overtime_over_60h = max(0, overtime_hours - 60) if overtime_hours > 60 else 0
            overtime_hours = min(overtime_hours, 60)

            # Extract payment data from rows 4-6
            payment_data = self._extract_kintaihyo_payments(ws, start_row + 4, col)

            # Calculate base salary from rate and hours if not provided
            base_salary = payment_data.get("base_salary", 0.0)
            if base_salary == 0 and base_rate > 0 and work_hours > 0:
                base_salary = base_rate * work_hours

            # Get other payment fields
            overtime_pay = payment_data.get("overtime_pay", 0.0)
            night_pay = payment_data.get("night_pay", 0.0)
            holiday_pay = payment_data.get("holiday_pay", 0.0)
            overtime_over_60h_pay = payment_data.get("overtime_over_60h_pay", 0.0)
            transport_allowance = payment_data.get("transport_allowance", 0.0)
            paid_leave_amount = payment_data.get("paid_leave_amount", 0.0)
            other_allowances = payment_data.get("other_allowances", 0.0)
            non_billable = payment_data.get("non_billable", 0.0)

            # Deductions
            social_insurance = payment_data.get("social_insurance", 0.0)
            welfare_pension = payment_data.get("welfare_pension", 0.0)
            employment_insurance = payment_data.get("employment_insurance", 0.0)
            income_tax = payment_data.get("income_tax", 0.0)
            resident_tax = payment_data.get("resident_tax", 0.0)

            # Totals
            gross_salary = payment_data.get("gross_salary", 0.0)
            net_salary = payment_data.get("net_salary", 0.0)

            # If gross_salary is 0, calculate from components
            if gross_salary == 0:
                gross_salary = (
                    base_salary
                    + overtime_pay
                    + night_pay
                    + holiday_pay
                    + overtime_over_60h_pay
                    + paid_leave_amount
                    + transport_allowance
                    + other_allowances
                    + non_billable
                )

            # Build record
            # Note: billing_rate is stored in Employee table, not PayrollRecord
            # The billing calculation is done by services.py using the employee's rate
            data = {
                "employee_id": employee_id,
                "employee_name": employee_name,
                "period": period,
                # Time data
                "work_days": int(work_days),
                "work_hours": work_hours,
                "overtime_hours": overtime_hours,
                "night_hours": night_hours,
                "holiday_hours": holiday_hours,
                "overtime_over_60h": overtime_over_60h,
                "paid_leave_days": payment_data.get("paid_leave_days", 0),
                "paid_leave_hours": 0,  # Not available in this format
                "paid_leave_amount": paid_leave_amount,
                # Salary
                "base_salary": base_salary,
                "overtime_pay": overtime_pay,
                "night_pay": night_pay,
                "holiday_pay": holiday_pay,
                "overtime_over_60h_pay": overtime_over_60h_pay,
                "other_allowances": other_allowances,
                "non_billable_allowances": non_billable,
                "transport_allowance": transport_allowance,
                "gross_salary": gross_salary,
                # Deductions
                "social_insurance": social_insurance,
                "welfare_pension": welfare_pension,
                "employment_insurance": employment_insurance,
                "income_tax": income_tax,
                "resident_tax": resident_tax,
                "rent_deduction": payment_data.get("rent_deduction", 0.0),
                "utilities_deduction": payment_data.get("utilities", 0.0),
                "meal_deduction": payment_data.get("meal_deduction", 0.0),
                "advance_payment": payment_data.get("advance_payment", 0.0),
                "year_end_adjustment": payment_data.get("year_end_adjustment", 0.0),
                "other_deductions": 0,
                "net_salary": net_salary,
                # Billing will be calculated by services.py using employee's billing_rate
                "billing_amount": 0,
                # Metadata
                "dispatch_company": sheet_name,
            }

            return PayrollRecordCreate(**data)

        except Exception as e:
            print(f"[KintaihyoParser] Error extracting worker at row {start_row}: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _extract_kintaihyo_daily_hours(self, ws, row: int, start_col: int) -> Dict[str, Any]:
        """
        Extract daily attendance hours from a Kintaihyo row.

        The daily hours are stored in columns to the right of the name,
        typically one column per day of the month.

        Format examples:
        - [8] = 8 hours worked
        - [8.5] = 8.5 hours worked
        - [0] = no work (absent or holiday)
        - [2] = 2 hours (partial day)

        Args:
            ws: Worksheet
            row: Row containing daily data
            start_col: Starting column (name column)

        Returns:
            Dict with total_hours, work_days, overtime_hours, etc.
        """
        result = {
            "total_hours": 0.0,
            "work_days": 0,
            "overtime_hours": 0.0,
            "night_hours": 0.0,
            "holiday_hours": 0.0,
        }

        # Standard work day hours (8 hours)
        STANDARD_DAY_HOURS = 8.0

        # Scan columns to the right (typically 31 days max + some padding)
        total_regular_hours = 0.0
        total_overtime = 0.0

        for col_offset in range(1, 35):  # Check up to 35 columns
            col = start_col + col_offset
            if col > ws.max_column:
                break

            cell_value = ws.cell(row=row, column=col).value
            if cell_value is None:
                continue

            try:
                # Try to parse as number
                hours = float(str(cell_value).strip().replace(",", ""))

                if hours > 0:
                    result["work_days"] += 1
                    result["total_hours"] += hours

                    # Calculate overtime (hours over 8 per day)
                    if hours > STANDARD_DAY_HOURS:
                        daily_overtime = hours - STANDARD_DAY_HOURS
                        total_overtime += daily_overtime
                        total_regular_hours += STANDARD_DAY_HOURS
                    else:
                        total_regular_hours += hours

            except (ValueError, TypeError):
                # Not a number, might be text label - skip
                continue

        result["overtime_hours"] = total_overtime

        return result

    def _extract_kintaihyo_payments(self, ws, start_row: int, col: int) -> Dict[str, Any]:
        """
        Extract payment and deduction data from Kintaihyo rows 4-6.

        These rows typically contain:
        - Work hours totals
        - Various allowances (overtime pay, night pay, etc.)
        - Deductions (taxes, insurance)
        - Gross and net salary

        The exact layout varies by company, so we scan for known labels.

        Args:
            ws: Worksheet
            start_row: First payment row (typically start_row + 4 of block)
            col: Column of the worker block

        Returns:
            Dict with payment amounts
        """
        result = {
            "base_salary": 0.0,
            "overtime_pay": 0.0,
            "night_pay": 0.0,
            "holiday_pay": 0.0,
            "overtime_over_60h_pay": 0.0,
            "transport_allowance": 0.0,
            "paid_leave_amount": 0.0,
            "paid_leave_days": 0,
            "other_allowances": 0.0,
            "non_billable": 0.0,
            "social_insurance": 0.0,
            "welfare_pension": 0.0,
            "employment_insurance": 0.0,
            "income_tax": 0.0,
            "resident_tax": 0.0,
            "rent_deduction": 0.0,
            "utilities": 0.0,
            "meal_deduction": 0.0,
            "advance_payment": 0.0,
            "year_end_adjustment": 0.0,
            "gross_salary": 0.0,
            "net_salary": 0.0,
        }

        # Scan rows 4-7 (4 rows) and multiple columns for labeled values
        # Labels might be in adjacent columns
        for row_offset in range(4):
            current_row = start_row + row_offset

            # Scan columns around the worker's column (might span multiple cols)
            for col_offset in range(-5, 40):  # Look left and right
                current_col = col + col_offset
                if current_col < 1 or current_col > ws.max_column:
                    continue

                cell_value = ws.cell(row=current_row, column=current_col).value
                if cell_value is None:
                    continue

                label = str(cell_value).strip()
                label_normalized = self._normalize_label(label)

                # Check if this is a known label
                field_type = self._identify_kintaihyo_field(label_normalized)
                if field_type:
                    # Look for value in next column
                    value_col = current_col + 1
                    if value_col <= ws.max_column:
                        value = self._get_numeric(ws, current_row, value_col)
                        if value != 0:
                            # Map field type to result key
                            self._assign_kintaihyo_value(result, field_type, value)

        return result

    def _identify_kintaihyo_field(self, label: str) -> Optional[str]:
        """
        Identify what type of field a Kintaihyo label represents.

        Args:
            label: Normalized label text

        Returns:
            Field type string or None if not recognized
        """
        # Check against standard field patterns
        for field_name, patterns in self.FIELD_PATTERNS.items():
            for pattern in patterns:
                pattern_norm = self._normalize_label(pattern)
                if pattern_norm in label or label in pattern_norm:
                    return field_name

        # Check dynamic zone labels
        for known_label, category in self.DYNAMIC_ZONE_LABELS.items():
            known_norm = self._normalize_label(known_label)
            if known_norm in label or label in known_norm:
                return category

        return None

    def _assign_kintaihyo_value(self, result: Dict, field_type: str, value: float) -> None:
        """
        Assign a value to the appropriate field in the result dict.

        Args:
            result: Result dictionary to update
            field_type: Field type from _identify_kintaihyo_field
            value: Numeric value to assign
        """
        # Map field types to result keys
        field_mapping = {
            "base_salary": "base_salary",
            "overtime_pay": "overtime_pay",
            "night_pay": "night_pay",
            "holiday_pay": "holiday_pay",
            "overtime_over_60h_pay": "overtime_over_60h_pay",
            "transport_allowance": "transport_allowance",
            "paid_leave_amount": "paid_leave_amount",
            "social_insurance": "social_insurance",
            "welfare_pension": "welfare_pension",
            "employment_insurance": "employment_insurance",
            "income_tax": "income_tax",
            "resident_tax": "resident_tax",
            "gross_salary": "gross_salary",
            "net_salary": "net_salary",
            "rent_deduction": "rent_deduction",
            "utilities": "utilities",
            "meal_deduction": "meal_deduction",
            "advance_payment": "advance_payment",
            "year_end_adjustment": "year_end_adjustment",
            # Dynamic zone categories
            "non_billable": "non_billable",
            "other_allowance": "other_allowances",
        }

        if field_type in field_mapping:
            key = field_mapping[field_type]
            # For some fields, accumulate values (allowances)
            if key in ["other_allowances", "non_billable"]:
                result[key] += value
            else:
                # For specific fields, only set if not already set
                if result.get(key, 0) == 0:
                    result[key] = value

    def _apply_template(self, template: Dict[str, Any]) -> None:
        """
        Apply a saved template to the parser state.

        Args:
            template: Template dict from TemplateManager
        """
        self.detected_fields = template.get("field_positions", {}).copy()
        self.detected_allowances = template.get("detected_allowances", {}).copy()
        self.current_column_offsets = template.get(
            "column_offsets", self.COLUMN_OFFSETS.copy()
        )

        # Handle non-billable allowances
        non_billable_list = template.get("non_billable_allowances", [])
        self.detected_non_billable = {}
        for name in non_billable_list:
            if name in self.detected_allowances:
                self.detected_non_billable[name] = self.detected_allowances[name]

    def _save_detected_template(self, ws, sheet_name: str, confidence: float) -> None:
        """
        Save the current detected positions as a template.

        Args:
            ws: Worksheet being parsed
            sheet_name: Factory/sheet name
            confidence: Detection confidence score
        """
        # Find sample employee ID and period for verification
        sample_emp_id = None
        sample_period = None

        emp_id_row = self.detected_fields.get(
            "employee_id"
        ) or self.FALLBACK_ROW_POSITIONS.get("employee_id", 6)
        period_row = self.detected_fields.get(
            "period"
        ) or self.FALLBACK_ROW_POSITIONS.get("period", 10)

        for col in range(1, min(50, ws.max_column + 1)):
            # Find employee ID
            if not sample_emp_id:
                cell_value = ws.cell(row=emp_id_row, column=col).value
                if cell_value:
                    emp_str = str(cell_value).strip()
                    if emp_str.isdigit() and len(emp_str) == 6:
                        sample_emp_id = emp_str

            # Find period
            if not sample_period:
                cell_value = ws.cell(row=period_row, column=col).value
                if cell_value:
                    from datetime import datetime

                    if isinstance(cell_value, datetime):
                        sample_period = f"{cell_value.year}年{cell_value.month}月"
                    else:
                        import re

                        match = re.search(r"(\d{4})年(\d{1,2})月", str(cell_value))
                        if match:
                            sample_period = f"{match.group(1)}年{int(match.group(2))}月"

        # Save template
        self.template_manager.save_template(
            factory_identifier=sheet_name,
            field_positions=self.detected_fields.copy(),
            column_offsets=self.current_column_offsets or self.COLUMN_OFFSETS.copy(),
            detected_allowances=self.detected_allowances.copy(),
            non_billable_allowances=list(self.detected_non_billable.keys()),
            employee_column_width=self.EMPLOYEE_COLUMN_WIDTH,
            detection_confidence=confidence,
            sample_employee_id=sample_emp_id,
            sample_period=sample_period,
            notes="Auto-generated from Excel parsing",
        )

    def _detect_field_positions(self, ws) -> None:
        """
        Scan worksheet to find row positions of known fields by their labels.
        Also detects any 手当 (allowances) dynamically.
        """
        self.detected_fields = {}
        self.detected_allowances = {}
        self.detected_non_billable = {}

        # Scan first 50 rows, looking at label columns
        # Labels are at base_col + COLUMN_OFFSETS['label'] for each employee block
        # Employee blocks start at: 1, 15, 29, 43, 57, ... (spacing = 14)
        label_offset = self.COLUMN_OFFSETS.get("label", 2)
        label_columns = []
        for block_start in range(1, 100, 14):  # Generate for first ~7 employees
            label_columns.append(block_start + label_offset)  # Column with labels

        for row in range(1, min(50, ws.max_row + 1)):
            for col in label_columns:
                cell_value = ws.cell(row=row, column=col).value
                if not cell_value:
                    continue

                label = str(cell_value).strip()
                # Normalize label by removing ALL spaces (both ASCII and full-width)
                label_normalized = label.replace(" ", "").replace("　", "")

                # Check against known field patterns (using normalized label)
                # IMPORTANTE: Usar matching EXACTO para evitar confusiones
                # Por ejemplo: '残業' no debe matchear '残業手当' (que es YEN, no horas)
                for field_name, patterns in self.FIELD_PATTERNS.items():
                    if field_name not in self.detected_fields:
                        for pattern in patterns:
                            pattern_normalized = pattern.replace(" ", "").replace(
                                "　", ""
                            )
                            # Usar EXACT MATCH o si el label COMIENZA con el pattern
                            # Esto evita que '残業' matchee '残業手当'
                            if label_normalized == pattern_normalized:
                                self.detected_fields[field_name] = row
                                break
                            # También aceptar si el label empieza con el pattern (para variantes)
                            # Pero SOLO si no contiene indicadores de pago (手当/代/割増/給)
                            elif label_normalized.startswith(pattern_normalized):
                                # Verificar que no sea un campo de pago disfrazado
                                payment_indicators = ["手当", "代", "割増", "給"]
                                is_payment_field = any(
                                    ind in label_normalized
                                    for ind in payment_indicators
                                )
                                # Si estamos buscando horas y el label tiene indicador de pago, ignorar
                                if "_hours" in field_name or "_days" in field_name:
                                    if not is_payment_field:
                                        self.detected_fields[field_name] = row
                                        break
                                else:
                                    self.detected_fields[field_name] = row
                                    break

                # Check for NON-BILLABLE allowances (通勤手当（非）, 業務手当, etc.)
                if label in self.NON_BILLABLE_ALLOWANCES:
                    if label not in self.detected_non_billable:
                        self.detected_non_billable[label] = row

                # Check for ANY 手当 (allowance) - dynamic detection
                elif self._is_allowance(label) and label not in self.KNOWN_ALLOWANCES:
                    if label not in self.detected_allowances:
                        self.detected_allowances[label] = row

    def _is_allowance(self, label: str) -> bool:
        """Check if a label represents an allowance (手当)"""
        for pattern in self.ALLOWANCE_PATTERNS:
            if re.match(pattern, label):
                return True
        return False

    def _normalize_label(self, label: Any) -> str:
        """
        Normaliza un label para comparación consistente.
        Integrado de ChinginGenerator para mejor detección.

        Handles:
        - Full-width spaces (　) and regular spaces
        - Parentheses and their contents
        - Japanese interpunct (・)
        """
        if label is None:
            return ""
        text = str(label)
        # Remover espacios japoneses (全角) y normales
        text = text.replace("\u3000", "").replace(" ", "").replace("　", "")
        # Remover paréntesis y contenido (ej: 通勤手当（非）→ 通勤手当)
        text = re.sub(r"[（(].*?[）)]", "", text)
        # Remover caracteres de formato japonés
        text = text.replace("・", "").replace("･", "")
        return text.strip()

    def _scan_dynamic_zone_for_employee(self, ws, base_col: int) -> Dict[str, Any]:
        """
        Scan rows 20-29 for a specific employee to find their allowances.

        This is the KEY method for v4.0 - it handles the case where different
        employees have different allowances in different rows.

        Args:
            ws: Worksheet
            base_col: Base column for this employee

        Returns:
            Dict with:
                - overtime_over_60h_pay: float
                - paid_leave_amount: float
                - paid_leave_days: float (NEW - extracted from 'days' column)
                - non_billable_total: float
                - non_billable_details: List[str]
                - other_allowances_total: float
                - other_allowances_details: List[str]
        """
        offsets = self.current_column_offsets or self.COLUMN_OFFSETS
        label_col = base_col + offsets.get("label", 1)
        value_col = base_col + offsets.get("value", 3)
        days_col = base_col + offsets.get("days", 5)  # Column for days (有給日数)

        result = {
            "overtime_over_60h_pay": 0.0,
            "paid_leave_amount": 0.0,
            "paid_leave_days": 0.0,  # NEW: 有給日数 from dynamic zone
            "non_billable_total": 0.0,
            "non_billable_details": [],
            "other_allowances_total": 0.0,
            "other_allowances_details": [],
            # Nuevas deducciones del ChinginGenerator
            "rent_deduction": 0.0,
            "utilities": 0.0,
            "advance_payment": 0.0,
            "meal_deduction": 0.0,
            "year_end_adjustment": 0.0,
        }

        # Scan rows 20-29 for this employee
        for row in range(self.DYNAMIC_ZONE_START, self.DYNAMIC_ZONE_END + 1):
            # Get the label in the employee's label column
            label_cell = ws.cell(row=row, column=label_col)
            label = label_cell.value

            if not label:
                continue

            label_str = str(label).strip()

            # Skip if empty or just whitespace
            if not label_str or label_str in ["", "給", "額"]:
                continue

            # Get the value (yen amount)
            value = self._get_numeric(ws, row, value_col)

            # Check against known labels
            label_normalized = label_str.replace(" ", "").replace("　", "")

            # Check known labels first
            matched = False
            for known_label, category in self.DYNAMIC_ZONE_LABELS.items():
                if known_label in label_normalized or label_normalized in known_label:
                    matched = True
                    if category == "overtime_over_60h_pay":
                        result["overtime_over_60h_pay"] += value
                    elif category == "paid_leave_amount":
                        # Extract BOTH the amount (value) AND the days from this row
                        result["paid_leave_amount"] += value
                        # Get days from 'days' column (same row, different column)
                        days_value = self._get_numeric(ws, row, days_col)
                        if days_value > 0:
                            result["paid_leave_days"] += days_value
                    elif category == "non_billable":
                        result["non_billable_total"] += value
                        result["non_billable_details"].append(
                            f"{label_str}=¥{value:,.0f}"
                        )
                    elif category == "other_allowance":
                        result["other_allowances_total"] += value
                        result["other_allowances_details"].append(
                            f"{label_str}=¥{value:,.0f}"
                        )
                    # NUEVAS categorías del ChinginGenerator (deducciones especiales)
                    elif category == "rent_deduction":
                        result["rent_deduction"] += value
                    elif category == "utilities":
                        result["utilities"] += value
                    elif category == "advance_payment":
                        result["advance_payment"] += value
                    elif category == "meal_deduction":
                        result["meal_deduction"] += value
                    elif category == "year_end_adjustment":
                        result["year_end_adjustment"] += value
                    break

            # If not matched but looks like an allowance, add to other_allowances
            if not matched and self._is_allowance(label_str) and value > 0:
                result["other_allowances_total"] += value
                result["other_allowances_details"].append(f"{label_str}=¥{value:,.0f}")

        return result

    def _detect_employee_columns(self, ws) -> List[int]:
        """
        Find column indices where employee blocks start.
        Employee IDs are 6-digit numbers.
        """
        columns = []

        # Use current column offsets (from template or default)
        offsets = self.current_column_offsets or self.COLUMN_OFFSETS

        # Determine which row has employee IDs
        emp_id_row = self.detected_fields.get(
            "employee_id"
        ) or self.FALLBACK_ROW_POSITIONS.get("employee_id", 6)

        # Scan for 6-digit numbers
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=emp_id_row, column=col).value

            if cell_value is None:
                continue

            try:
                emp_id_str = str(cell_value).strip()
                if emp_id_str.isdigit() and len(emp_id_str) == 6:
                    # Found an employee ID - calculate base column
                    base_col = col - offsets.get("employee_id", 9)
                    if base_col > 0 and base_col not in columns:
                        columns.append(base_col)
            except (ValueError, AttributeError):
                continue

        return sorted(columns)

    def _extract_employee_data(
        self, ws, base_col: int, sheet_name: str
    ) -> Optional[PayrollRecordCreate]:
        """Extract data for one employee using intelligent field detection or template"""
        try:
            # Use current column offsets (from template or default)
            offsets = self.current_column_offsets or self.COLUMN_OFFSETS

            # Get period
            period_row = (
                self.detected_fields.get("period")
                or self.FALLBACK_ROW_POSITIONS["period"]
            )
            period_col = base_col + offsets.get("period", 8)
            period_cell = ws.cell(row=period_row, column=period_col)

            period = self._parse_period(period_cell.value)
            if not period:
                return None

            # Get employee_id
            emp_id_row = self.detected_fields.get(
                "employee_id"
            ) or self.FALLBACK_ROW_POSITIONS.get("employee_id", 6)
            emp_id_col = base_col + offsets.get("employee_id", 9)
            emp_id_cell = ws.cell(row=emp_id_row, column=emp_id_col)
            employee_id = str(emp_id_cell.value or "").strip()

            if not employee_id or not employee_id.isdigit():
                return None

            # Filter out invalid employee IDs (0, 000000)
            if int(employee_id) == 0:
                return None

            # Extract all standard fields
            # work_days usa columna 'days' (offset 5), no 'value'
            work_days_row = self.detected_fields.get(
                "work_days"
            ) or self.FALLBACK_ROW_POSITIONS.get("work_days", 11)
            work_days = self._get_numeric(
                ws, work_days_row, base_col + offsets.get("days", 5)
            )

            paid_leave_days = self._get_field_value(ws, "paid_leave_days", base_col)
            # Use _get_hours_with_minutes for hour fields to include minutes (73h 30m -> 73.5)
            work_hours = self._get_hours_with_minutes(ws, "work_hours", base_col)
            overtime_hours = self._get_hours_with_minutes(
                ws, "overtime_hours", base_col
            )
            night_hours = self._get_hours_with_minutes(ws, "night_hours", base_col)
            holiday_hours = self._get_hours_with_minutes(ws, "holiday_hours", base_col)

            # NOTE: overtime_over_60h_pay is in DYNAMIC ZONE
            # overtime_over_60h (hours) is CALCULATED from overtime_hours when > 60
            overtime_over_60h_pay = 0  # Will be set from dynamic zone

            # Calculate overtime_over_60h hours:
            # If overtime_hours > 60, the excess goes to overtime_over_60h
            # Example: 73h overtime → 60h normal overtime + 13h over-60h
            # IMPORTANT: overtime_hours debe ser máximo 60, el resto va a overtime_over_60h
            overtime_over_60h = (
                max(0, overtime_hours - 60) if overtime_hours > 60 else 0
            )
            # Cap overtime_hours at 60 (excess already moved to overtime_over_60h)
            overtime_hours = min(overtime_hours, 60)

            base_salary = self._get_field_value(ws, "base_salary", base_col)
            overtime_pay = self._get_field_value(ws, "overtime_pay", base_col)
            night_pay = self._get_field_value(ws, "night_pay", base_col)
            holiday_pay = self._get_field_value(ws, "holiday_pay", base_col)
            transport_allowance = self._get_field_value(
                ws, "transport_allowance", base_col
            )

            # NOTE: paid_leave_amount is in DYNAMIC ZONE (有給休暇)
            # It will be set from dynamic zone scanning below
            paid_leave_amount = 0  # Will be set from dynamic zone

            # Extract deductions
            social_insurance = self._get_field_value(ws, "social_insurance", base_col)
            welfare_pension = self._get_field_value(ws, "welfare_pension", base_col)
            employment_insurance = self._get_field_value(
                ws, "employment_insurance", base_col
            )
            income_tax = self._get_field_value(ws, "income_tax", base_col)
            resident_tax = self._get_field_value(ws, "resident_tax", base_col)

            # Get totals from Excel
            gross_salary_excel = self._get_field_value(ws, "gross_salary", base_col)
            net_salary = self._get_field_value(ws, "net_salary", base_col)

            # ================================================================
            # DYNAMIC ZONE SCANNING (Rows 20-29)
            # ================================================================
            # Scan for employee-specific allowances in the dynamic zone
            dynamic_data = self._scan_dynamic_zone_for_employee(ws, base_col)

            # Extract values from dynamic zone
            if "overtime_over_60h_pay" in dynamic_data:
                overtime_over_60h_pay = dynamic_data["overtime_over_60h_pay"]

            if "paid_leave_amount" in dynamic_data:
                paid_leave_amount = dynamic_data["paid_leave_amount"]

            # NEW: Get paid_leave_days from dynamic zone if found
            # (It's in the same row as 有給/有給休暇 but in the 'days' column)
            if dynamic_data.get("paid_leave_days", 0) > 0:
                paid_leave_days = dynamic_data["paid_leave_days"]

            other_allowances_total = dynamic_data.get("other_allowances_total", 0)
            non_billable_total = dynamic_data.get("non_billable_total", 0)

            # paid_leave_hours not available in this Excel format
            paid_leave_hours = 0

            # Calculate gross_salary from components if Excel value is missing
            gross_salary = gross_salary_excel or (
                base_salary
                + overtime_pay
                + night_pay
                + holiday_pay
                + overtime_over_60h_pay
                + paid_leave_amount
                + other_allowances_total
                + transport_allowance
                + non_billable_total
            )

            data = {
                "employee_id": employee_id,
                "period": period,
                # Time data
                "work_days": int(work_days),
                "work_hours": work_hours,
                "overtime_hours": overtime_hours,
                "night_hours": night_hours,
                "holiday_hours": holiday_hours,
                "overtime_over_60h": overtime_over_60h,
                "paid_leave_days": paid_leave_days,
                "paid_leave_hours": paid_leave_hours,
                "paid_leave_amount": paid_leave_amount,
                # Salary
                "base_salary": base_salary,
                "overtime_pay": overtime_pay,
                "night_pay": night_pay,
                "holiday_pay": holiday_pay,
                "overtime_over_60h_pay": overtime_over_60h_pay,
                "other_allowances": other_allowances_total,  # Only billable allowances
                "non_billable_allowances": non_billable_total,  # 通勤手当（非）、業務手当等 - company cost only
                "transport_allowance": transport_allowance,
                "gross_salary": gross_salary,
                # Deductions
                "social_insurance": social_insurance,
                "welfare_pension": welfare_pension,
                "employment_insurance": employment_insurance,
                "income_tax": income_tax,
                "resident_tax": resident_tax,
                "rent_deduction": dynamic_data.get("rent_deduction", 0),
                "utilities_deduction": dynamic_data.get("utilities", 0),
                "meal_deduction": dynamic_data.get("meal_deduction", 0),
                "advance_payment": dynamic_data.get("advance_payment", 0),
                "year_end_adjustment": dynamic_data.get("year_end_adjustment", 0),
                "other_deductions": 0,
                "net_salary": net_salary,
                # Billing will be calculated by services.py
                "billing_amount": 0,
                # Extra: dispatch company from sheet name
                "dispatch_company": sheet_name,
            }

            return PayrollRecordCreate(**data)

        except Exception as e:
            print(
                f"  [ERROR] Error extracting data for employee at column {base_col}: {e}"
            )
            return None

    def _get_field_value(self, ws, field_name: str, base_col: int) -> float:
        """Get value for a field, using detected position or fallback"""
        # Try detected position first
        if field_name in self.detected_fields:
            row = self.detected_fields[field_name]
        elif field_name in self.FALLBACK_ROW_POSITIONS:
            row = self.FALLBACK_ROW_POSITIONS[field_name]
        else:
            return 0.0

        # Use current column offsets (from template or default)
        offsets = self.current_column_offsets or self.COLUMN_OFFSETS
        col = base_col + offsets.get("value", 3)
        return self._get_numeric(ws, row, col)

    def _parse_time_string(self, value) -> Optional[float]:
        """
        Parse time string formats to decimal hours.

        Supported formats:
        - "HH:MM" or "H:MM" (e.g., "08:30" → 8.5, "168:45" → 168.75)
        - "HHhMMm" Japanese format (e.g., "8時30分" → 8.5)
        - "HH時間MM分" extended format
        - Excel time values (datetime.time)

        Returns: decimal hours or None if not a time format
        """
        if value is None or value == "":
            return None

        # Handle Excel time values (datetime.time)
        from datetime import time, datetime, timedelta

        if isinstance(value, time):
            return value.hour + (value.minute / 60.0)

        if isinstance(value, datetime):
            # Excel sometimes stores times as datetime with date 1899-12-30
            return value.hour + (value.minute / 60.0)

        if isinstance(value, timedelta):
            # Excel duration format
            total_seconds = value.total_seconds()
            return total_seconds / 3600.0

        if isinstance(value, (int, float)):
            return None  # Not a time string

        value_str = str(value).strip()

        # Normalize Unicode
        value_str = unicodedata.normalize("NFKC", value_str)

        # Pattern 1: HH:MM or HHH:MM (colon format)
        match = re.match(r"^(\d{1,3}):(\d{2})$", value_str)
        if match:
            hours = int(match.group(1))
            minutes = int(match.group(2))
            return hours + (minutes / 60.0)

        # Pattern 2: Japanese format "8時30分" or "8時間30分"
        match = re.match(r"^(\d{1,3})時間?(\d{1,2})分$", value_str)
        if match:
            hours = int(match.group(1))
            minutes = int(match.group(2))
            return hours + (minutes / 60.0)

        # Pattern 3: Hours only "8時間" or "8h"
        match = re.match(r"^(\d{1,3})(時間?|h)$", value_str)
        if match:
            return float(match.group(1))

        return None

    def _get_hours_with_minutes(self, ws, field_name: str, base_col: int) -> float:
        """Get hours value including minutes (HH:MM -> decimal hours)

        Handles FOUR Excel formats:
        1. Separate columns: Hours in 'value' (col 4), Minutes in 'minutes' (col 10)
           Example: 73h in col 4, 30m in col 10 -> 73.5
        2. Single decimal value: Already decimal in 'value' column
           Example: 13.5 in col 4 -> 13.5
        3. Total minutes only: Hours=0, Minutes=total minutes (プレテック format)
           Example: 0h in col 4, 10080m in col 10 -> 168.0 (10080/60)
        4. Time string format: "HH:MM" or "HH時MM分" directly in cell
           Example: "08:30" in col 4 -> 8.5

        Returns: hours as decimal (e.g., 73h 30m -> 73.5)
        """
        # Get the row for this field
        if field_name in self.detected_fields:
            row = self.detected_fields[field_name]
        elif field_name in self.FALLBACK_ROW_POSITIONS:
            row = self.FALLBACK_ROW_POSITIONS[field_name]
        else:
            return 0.0

        offsets = self.current_column_offsets or self.COLUMN_OFFSETS

        # Get raw cell value first to check for time string format
        hours_col = base_col + offsets.get("value", 3)
        cell = ws.cell(row=row, column=hours_col)
        raw_value = cell.value

        # Format 4: Check if it's a time string format first
        time_parsed = self._parse_time_string(raw_value)
        if time_parsed is not None:
            return time_parsed

        # Get numeric value for other formats
        hours = self._get_numeric(ws, row, hours_col)

        # Check if hours already has decimal (e.g., 13.5)
        # If yes, it's already in decimal format - don't add minutes
        # Use tolerance-based comparison to avoid floating point issues
        if abs(hours - round(hours)) > 0.001:
            # Already decimal (e.g., 13.5), return as-is
            return hours

        # Hours is whole number, check for separate minutes column
        minutes_col = base_col + offsets.get("minutes", 9)
        minutes = self._get_numeric(ws, row, minutes_col)

        # Format 3: Total minutes only (プレテック style)
        # If hours is 0 and minutes is large (>=60), treat minutes as TOTAL minutes
        if hours == 0 and minutes >= 60:
            return minutes / 60.0

        # Format 1: Normal HH:MM format (minutes is 0-59)
        if 0 <= minutes < 60:
            return hours + (minutes / 60.0)

        # Minutes value is invalid (negative), ignore it
        return hours

    def _parse_period(self, value) -> str:
        """Convert period string to standard format (YYYY年M月)"""
        if value is None or value == "":
            return ""

        # Handle datetime objects from Excel
        from datetime import datetime

        if isinstance(value, datetime):
            return f"{value.year}年{value.month}月"

        value_str = str(value)
        match = re.search(r"(\d{4})年(\d{1,2})月", value_str)
        if match:
            year = match.group(1)
            month = match.group(2)
            return f"{year}年{int(month)}月"

        return ""

    def _get_numeric(self, ws, row: int, col: int) -> float:
        """
        Safely extract numeric value from a cell.

        Enhanced to handle Japanese Excel formats:
        - NFKC normalization: Full-width → Half-width (０１２３ → 0123)
        - Japanese negatives: △123, ▲123 → -123
        - Parenthesized negatives: (123) → -123
        - Full-width yen: ￥ → removed
        - Comma separators: 1,234 → 1234
        """
        try:
            cell = ws.cell(row=row, column=col)
            value = cell.value

            if value is None or value == "":
                return 0.0

            # Numeric values pass through directly
            if isinstance(value, (int, float)):
                return float(value)

            value_str = str(value).strip()
            if not value_str:
                return 0.0

            # NFKC normalization: Convert full-width to half-width
            # ０１２３４５６７８９ → 0123456789
            # ．→ .  ，→ ,  （）→ ()
            value_str = unicodedata.normalize("NFKC", value_str)

            # Detect Japanese negative indicators BEFORE removing them
            is_negative = False

            # △ and ▲ indicate negative in Japanese accounting
            if value_str.startswith("△") or value_str.startswith("▲"):
                is_negative = True
                value_str = value_str[1:]  # Remove the symbol
            elif "△" in value_str or "▲" in value_str:
                is_negative = True
                value_str = value_str.replace("△", "").replace("▲", "")

            # Parentheses also indicate negative: (123) → -123
            if value_str.startswith("(") and value_str.endswith(")"):
                is_negative = True
                value_str = value_str[1:-1]

            # Handle minus sign variants
            if value_str.startswith("-") or value_str.startswith("−"):  # ASCII minus and Unicode minus
                is_negative = True
                value_str = value_str[1:]

            # Remove formatting characters
            value_str = (
                value_str
                .replace(",", "")      # Thousands separator
                .replace("¥", "")      # Half-width yen
                .replace("￥", "")     # Full-width yen
                .replace(" ", "")      # Spaces
                .replace("　", "")     # Full-width space
                .replace("円", "")     # Yen kanji
            )

            if not value_str:
                return 0.0

            result = float(value_str)
            return -result if is_negative else result

        except (ValueError, TypeError, AttributeError):
            return 0.0


# ================================================================
# LEGACY PARSER - Para compatibilidad con formato antiguo
# ================================================================
class SalaryStatementParserLegacy:
    """Parser legacy con posiciones fijas (para compatibilidad)"""

    ROW_POSITIONS = {
        "period": 5,
        "employee_id": 6,
        "name": 7,
        "work_days": 11,
        "paid_leave_days": 12,
        "work_hours": 13,
        "overtime_hours": 14,
        "night_hours": 15,
        "holiday_hours": 16,
        "overtime_over_60h": 17,
        "base_salary": 18,
        "overtime_pay": 19,
        "night_pay": 20,
        "holiday_pay": 21,
        "overtime_over_60h_pay": 22,
        "other_allowances": 23,
        "transport_allowance": 24,
        "paid_leave_amount": 25,
        "social_insurance": 26,
        "employment_insurance": 27,
        "income_tax": 28,
        "resident_tax": 29,
        "gross_salary": 30,
        "net_salary": 31,
    }

    COLUMN_OFFSETS = {
        "period": 2,
        "employee_id": 9,
        "name": 2,
        "value": 0,  # FIXED: Values are in same column as employee_id
    }

    def parse(self, content: bytes) -> List[PayrollRecordCreate]:
        """Use the new intelligent parser but with fallback mode"""
        parser = SalaryStatementParser(use_intelligent_mode=False)
        parser.FALLBACK_ROW_POSITIONS = self.ROW_POSITIONS
        return parser.parse(content)
