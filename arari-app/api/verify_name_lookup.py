import io

import openpyxl

from salary_parser import SalaryStatementParser


def create_mock_vertical_file_no_id():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VerticalSheet"

    # Set Vertical Layout Indicators (Column A)
    indicators = ["基本給", "残業手当", "総支給額", "労働日数", "労働時間"]
    for i, ind in enumerate(indicators, start=1):
        ws.cell(row=i, column=1, value=ind)

    # Set Header Row (Row 2 usually)
    # Col 1: Label, Col 2: Label, Col 3: Name Header, Col 4: Name Value (Expected)
    ws.cell(row=2, column=3, value="氏名")
    ws.cell(row=2, column=4, value="テスト 太郎")  # Name found here

    # NO ID provided
    ws.cell(row=2, column=6, value="No.")
    ws.cell(row=2, column=7, value="")  # Empty ID

    # Set Month Header (Row 10)
    ws.cell(row=10, column=3, value="4月分")
    ws.cell(row=10, column=4, value="5月分")
    ws.cell(row=10, column=5, value="6月分")

    # Set Data
    # Row 1 (Base Salary)
    ws.cell(row=1, column=3, value=250000)
    # Row 3 (Gross Salary) - indicators index 3 (start=1 -> 1,2,3)
    ws.cell(row=3, column=3, value=300000)

    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def verify():
    content = create_mock_vertical_file_no_id()

    # Mock Map
    name_map = {"テスト 太郎": "EMP999"}

    parser = SalaryStatementParser(
        use_intelligent_mode=True, employee_name_map=name_map
    )
    records = parser.parse(content)

    print(f"Parsed {len(records)} records")
    for r in records:
        print(f"Record: {r.employee_name} ({r.employee_id}) - {r.period}")
        if r.employee_id == "EMP999" and r.employee_name == "テスト 太郎":
            print("SUCCESS: Resolved ID via Name Lookup")
        else:
            print(f"FAILURE: Expected EMP999, got {r.employee_id}")


if __name__ == "__main__":
    try:
        verify()
    except Exception:
        import traceback

        traceback.print_exc()
