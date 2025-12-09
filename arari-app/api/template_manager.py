"""
Template Manager for Excel Parser
==================================
Manages factory-specific templates for Excel parsing.

Flow:
1. First upload → Try intelligent detection by labels
2. If successful → Generate and save template
3. Future uploads → Load template by factory/sheet name
4. If template fails → Fallback to intelligent detection
"""

import json
import sqlite3
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
from pathlib import Path


class TemplateManager:
    """
    Manages Excel parsing templates per factory (派遣先).

    Templates store:
    - Field name → row position mappings
    - Column layout information
    - Detected allowances
    - Non-billable allowances
    """

    def __init__(self, db_path: Optional[Path] = None):
        """
        Initialize template manager.

        Args:
            db_path: Path to SQLite database. If None, uses default.
        """
        self.db_path = db_path or Path(__file__).parent / "arari_pro.db"
        self._ensure_table_exists()

    def _get_connection(self) -> sqlite3.Connection:
        """Get database connection"""
        conn = sqlite3.connect(str(self.db_path), check_same_thread=False)
        conn.row_factory = sqlite3.Row
        return conn

    def _ensure_table_exists(self) -> None:
        """Create factory_templates table if not exists"""
        conn = self._get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS factory_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                factory_identifier TEXT UNIQUE NOT NULL,
                template_name TEXT,
                field_positions JSON NOT NULL,
                column_offsets JSON NOT NULL,
                detected_allowances JSON DEFAULT '{}',
                non_billable_allowances JSON DEFAULT '[]',
                employee_column_width INTEGER DEFAULT 14,
                detection_confidence REAL DEFAULT 0.0,
                sample_employee_id TEXT,
                sample_period TEXT,
                is_active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                notes TEXT
            )
        """)

        # Create index for faster lookups
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_factory_templates_identifier
            ON factory_templates(factory_identifier)
        """)

        conn.commit()
        conn.close()

    def save_template(
        self,
        factory_identifier: str,
        field_positions: Dict[str, int],
        column_offsets: Dict[str, int],
        detected_allowances: Optional[Dict[str, int]] = None,
        non_billable_allowances: Optional[List[str]] = None,
        employee_column_width: int = 14,
        detection_confidence: float = 0.0,
        sample_employee_id: Optional[str] = None,
        sample_period: Optional[str] = None,
        template_name: Optional[str] = None,
        notes: Optional[str] = None
    ) -> bool:
        """
        Save or update a factory template.

        Args:
            factory_identifier: Unique identifier (usually sheet name or factory name)
            field_positions: Dict mapping field names to row positions
            column_offsets: Dict mapping field types to column offsets
            detected_allowances: Dict of detected 手当 names to row positions
            non_billable_allowances: List of non-billable allowance names
            employee_column_width: Number of columns per employee block
            detection_confidence: How confident the auto-detection was (0-1)
            sample_employee_id: Example employee ID for verification
            sample_period: Example period for verification
            template_name: Human-readable name for the template
            notes: Additional notes about the template

        Returns:
            True if successful, False otherwise
        """
        conn = self._get_connection()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                INSERT INTO factory_templates (
                    factory_identifier, template_name, field_positions, column_offsets,
                    detected_allowances, non_billable_allowances, employee_column_width,
                    detection_confidence, sample_employee_id, sample_period, notes,
                    updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(factory_identifier) DO UPDATE SET
                    template_name = excluded.template_name,
                    field_positions = excluded.field_positions,
                    column_offsets = excluded.column_offsets,
                    detected_allowances = excluded.detected_allowances,
                    non_billable_allowances = excluded.non_billable_allowances,
                    employee_column_width = excluded.employee_column_width,
                    detection_confidence = excluded.detection_confidence,
                    sample_employee_id = excluded.sample_employee_id,
                    sample_period = excluded.sample_period,
                    notes = excluded.notes,
                    updated_at = excluded.updated_at
            """, (
                factory_identifier,
                template_name or factory_identifier,
                json.dumps(field_positions, ensure_ascii=False),
                json.dumps(column_offsets, ensure_ascii=False),
                json.dumps(detected_allowances or {}, ensure_ascii=False),
                json.dumps(non_billable_allowances or [], ensure_ascii=False),
                employee_column_width,
                detection_confidence,
                sample_employee_id,
                sample_period,
                notes,
                datetime.now().isoformat()
            ))

            conn.commit()
            print(f"[Template] Saved template for '{factory_identifier}' "
                  f"({len(field_positions)} fields, confidence={detection_confidence:.2f})")
            return True

        except Exception as e:
            print(f"[Template ERROR] Failed to save template: {e}")
            conn.rollback()
            return False

        finally:
            conn.close()

    def load_template(self, factory_identifier: str) -> Optional[Dict[str, Any]]:
        """
        Load a template by factory identifier.

        Args:
            factory_identifier: Factory/sheet name to look up

        Returns:
            Template dict or None if not found
        """
        conn = self._get_connection()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                SELECT * FROM factory_templates
                WHERE factory_identifier = ? AND is_active = 1
            """, (factory_identifier,))

            row = cursor.fetchone()
            if not row:
                return None

            template = {
                'id': row['id'],
                'factory_identifier': row['factory_identifier'],
                'template_name': row['template_name'],
                'field_positions': json.loads(row['field_positions']),
                'column_offsets': json.loads(row['column_offsets']),
                'detected_allowances': json.loads(row['detected_allowances'] or '{}'),
                'non_billable_allowances': json.loads(row['non_billable_allowances'] or '[]'),
                'employee_column_width': row['employee_column_width'],
                'detection_confidence': row['detection_confidence'],
                'sample_employee_id': row['sample_employee_id'],
                'sample_period': row['sample_period'],
                'created_at': row['created_at'],
                'updated_at': row['updated_at'],
                'notes': row['notes'],
            }

            print(f"[Template] Loaded template for '{factory_identifier}' "
                  f"({len(template['field_positions'])} fields)")
            return template

        except Exception as e:
            print(f"[Template ERROR] Failed to load template: {e}")
            return None

        finally:
            conn.close()

    def find_matching_template(self, sheet_name: str) -> Optional[Dict[str, Any]]:
        """
        Find a template that matches the sheet name.
        Uses fuzzy matching for factory names.

        Args:
            sheet_name: Sheet name from Excel file

        Returns:
            Best matching template or None
        """
        # First, try exact match
        template = self.load_template(sheet_name)
        if template:
            return template

        # Try partial match (factory name might be substring)
        conn = self._get_connection()
        cursor = conn.cursor()

        try:
            # Get all active templates
            cursor.execute("""
                SELECT factory_identifier FROM factory_templates
                WHERE is_active = 1
            """)

            for row in cursor.fetchall():
                factory_id = row['factory_identifier']

                # Check if sheet_name contains factory_identifier or vice versa
                if factory_id in sheet_name or sheet_name in factory_id:
                    return self.load_template(factory_id)

            return None

        except Exception as e:
            print(f"[Template ERROR] Failed to find matching template: {e}")
            return None

        finally:
            conn.close()

    def list_templates(self, include_inactive: bool = False) -> List[Dict[str, Any]]:
        """
        List all templates.

        Args:
            include_inactive: Whether to include inactive templates

        Returns:
            List of template summaries
        """
        conn = self._get_connection()
        cursor = conn.cursor()

        try:
            if include_inactive:
                cursor.execute("SELECT * FROM factory_templates ORDER BY updated_at DESC")
            else:
                cursor.execute("""
                    SELECT * FROM factory_templates
                    WHERE is_active = 1
                    ORDER BY updated_at DESC
                """)

            templates = []
            for row in cursor.fetchall():
                field_positions = json.loads(row['field_positions'])
                templates.append({
                    'id': row['id'],
                    'factory_identifier': row['factory_identifier'],
                    'template_name': row['template_name'],
                    'field_count': len(field_positions),
                    'detection_confidence': row['detection_confidence'],
                    'sample_employee_id': row['sample_employee_id'],
                    'sample_period': row['sample_period'],
                    'is_active': bool(row['is_active']),
                    'created_at': row['created_at'],
                    'updated_at': row['updated_at'],
                })

            return templates

        except Exception as e:
            print(f"[Template ERROR] Failed to list templates: {e}")
            return []

        finally:
            conn.close()

    def delete_template(self, factory_identifier: str, hard_delete: bool = False) -> bool:
        """
        Delete a template.

        Args:
            factory_identifier: Factory to delete
            hard_delete: If True, actually delete. If False, just deactivate.

        Returns:
            True if successful
        """
        conn = self._get_connection()
        cursor = conn.cursor()

        try:
            if hard_delete:
                cursor.execute("""
                    DELETE FROM factory_templates
                    WHERE factory_identifier = ?
                """, (factory_identifier,))
            else:
                cursor.execute("""
                    UPDATE factory_templates
                    SET is_active = 0, updated_at = ?
                    WHERE factory_identifier = ?
                """, (datetime.now().isoformat(), factory_identifier))

            conn.commit()

            if cursor.rowcount > 0:
                action = "deleted" if hard_delete else "deactivated"
                print(f"[Template] {action.capitalize()} template for '{factory_identifier}'")
                return True
            return False

        except Exception as e:
            print(f"[Template ERROR] Failed to delete template: {e}")
            conn.rollback()
            return False

        finally:
            conn.close()

    def get_template_stats(self) -> Dict[str, Any]:
        """
        Get statistics about templates.

        Returns:
            Dict with template statistics
        """
        conn = self._get_connection()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                SELECT
                    COUNT(*) as total,
                    SUM(CASE WHEN is_active = 1 THEN 1 ELSE 0 END) as active,
                    AVG(detection_confidence) as avg_confidence,
                    MAX(updated_at) as last_updated
                FROM factory_templates
            """)

            row = cursor.fetchone()

            return {
                'total_templates': row['total'] or 0,
                'active_templates': row['active'] or 0,
                'average_confidence': row['avg_confidence'] or 0.0,
                'last_updated': row['last_updated'],
            }

        except Exception as e:
            print(f"[Template ERROR] Failed to get stats: {e}")
            return {'total_templates': 0, 'active_templates': 0}

        finally:
            conn.close()


class TemplateGenerator:
    """
    Generates templates from Excel analysis.
    Scans Excel files to detect field positions and create templates.
    """

    # Field patterns to search for (Japanese labels)
    FIELD_PATTERNS = {
        # Time data (時間データ)
        'work_days': ['出勤日数', '出勤日', '勤務日数', '労働日数'],
        'paid_leave_days': ['有給日数', '有休日数', '有給休暇日数'],
        'work_hours': ['労働時間', '勤務時間', '所定時間', '基本時間', '総労働時間', '出勤時間', '就業時間', '実働時間', '実働'],
        'overtime_hours': ['残業時間', '時間外', '残業', '普通残業', '早出残業', '所定外', '所定時間外労働'],
        'night_hours': ['深夜時間', '深夜', '深夜労働', '深夜時間労働'],
        'holiday_hours': ['休日時間', '休日出勤', '休出時間', '法定休日', '公休出勤'],
        'overtime_over_60h': ['60H過', '60時間超', '60h超', '60H超残業'],

        # Salary amounts (給与)
        'base_salary': ['基本給', '基本賃金', '本給'],
        'overtime_pay': ['残業代', '残業手当', '時間外手当'],
        'night_pay': ['深夜手当', '深夜割増', '深夜代'],
        'holiday_pay': ['休日手当', '休日割増', '休出手当'],
        'overtime_over_60h_pay': ['60H過手当', '60時間超手当', '60H超手当'],
        'transport_allowance': ['通勤費', '交通費', '通勤手当'],
        'paid_leave_amount': ['有給金額', '有休金額', '有給手当', '有給支給'],

        # Deductions (控除)
        'social_insurance': ['社会保険', '社保', '健康保険'],
        'employment_insurance': ['雇用保険'],
        'income_tax': ['所得税', '源泉税'],
        'resident_tax': ['住民税', '市民税'],

        # Totals
        'gross_salary': ['総支給額', '支給合計', '総支給', '給与総額'],
        'net_salary': ['差引支給額', '手取り', '振込額', '差引額'],

        # Identifiers
        'employee_id': ['社員番号', '従業員番号', 'ID', '社員ID', '従業員ID'],
        'period': ['期間', '給与月', '対象月', '支給月'],
    }

    # Non-billable allowance names
    NON_BILLABLE_NAMES = [
        '通勤手当', '通勤手当（非）', '通勤費', '業務手当',
    ]

    def __init__(self):
        self.detected_fields: Dict[str, Tuple[int, int]] = {}  # field -> (row, col)
        self.detected_allowances: Dict[str, Tuple[int, int]] = {}
        self.non_billable_found: List[str] = []
        self.confidence_score = 0.0

    def analyze_worksheet(self, ws, sheet_name: str) -> Optional[Dict[str, Any]]:
        """
        Analyze a worksheet and generate a template.

        Args:
            ws: openpyxl worksheet object
            sheet_name: Name of the sheet (used as factory identifier)

        Returns:
            Template dict or None if analysis failed
        """
        import re

        self.detected_fields = {}
        self.detected_allowances = {}
        self.non_billable_found = []

        # Scan rows 1-60 for labels
        max_scan_row = min(60, ws.max_row)
        max_scan_col = min(50, ws.max_column)

        label_positions = []  # Track where we find labels

        for row in range(1, max_scan_row + 1):
            for col in range(1, max_scan_col + 1):
                cell_value = ws.cell(row=row, column=col).value
                if not cell_value:
                    continue

                label = str(cell_value).strip()
                if not label or len(label) > 30:  # Skip very long text
                    continue

                # Normalize: remove spaces
                label_normalized = label.replace(' ', '').replace('　', '')

                # Check against known field patterns
                for field_name, patterns in self.FIELD_PATTERNS.items():
                    if field_name in self.detected_fields:
                        continue  # Already found

                    for pattern in patterns:
                        pattern_normalized = pattern.replace(' ', '').replace('　', '')
                        if pattern_normalized in label_normalized or label_normalized == pattern_normalized:
                            self.detected_fields[field_name] = (row, col)
                            label_positions.append((row, col, field_name))
                            break

                # Check for non-billable allowances
                if label in self.NON_BILLABLE_NAMES:
                    if label not in self.non_billable_found:
                        self.non_billable_found.append(label)
                        self.detected_allowances[label] = (row, col)

                # Check for other 手当 (allowances)
                elif re.match(r'.*手当.*', label) or re.match(r'.*割増.*', label):
                    if label not in self.detected_allowances:
                        self.detected_allowances[label] = (row, col)

        # Calculate confidence based on how many required fields we found
        required_fields = ['employee_id', 'gross_salary', 'base_salary', 'work_hours']
        found_required = sum(1 for f in required_fields if f in self.detected_fields)
        self.confidence_score = found_required / len(required_fields)

        if self.confidence_score < 0.5:
            print(f"[TemplateGen] Low confidence ({self.confidence_score:.2f}) for '{sheet_name}'")
            return None

        # Determine column offsets based on detected positions
        column_offsets = self._calculate_column_offsets(label_positions)

        # Convert field positions to row-only (for the template)
        field_positions = {
            field: pos[0] for field, pos in self.detected_fields.items()
        }

        # Detect employee column width
        employee_width = self._detect_employee_width(ws, field_positions.get('employee_id'))

        # Find sample data for verification
        sample_emp_id, sample_period = self._find_sample_data(ws, field_positions)

        template = {
            'factory_identifier': sheet_name,
            'template_name': sheet_name,
            'field_positions': field_positions,
            'column_offsets': column_offsets,
            'detected_allowances': {
                name: pos[0] for name, pos in self.detected_allowances.items()
            },
            'non_billable_allowances': self.non_billable_found,
            'employee_column_width': employee_width,
            'detection_confidence': self.confidence_score,
            'sample_employee_id': sample_emp_id,
            'sample_period': sample_period,
        }

        print(f"[TemplateGen] Generated template for '{sheet_name}': "
              f"{len(field_positions)} fields, {len(self.detected_allowances)} allowances, "
              f"confidence={self.confidence_score:.2f}")

        return template

    def _calculate_column_offsets(self, label_positions: List[Tuple[int, int, str]]) -> Dict[str, int]:
        """
        Calculate column offsets based on where labels were found.
        """
        # Default offsets
        offsets = {
            'label': 1,      # Where labels typically are
            'value': 3,      # Where values typically are (relative to employee base)
            'days': 5,       # Where day counts are
            'period': 8,     # Where period is
            'employee_id': 9,  # Where employee ID is
        }

        # Try to improve offsets based on detected positions
        label_cols = [pos[1] for pos in label_positions]
        if label_cols:
            # Most common label column
            from collections import Counter
            col_counts = Counter(label_cols)
            most_common_label_col = col_counts.most_common(1)[0][0]
            offsets['label'] = most_common_label_col

        return offsets

    def _detect_employee_width(self, ws, emp_id_row: Optional[int]) -> int:
        """
        Detect how many columns each employee block occupies.
        """
        if not emp_id_row:
            return 14  # Default

        # Find all 6-digit IDs in the employee ID row
        emp_columns = []
        for col in range(1, min(100, ws.max_column + 1)):
            cell_value = ws.cell(row=emp_id_row, column=col).value
            if cell_value:
                try:
                    emp_str = str(cell_value).strip()
                    if emp_str.isdigit() and len(emp_str) == 6:
                        emp_columns.append(col)
                except:
                    pass

        if len(emp_columns) >= 2:
            # Calculate average distance between employees
            distances = [emp_columns[i+1] - emp_columns[i] for i in range(len(emp_columns)-1)]
            avg_distance = sum(distances) / len(distances)
            return int(round(avg_distance))

        return 14  # Default

    def _find_sample_data(self, ws, field_positions: Dict[str, int]) -> Tuple[Optional[str], Optional[str]]:
        """
        Find sample employee ID and period for verification.
        """
        sample_emp_id = None
        sample_period = None

        # Find employee ID
        emp_id_row = field_positions.get('employee_id')
        if emp_id_row:
            for col in range(1, min(50, ws.max_column + 1)):
                cell_value = ws.cell(row=emp_id_row, column=col).value
                if cell_value:
                    try:
                        emp_str = str(cell_value).strip()
                        if emp_str.isdigit() and len(emp_str) == 6:
                            sample_emp_id = emp_str
                            break
                    except:
                        pass

        # Find period
        period_row = field_positions.get('period')
        if period_row:
            from datetime import datetime
            import re

            for col in range(1, min(50, ws.max_column + 1)):
                cell_value = ws.cell(row=period_row, column=col).value
                if cell_value:
                    # Handle datetime
                    if isinstance(cell_value, datetime):
                        sample_period = f"{cell_value.year}年{cell_value.month}月"
                        break
                    # Handle string
                    val_str = str(cell_value)
                    match = re.search(r'(\d{4})年(\d{1,2})月', val_str)
                    if match:
                        sample_period = f"{match.group(1)}年{int(match.group(2))}月"
                        break

        return sample_emp_id, sample_period


def create_template_from_excel(file_content: bytes, template_manager: TemplateManager) -> Dict[str, Any]:
    """
    Convenience function to analyze an Excel file and save templates.

    Args:
        file_content: Binary content of Excel file
        template_manager: TemplateManager instance

    Returns:
        Dict with results (templates created, errors, etc.)
    """
    import openpyxl
    from io import BytesIO

    results = {
        'templates_created': [],
        'templates_failed': [],
        'errors': [],
    }

    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    except Exception as e:
        results['errors'].append(f"Failed to load Excel file: {e}")
        return results

    generator = TemplateGenerator()

    for sheet_name in wb.sheetnames:
        # Skip summary sheets
        if sheet_name in ['集計', 'Summary', '目次', 'Index']:
            continue

        try:
            ws = wb[sheet_name]
            template = generator.analyze_worksheet(ws, sheet_name)

            if template:
                # Save the template
                success = template_manager.save_template(
                    factory_identifier=template['factory_identifier'],
                    field_positions=template['field_positions'],
                    column_offsets=template['column_offsets'],
                    detected_allowances=template['detected_allowances'],
                    non_billable_allowances=template['non_billable_allowances'],
                    employee_column_width=template['employee_column_width'],
                    detection_confidence=template['detection_confidence'],
                    sample_employee_id=template.get('sample_employee_id'),
                    sample_period=template.get('sample_period'),
                )

                if success:
                    results['templates_created'].append(sheet_name)
                else:
                    results['templates_failed'].append(sheet_name)
            else:
                results['templates_failed'].append(sheet_name)

        except Exception as e:
            results['errors'].append(f"Sheet '{sheet_name}': {e}")

    return results
