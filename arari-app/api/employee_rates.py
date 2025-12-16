"""
Load and cache employee hourly rates (時給) and billing rates (単価) from 社員台帳
"""

import os
from typing import Dict, Tuple

from openpyxl import load_workbook


class EmployeeRatesLoader:
    """Load employee rates from 社員台帳 (Employee Master sheet)"""

    # 社員台帳 file location
    SYAIN_FILE = r"D:\【新】社員台帳(UNS)T　2022.04.05～.xlsm"
    SHEET_NAME = "DBGenzaiX"

    # Column positions in DBGenzaiX (1-indexed)
    COL_EMPLOYEE_ID = 2  # Column B: 社員№
    COL_EMPLOYEE_NAME = 8  # Column H: 氏名
    COL_HOURLY_RATE = 14  # Column N: 現給 (時給) - what we PAY the employee
    COL_BILLING_RATE = 16  # Column P: 現単価 (単価) - what factory PAYS us

    def __init__(self):
        """Initialize loader and cache employee rates"""
        self._rates_cache: Dict[str, Tuple[float, float]] = {}
        self._load_rates()

    def _load_rates(self):
        """Load all employee rates from 社員台帳 into cache"""
        if not os.path.exists(self.SYAIN_FILE):
            print(f"Warning: 社員台帳 not found at {self.SYAIN_FILE}")
            return

        try:
            wb = load_workbook(self.SYAIN_FILE, data_only=True)
            if self.SHEET_NAME not in wb.sheetnames:
                print(f"Warning: Sheet '{self.SHEET_NAME}' not found in 社員台帳")
                return

            ws = wb[self.SHEET_NAME]

            # Start from row 2 (row 1 is header)
            for row_num in range(2, ws.max_row + 1):
                emp_id_cell = ws.cell(row=row_num, column=self.COL_EMPLOYEE_ID)
                emp_name_cell = ws.cell(row=row_num, column=self.COL_EMPLOYEE_NAME)
                hourly_cell = ws.cell(row=row_num, column=self.COL_HOURLY_RATE)
                billing_cell = ws.cell(row=row_num, column=self.COL_BILLING_RATE)

                emp_id = emp_id_cell.value
                emp_name = emp_name_cell.value
                hourly_rate = hourly_cell.value
                billing_rate = billing_cell.value

                # Skip if missing employee ID
                if not emp_id:
                    continue

                emp_id_str = str(emp_id).strip()

                # Convert rates to float, default to 0 if missing
                try:
                    hourly = float(hourly_rate) if hourly_rate else 0.0
                    billing = float(billing_rate) if billing_rate else 0.0
                except (ValueError, TypeError):
                    hourly = 0.0
                    billing = 0.0

                # Cache the rates
                self._rates_cache[emp_id_str] = (hourly, billing)

                if hourly > 0 or billing > 0:
                    print(
                        f"Loaded rates for {emp_id_str}: 時給={hourly}, 単価={billing}"
                    )

            print(f"Loaded {len(self._rates_cache)} employee rates from 社員台帳")

        except Exception as e:
            print(f"Error loading 社員台帳: {e}")

    def get_rates(self, employee_id: str) -> Tuple[float, float]:
        """
        Get hourly and billing rates for an employee

        Args:
            employee_id: Employee ID (社員№)

        Returns:
            Tuple of (hourly_rate, billing_rate)
            Returns (0.0, 0.0) if not found
        """
        emp_id_str = str(employee_id).strip()
        return self._rates_cache.get(emp_id_str, (0.0, 0.0))

    def get_hourly_rate(self, employee_id: str) -> float:
        """Get hourly rate (時給) for an employee"""
        return self.get_rates(employee_id)[0]

    def get_billing_rate(self, employee_id: str) -> float:
        """Get billing rate (単価) for an employee"""
        return self.get_rates(employee_id)[1]


# Global instance - load once on startup
_rates_loader = None


def get_rates_loader() -> EmployeeRatesLoader:
    """Get global rates loader instance"""
    global _rates_loader
    if _rates_loader is None:
        _rates_loader = EmployeeRatesLoader()
    return _rates_loader
