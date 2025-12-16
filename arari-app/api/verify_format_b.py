import os
import sys
from io import BytesIO

import openpyxl

# Add current directory to path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from salary_parser import SalaryStatementParser


def create_mock_format_b():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MockFormatB"

    # 1. Setup Headers (Vertical Layout)
    ws.cell(row=1, column=1).value = "基本情報"
    ws.cell(row=2, column=2).value = "氏名"
    ws.cell(row=2, column=3).value = "Test User"
    ws.cell(row=2, column=5).value = "No"
    ws.cell(row=2, column=6).value = "999999"

    # 2. Setup Month Header (Row 5)
    months = [4, 5, 6]
    start_col = 3
    for i, m in enumerate(months):
        ws.cell(row=5, column=start_col + i).value = f"{m}月分"

    # 3. Setup Row Labels (Col 1)
    labels = {
        "work_days": "労働日数",
        "work_hours": "労働時間",
        "base_salary": "基本給",
        "overtime_pay": "残業手当",
        "gross_salary": "総支給額",
        "net_salary": "差引支給額",
    }

    row_map = {}
    current_row = 6
    for key, label in labels.items():
        ws.cell(row=current_row, column=1).value = label
        row_map[key] = current_row
        current_row += 1

    # 4. Fill Data for April (Col 3)
    # 4月 (Index 0 in loop -> Col 3)
    c = 3
    ws.cell(row=row_map["work_days"], column=c).value = 20
    ws.cell(row=row_map["work_hours"], column=c).value = 160.5
    ws.cell(row=row_map["base_salary"], column=c).value = 300000
    ws.cell(row=row_map["overtime_pay"], column=c).value = 50000
    ws.cell(row=row_map["gross_salary"], column=c).value = 350000
    ws.cell(row=row_map["net_salary"], column=c).value = 280000

    # Save to bytes
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def test_format_b():
    print("Generating Mock Format B file...")
    content = create_mock_format_b()

    print("Parsing...")
    parser = SalaryStatementParser()
    records = parser.parse(content)

    print(f"Parsed {len(records)} records.")

    found_april = False
    for r in records:
        print(f"--- Record: {r.period} ({r.employee_name}) ---")
        print(f"  Work Days: {r.work_days}")
        print(f"  Work Hours: {r.work_hours}")
        print(f"  Base Salary: {r.base_salary}")
        print(f"  Overtime: {r.overtime_pay}")
        print(f"  Gross: {r.gross_salary}")

        if "4月" in r.period and r.gross_salary == 350000:
            found_april = True

    if found_april:
        print("SUCCESS: Found expected April record.")
    else:
        print("FAILURE: Did not find expected April record.")


if __name__ == "__main__":
    # Redirect stdout to capture logs from parser
    with open("verify_log_mock.txt", "w", encoding="utf-8") as f:
        sys.stdout = f
        test_format_b()
