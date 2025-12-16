"""Debug Excel structure to find correct columns"""

from pathlib import Path

import openpyxl

excel_path = Path(r"D:\給料明細\給与明細(派遣社員)2025.1(0217支給).xlsm")
wb = openpyxl.load_workbook(excel_path, data_only=False)
sheet = wb[wb.sheetnames[1]]  # PMI sheet

print(f"Sheet: {sheet.title}")
print("=" * 100)

# Find first employee ID
emp_id_found = None
emp_col = None
for row in range(1, 20):
    for col in range(1, 30):
        val = str(sheet.cell(row, col).value or "")
        if val.isdigit() and len(val) == 6:  # Employee IDs are 6 digits
            emp_id_found = val
            emp_col = col
            print(f"Employee ID '{emp_id_found}' found at row {row}, column {col}")
            break
    if emp_id_found:
        break

if not emp_col:
    print("Employee ID not found!")
    exit(1)

# Print grid around employee column
print(f"\nGrid around employee column {emp_col} (columns {emp_col-2} to {emp_col+6}):")
print("=" * 100)
for row in range(1, 35):
    cells = []
    for col_offset in range(-2, 7):
        col = emp_col + col_offset
        val = sheet.cell(row, col).value
        # Format value for display
        if val is None:
            cells.append(".")
        elif isinstance(val, (int, float)):
            cells.append(f"{val:.0f}" if val == int(val) else f"{val:.2f}")
        else:
            val_str = str(val)[:15]
            cells.append(val_str)

    print(f"Row {row:2d}: {' | '.join(f'{c:>15}' for c in cells)}")

wb.close()
