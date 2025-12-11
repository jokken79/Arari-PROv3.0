
import openpyxl
import os
import json

template_path = r"d:\Arari-PROv1.0.25.12.11\Arari-PROv1.0\Arari-PRO\arari-app\api\templates\template_format_b.xlsx"
wb = openpyxl.load_workbook(template_path)
ws = wb.active

mapping = {
    "headers": {},
    "months": {},
    "rows": {}
}

# 1. Find Headers (Name, ID, etc)
for row in range(1, 10):
    for col in range(1, 20):
        val = str(ws.cell(row=row, column=col).value or "").strip().replace(" ", "")
        if "氏名" in val: mapping["headers"]["name"] = (row, col)
        if "No" in val or "社員番号" in val: mapping["headers"]["id"] = (row, col)
        if "部門" in val: mapping["headers"]["dept"] = (row, col)
        if "入社" in val: mapping["headers"]["hire_date"] = (row, col)
        if "年" in val and "度" in val: mapping["headers"]["year"] = (row, col)

# 2. Find Month Columns (Look for N月)
# Usually in one row e.g. Row 4 or 5
month_row_candidate = None
for row in range(1, 10):
    months_found = 0
    for col in range(1, 20):
        val = str(ws.cell(row=row, column=col).value or "").strip()
        if val.endswith("月") and len(val) <= 3:
            try:
                m = int(val.replace("月", ""))
                mapping["months"][m] = col
                months_found += 1
            except: pass
    if months_found > 3:
        month_row_candidate = row
        break

# 3. Find Data Rows (Base Salary etc)
# Use column A and B (1 and 2)
start_row = (month_row_candidate or 5) + 1
for row in range(start_row, 50):
    val = str(ws.cell(row=row, column=1).value or "") + str(ws.cell(row=row, column=2).value or "")
    val = val.strip().replace(" ", "").replace("　", "")
    
    if "日" in val and "勤" in val: mapping["rows"]["work_days"] = row
    if "労働時間" in val: mapping["rows"]["work_hours"] = row
    if "残業" in val and "手当" not in val: mapping["rows"]["overtime_hours"] = row
    
    if "基本給" in val: mapping["rows"]["base_salary"] = row
    if "残業手当" in val: mapping["rows"]["overtime_pay"] = row
    if "通勤" in val: mapping["rows"]["transport_allowance"] = row
    
    if "総支給" in val or "支給計" in val: mapping["rows"]["gross_salary"] = row
    
    if "健康" in val or "健保" in val: mapping["rows"]["social_insurance"] = row
    if "厚生" in val: mapping["rows"]["welfare_pension"] = row
    if "雇用" in val: mapping["rows"]["employment_insurance"] = row
    if "所得" in val: mapping["rows"]["income_tax"] = row
    if "住民" in val: mapping["rows"]["resident_tax"] = row
    
    if "差引" in val or "手取" in val: mapping["rows"]["net_salary"] = row

print(json.dumps(mapping, indent=2))
