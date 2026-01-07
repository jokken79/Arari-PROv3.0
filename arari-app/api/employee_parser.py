#!/usr/bin/env python3
"""
Parser for Excel files containing employee data from DBGenzaiX sheet.
Supports both .xls and .xlsm files.
"""

import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import openpyxl


@dataclass
class EmployeeRecord:
    """Employee record to import"""

    employee_id: str
    name: str
    name_kana: Optional[str] = None
    hourly_rate: float = 0.0
    billing_rate: float = 0.0
    dispatch_company: Optional[str] = None
    status: str = "active"
    hire_date: Optional[str] = None
    department: Optional[str] = None
    employee_type: str = "haken"
    # NEW FIELDS - 2025-12-11
    gender: Optional[str] = None  # 性別: 男/女 or M/F
    birth_date: Optional[str] = None  # 生年月日: Date of birth
    termination_date: Optional[str] = None  # 退社日: Resignation/termination date
    # NEW FIELD - 2026-01-07
    nationality: Optional[str] = None  # 国籍: Vietnam, Philippines, etc.


class DBGenzaiXParser:
    """Parser for DBGenzaiX sheet containing employee master data"""

    # Column name mappings (case-insensitive, supports multiple names)
    COLUMN_MAPPINGS = {
        "employee_id": [
            "employee_id",
            "emp_id",
            "社員id",
            "社員ID",
            "社員番号",
            "empid",
            "社員№",
            "社員no",
            "社員ｎｏ",
            "従業員id",
            "従業員番号",
            "id",
            "no",
            "no.",
            "code",
            "社員コード",
            "従業員コード",
        ],
        "name": ["name", "氏名", "名前", "社員名", "従業員名"],
        "name_kana": ["name_kana", "kana", "フリガナ", "カナ", "氏名カナ", "氏名読み"],
        "hourly_rate": ["hourly_rate", "時給", "時間給", "時給額", "基本時給"],
        "billing_rate": [
            "billing_rate",
            "単価",
            "請求単価",
            "billing",
            "派遣単価",
            "売上単価",
            "請求単価(円)",
            "単価(円)",
            "支払単価",
            "時間単価",
        ],
        "dispatch_company": [
            "dispatch_company",
            "派遣会社",
            "company",
            "companies",
            "派遣先",
            "派遣先名",
            "就業先",
        ],
        "status": ["status", "ステータス", "状態", "稼働状態", "現在", "在籍状況"],
        "hire_date": ["hire_date", "入社日", "hire", "雇用日", "採用日"],
        "department": ["department", "部署", "所属", "部門", "配属先"],
        # NEW FIELDS - 2025-12-11
        "gender": ["gender", "性別", "sex", "男女", "男/女"],
        "birth_date": [
            "birth_date",
            "生年月日",
            "birthday",
            "誕生日",
            "生日",
            "dob",
            "date_of_birth",
        ],
        "termination_date": [
            "termination_date",
            "退社日",
            "退職日",
            "resignation_date",
            "end_date",
            "終了日",
            "離職日",
        ],
        # NEW FIELD - 2026-01-07
        "nationality": [
            "nationality",
            "国籍",
            "国",
            "country",
            "出身国",
            "出身地",
        ],
    }

    # Active statuses (mapped to 'active')
    ACTIVE_STATUSES = ["在籍", "在職中", "アクティブ", "勤務中", "active", "在勤"]

    # Inactive statuses (mapped to 'inactive')
    INACTIVE_STATUSES = ["退社", "退職", "休職", "inactive"]

    def __init__(self):
        self.errors: List[str] = []
        self.warnings: List[str] = []

    def parse_employees(
        self, file_path: str
    ) -> Tuple[List[EmployeeRecord], Dict[str, int]]:
        """
        Parse employees from Excel file.

        Logic:
        1. Try to find 'DBGenzaiX' sheet.
        2. If found, look for headers in first 10 rows.
        3. If sheet not found (or no headers found), search ALL sheets.
        4. First sheet with valid 'employee_id' header wins.
        """
        self.errors = []
        self.warnings = []
        employees = []
        stats = {
            "total_rows": 0,
            "employees_found": 0,
            "rows_skipped": 0,
            "errors": 0,
        }

        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)

            target_sheet = None
            header_row = None
            col_indices = {}

            # Strategy 1: Look for DBGenzaiX sheet specifically
            sheet_name = self._find_sheet(wb, "DBGenzaiX")
            if sheet_name:
                sheet = wb[sheet_name]
                found_row, indices = self._find_header_row(sheet)
                if found_row and indices.get("employee_id"):
                    target_sheet = sheet
                    header_row = found_row
                    col_indices = indices

            # Strategy 2: If no valid DBGenzaiX, search ALL sheets
            if not target_sheet:
                print(
                    f"[DEBUG] DBGenzaiX not found. Scanning {len(wb.sheetnames)} sheets: {wb.sheetnames}"
                )
                for name in wb.sheetnames:
                    sheet = wb[name]
                    print(f"[DEBUG] Checking sheet: {name}")
                    found_row, indices = self._find_header_row(sheet)
                    if found_row:
                        print(
                            f"[DEBUG] Found potential header in {name} at row {found_row}. Indices: {indices}"
                        )
                        if indices.get("employee_id"):
                            target_sheet = sheet
                            header_row = found_row
                            col_indices = indices
                            print(
                                f"[DEBUG] VALID HEADER FOUND in {name} at row {found_row}"
                            )
                            break
                    else:
                        print(f"[DEBUG] No header found in {name}")

            if not target_sheet:
                print("[DEBUG] CRITICAL: No suitable sheet found after scanning all.")
                self.errors.append(
                    "No se encontró ninguna hoja con columna '社員番号' (Employee ID)"
                )
                return employees, stats

            print(
                f"[DEBUG] Processing sheet '{target_sheet.title}' from row {header_row + 1} to {target_sheet.max_row}"
            )

            # Iterate through data rows (start after header_row)
            for row_num in range(header_row + 1, target_sheet.max_row + 1):
                stats["total_rows"] += 1

                # Debug first few rows
                if row_num < header_row + 5:
                    print(f"[DEBUG] Processing row {row_num}...")

                try:
                    # Get values from row
                    row_data = {}
                    for field, col_idx in col_indices.items():
                        if col_idx:
                            cell_value = target_sheet.cell(
                                row=row_num, column=col_idx
                            ).value
                            row_data[field] = cell_value

                    if row_num < header_row + 5:
                        print(f"[DEBUG] Row {row_num} raw data: {row_data}")

                    # Check if employee_id exists (required field)
                    emp_id = str(row_data.get("employee_id", "")).strip()
                    if not emp_id or emp_id == "None":
                        if row_num < header_row + 5:
                            print(f"[DEBUG] Row {row_num} skipped: No Employee ID")
                        stats["rows_skipped"] += 1
                        continue

                    # Build employee record
                    # Determine status based on termination_date if not explicitly set
                    termination_date = self._format_date(
                        row_data.get("termination_date")
                    )
                    explicit_status = row_data.get("status")

                    # If termination_date exists, employee is inactive
                    if termination_date:
                        status = "inactive"
                    elif explicit_status:
                        status = self._map_status(explicit_status)
                    else:
                        status = "active"

                    emp = EmployeeRecord(
                        employee_id=emp_id,
                        name=str(row_data.get("name", "")).strip()
                        or f"Employee {emp_id}",
                        name_kana=self._clean_value(row_data.get("name_kana")),
                        hourly_rate=self._to_float(row_data.get("hourly_rate")),
                        billing_rate=self._to_float(row_data.get("billing_rate")),
                        dispatch_company=self._clean_value(
                            row_data.get("dispatch_company")
                        ),
                        status=status,
                        hire_date=self._format_date(row_data.get("hire_date")),
                        department=self._clean_value(row_data.get("department")),
                        employee_type=self._detect_employee_type(
                            row_data.get("billing_rate")
                        ),
                        # NEW FIELDS
                        gender=self._map_gender(row_data.get("gender")),
                        birth_date=self._format_date(row_data.get("birth_date")),
                        termination_date=termination_date,
                        nationality=self._normalize_nationality(row_data.get("nationality")),
                    )

                    employees.append(emp)
                    stats["employees_found"] += 1
                    print(f"[DEBUG] Added employee: {emp_id}")

                except Exception as e:
                    stats["errors"] += 1
                    print(f"[DEBUG] Error in row {row_num}: {e}")
                    self.errors.append(f"Fila {row_num}: {str(e)}")

            wb.close()

        except FileNotFoundError:
            self.errors.append(f"Archivo no encontrado: {file_path}")
        except Exception as e:
            self.errors.append(f"Error al leer archivo Excel: {str(e)}")

        return employees, stats

    def _find_sheet(self, workbook, sheet_name: str) -> Optional[str]:
        """Find sheet by name (case-insensitive)"""
        for name in workbook.sheetnames:
            if name.lower() == sheet_name.lower():
                return name
        return None

    def _find_header_row(self, sheet) -> Tuple[Optional[int], Dict[str, Optional[int]]]:
        """
        Scan first 15 rows to find a valid header row.
        STRICT VALIDATION: Row must contain 'employee_id' AND at least one other field
        (name, status, dispatch_company, etc.) to be considered a header.
        This prevents false positives on rows that just have "No" or "ID".
        """
        for row in range(1, min(20, sheet.max_row + 1)):
            col_indices = self._detect_columns_in_row(sheet, row)

            # Check for employee_id
            has_id = col_indices.get("employee_id")

            # Check for at least one other critical field
            has_other_field = any(
                [
                    col_indices.get("name"),
                    col_indices.get("status"),
                    col_indices.get("billing_rate"),
                    col_indices.get("dispatch_company"),
                    col_indices.get("hourly_rate"),
                ]
            )

            if has_id and has_other_field:
                return row, col_indices

        # Last resort: if we only found ID but nothing else, maybe return it?
        # But safest is to return None to avoid garbage data.
        return None, {}

    def _detect_columns_in_row(self, sheet, row_num: int) -> Dict[str, Optional[int]]:
        """Detect column indices from a specific row"""
        col_indices: Dict[str, Optional[int]] = {
            field: None for field in self.COLUMN_MAPPINGS
        }

        # Scan columns
        for col_idx in range(1, sheet.max_column + 1):
            header = sheet.cell(row=row_num, column=col_idx).value
            if not header:
                continue

            header_lower = str(header).strip().lower()

            # Check each field mapping
            for field, aliases in self.COLUMN_MAPPINGS.items():
                # Skip if already found
                if col_indices[field]:
                    continue

                for alias in aliases:
                    if header_lower == alias.lower():
                        col_indices[field] = col_idx
                        break

        return col_indices

    def _map_status(self, status_value) -> str:
        """Map status value to 'active' or 'inactive'"""
        if not status_value:
            return "active"

        status_str = str(status_value).strip()
        if status_str in self.ACTIVE_STATUSES:
            return "active"
        elif status_str in self.INACTIVE_STATUSES:
            return "inactive"
        # Default: if not recognized, assume active
        return "active"

    def _detect_employee_type(self, billing_rate) -> str:
        """Determine employee type based on billing_rate"""
        rate = self._to_float(billing_rate)
        return "haken" if rate > 0 else "ukeoi"

    def _to_float(self, value) -> float:
        """Convert value to float, return 0 if invalid"""
        if value is None or value == "" or str(value).lower() == "none":
            return 0.0
        try:
            # Handle Japanese currency format
            if isinstance(value, str):
                value = value.replace("¥", "").replace(",", "").strip()
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def _clean_value(self, value) -> Optional[str]:
        """Clean and validate string value"""
        if value is None or value == "" or str(value).lower() == "none":
            return None
        cleaned = str(value).strip()
        return cleaned if cleaned else None

    def _map_gender(self, value) -> Optional[str]:
        """Map gender value to standardized format (M/F)"""
        if value is None or value == "" or str(value).lower() == "none":
            return None

        gender_str = str(value).strip().upper()

        # Japanese mappings
        if gender_str in ["男", "男性", "M", "MALE", "♂"]:
            return "M"
        elif gender_str in ["女", "女性", "F", "FEMALE", "♀"]:
            return "F"

        return None

    def _format_date(self, value) -> Optional[str]:
        """Format date value to YYYY-MM-DD string"""
        if value is None or value == "" or str(value).lower() == "none":
            return None

        from datetime import datetime

        # If already a datetime object (from Excel)
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")

        # Try to parse string formats
        date_str = str(value).strip()

        # Common Japanese formats: 1990/01/15, 1990-01-15, 1990年1月15日
        patterns = [
            (r"(\d{4})/(\d{1,2})/(\d{1,2})", "%Y/%m/%d"),
            (r"(\d{4})-(\d{1,2})-(\d{1,2})", "%Y-%m-%d"),
            (r"(\d{4})年(\d{1,2})月(\d{1,2})日", None),  # Special handling
        ]

        for pattern, fmt in patterns:
            match = re.match(pattern, date_str)
            if match:
                if fmt:
                    try:
                        parsed = datetime.strptime(date_str, fmt)
                        return parsed.strftime("%Y-%m-%d")
                    except ValueError:
                        pass
                else:
                    # Handle Japanese format 1990年1月15日
                    year, month, day = match.groups()
                    return f"{year}-{int(month):02d}-{int(day):02d}"

        # Try Excel serial date (days since 1900-01-01)
        try:
            serial = float(date_str)
            if 1 < serial < 100000:  # Reasonable range for dates
                from datetime import timedelta

                base = datetime(1899, 12, 30)  # Excel's epoch
                result = base + timedelta(days=serial)
                return result.strftime("%Y-%m-%d")
        except ValueError:
            pass

        return None

    def _normalize_nationality(self, value) -> Optional[str]:
        """
        Normalize nationality value to standard format.
        Maps various representations to consistent country names.
        """
        if value is None or value == "" or str(value).lower() == "none":
            return None

        nationality_str = str(value).strip()

        # Mapping of various representations to normalized names
        NATIONALITY_MAP = {
            # Vietnamese
            "ベトナム": "Vietnam",
            "べとなむ": "Vietnam",
            "vietnam": "Vietnam",
            "viet nam": "Vietnam",
            "vietnamese": "Vietnam",
            "越南": "Vietnam",
            "vn": "Vietnam",
            # Philippines
            "フィリピン": "Philippines",
            "ふぃりぴん": "Philippines",
            "philippines": "Philippines",
            "philippine": "Philippines",
            "filipino": "Philippines",
            "ph": "Philippines",
            # Indonesia
            "インドネシア": "Indonesia",
            "indonesia": "Indonesia",
            "indonesian": "Indonesia",
            "id": "Indonesia",
            # China
            "中国": "China",
            "china": "China",
            "chinese": "China",
            "cn": "China",
            # Thailand
            "タイ": "Thailand",
            "thailand": "Thailand",
            "thai": "Thailand",
            "th": "Thailand",
            # Myanmar
            "ミャンマー": "Myanmar",
            "myanmar": "Myanmar",
            "burma": "Myanmar",
            "burmese": "Myanmar",
            "mm": "Myanmar",
            # Nepal
            "ネパール": "Nepal",
            "nepal": "Nepal",
            "nepalese": "Nepal",
            "np": "Nepal",
            # Brazil
            "ブラジル": "Brazil",
            "brazil": "Brazil",
            "brazilian": "Brazil",
            "br": "Brazil",
            # Peru
            "ペルー": "Peru",
            "peru": "Peru",
            "peruvian": "Peru",
            "pe": "Peru",
            # Japan
            "日本": "Japan",
            "japan": "Japan",
            "japanese": "Japan",
            "jp": "Japan",
            # Korea
            "韓国": "Korea",
            "korea": "Korea",
            "korean": "Korea",
            "kr": "Korea",
            # Cambodia
            "カンボジア": "Cambodia",
            "cambodia": "Cambodia",
            "cambodian": "Cambodia",
            "kh": "Cambodia",
            # Sri Lanka
            "スリランカ": "Sri Lanka",
            "sri lanka": "Sri Lanka",
            "srilanka": "Sri Lanka",
            "lk": "Sri Lanka",
            # Bangladesh
            "バングラデシュ": "Bangladesh",
            "bangladesh": "Bangladesh",
            "bangladeshi": "Bangladesh",
            "bd": "Bangladesh",
        }

        # Try to find in map (case-insensitive)
        nationality_lower = nationality_str.lower()
        if nationality_lower in NATIONALITY_MAP:
            return NATIONALITY_MAP[nationality_lower]

        # Check for partial matches (e.g., "ベトナム人" contains "ベトナム")
        for key, normalized in NATIONALITY_MAP.items():
            if key in nationality_lower or nationality_lower in key:
                return normalized

        # If not found in map, return original (cleaned)
        return nationality_str
