import os

import openpyxl

from employee_parser import DBGenzaiXParser


def debug_import(file_path):
    print(f"--- DEBUGGING IMPORT FOR: {file_path} ---")

    if not os.path.exists(file_path):
        print(f"ERROR: File not found at {file_path}")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Workbook Loaded. Sheet names: {wb.sheetnames}")

    parser = DBGenzaiXParser()

    # 1. Test Sheet Search
    targets = ["DBGenzaiX", "DBUkeoiX"]
    found_sheets = []

    for t in targets:
        actual = parser._find_sheet(wb, t)
        print(f"Target '{t}' -> Found: '{actual}'")
        if actual:
            found_sheets.append(wb[actual])

    if not found_sheets:
        print("No target sheets found. Checking all...")
        found_sheets = wb.worksheets

    # 2. Test Header and Rows
    for sheet in found_sheets:
        print(f"\n--- Analysing Sheet: {sheet.title} ---")
        print(f"Max Row: {sheet.max_row}, Max Col: {sheet.max_column}")

        found_row, indices = parser._find_header_row(sheet)
        print(f"Header Search Result: Row={found_row}")
        print(f"Indices found: {indices}")

        if not found_row:
            print("SKIPPING: No header found.")
            continue

        print(
            f"Header Row Content: {[sheet.cell(found_row, c).value for c in range(1, 15)]}"
        )

        valid_rows = 0
        skipped_rows = 0

        print("\nScanning first 10 data rows...")
        for r in range(found_row + 1, min(found_row + 11, sheet.max_row + 1)):
            row_vals = []
            for field in ["employee_id", "name", "status"]:
                idx = indices.get(field)
                val = sheet.cell(r, idx).value if idx else "N/A"
                row_vals.append(f"{field}={val}")
            print(f"Row {r}: {', '.join(row_vals)}")

        print("\nCounting valid rows in entire sheet...")
        for r in range(found_row + 1, sheet.max_row + 1):
            idx = indices.get("employee_id")
            if idx:
                val = sheet.cell(r, idx).value
                if val and str(val).strip():
                    valid_rows += 1
                else:
                    skipped_rows += 1

    # 3. Test Full Parse
    print("\n--- Testing Full parse_employees() Method ---")
    employees, stats = parser.parse_employees(file_path)
    print(f"Total Employees Parsed: {len(employees)}")
    print(f"Stats: {stats}")

    with open("debug_import_log.txt", "w", encoding="utf-8") as f:
        f.write(f"Total Employees: {len(employees)}\n")
        f.write(f"Stats: {stats}\n")
        for e in employees[:5]:
            f.write(f"Sample: {e.employee_id} - {e.name}\n")


if __name__ == "__main__":
    file_path = r"D:\【新】社員台帳(UNS)T　2022.04.05～.xlsm"
    debug_import(file_path)
