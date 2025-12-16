"""Debug complete row mapping for salary fields"""

import sys
from pathlib import Path

import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

excel_path = Path(r"D:\給料明細\給与明細(派遣社員)2025.1(0217支給).xlsm")
wb = openpyxl.load_workbook(excel_path, data_only=True)  # Read calculated values
sheet = wb[wb.sheetnames[1]]  # PMI sheet

print(f"Sheet: {sheet.title}")
print("=" * 100)

# First employee starts at base_col = 1
base_col = 1
emp_col = base_col + 9  # Employee ID at offset 9
value_col = base_col + 3  # Values at offset 3
days_col = base_col + 5  # Days at offset 5
label_col = base_col + 1  # Labels at offset 1

print(f"\nEmployee ID (row 6, col {emp_col}): {sheet.cell(6, emp_col).value}")
print(f"Period (row 10, col {emp_col - 1}): {sheet.cell(10, emp_col - 1).value}")
print()

# Scan rows 1-50 for all data
print("Complete field mapping:")
print("=" * 100)
for row in range(1, 51):
    label = sheet.cell(row, label_col).value
    value = sheet.cell(row, value_col).value
    days = sheet.cell(row, days_col).value

    # Show rows with any data
    if label or value or days:
        label_str = str(label)[:30] if label else ""
        value_str = (
            f"{value:,.0f}"
            if isinstance(value, (int, float)) and value
            else str(value) if value else ""
        )
        days_str = f"{days}" if days else ""

        print(
            f"Row {row:2d}: Label='{label_str:30s}' | Value={value_str:>15s} | Days={days_str:>5s}"
        )

wb.close()
