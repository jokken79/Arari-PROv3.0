
import os
import shutil
import tempfile
import zipfile
from pathlib import Path
from typing import List, Optional, Dict
import openpyxl
from datetime import datetime

class WageLedgerExportService:
    """
    Service to handle Wage Ledger (賃金台帳) exports.
    Supports:
    - Filling Excel templates (Format B, Format C)
    - Single employee export
    - Batch export (ZIP)
    """

    def __init__(self, template_dir: str):
        self.template_dir = Path(template_dir)

    def generate_ledger(self, employee: Dict, payroll_records: List[Dict], template_name: str, year: int) -> str:
        """
        Generate a single Excel file for an employee using the specified template.
        Returns the path to the generated temporary file.
        """
        template_path = self.template_dir / f"template_{template_name}.xlsx"
        
        # Check if template exists
        if not template_path.exists():
            # Fallback check for variations users might have
            if template_name == "format_b":
                 potential_matches = list(self.template_dir.glob("*format_b*.xlsx"))
                 if potential_matches: template_path = potential_matches[0]
            elif template_name == "format_c":
                 potential_matches = list(self.template_dir.glob("*format_c*.xlsx"))
                 if potential_matches: template_path = potential_matches[0]
        
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_name} (Expected at {template_path})")

        # Create a temp copy to work on
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        
        shutil.copy(template_path, tmp_path)

        try:
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active # Assume data goes into the active sheet

            # ---------------------------------------------------------
            # MAP DATA TO CELLS
            # This mapping depends heavily on the template structure.
            # We will implement a smart mapping that looks for {{placeholders}} 
            # or uses standard positions if placeholders aren't found.
            # ---------------------------------------------------------
            
            # HEADER INFO
            # We'll try to find cells that look like labels and put values next to them
            self._fill_header_info(ws, employee, year)

            # PAYROLL DATA (Jan-Dec)
            # We assume a standard layout where months represent columns or rows.
            # For standard Wage Ledger, months are usually columns.
            self._fill_monthly_data(ws, payroll_records, year)

            wb.save(tmp_path)
            return tmp_path

        except Exception as e:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
            raise e



    def _fill_monthly_data(self, ws, records: List[Dict], year: int):
        """
        Fill 12 months of data using Robust Mapping for Format B.
        """
        # =================================================================================
        # MAPPING CONFIGURATION (Matched to User Image)
        # =================================================================================
        
        # Explicit row numbers from Image Analysis
        ROW_MAPPING = {
            'work_days': 6,             # 労働日数
            'work_hours': 7,            # 労働時間数 (Image Row 7)
            'overtime_hours': 8,        # 時間外労働 (Image Row 8)
            
            'base_salary': 13,          # 基本給 (Image Row 13)
            'overtime_pay': 15,         # 手当 (Row 15 - typically used for Overtime Pay in this layout)
            
            'gross_salary': 26,         # 総支給合計 (Image Row 26)
            
            'employment_insurance': 34, # 雇用保険 (Image Row 34, from bottom up 34 seems right for Emp Ins)
            'social_insurance': 31,     # 社会保険 (Generic/Total?)
            
            # The following are estimated/standard relative positions
            'welfare_pension': 33,      # 厚生年金
            'income_tax': 38,           # 所得税 (Image Row 38 seems to be Tax?)
                                        # Wait, Image Row 40 is Net.
                                        # Let's use search for safety on these lower ones.

            'net_salary': 40            # 差引支給額 (Image Row 40)
        }
        
        # Dynamic Row Adjustment search (Refining the constants)
        # We scan column A (1) for labels to confirm.
        for row in range(1, 45):
            val = str(ws.cell(row=row, column=1).value or "").replace(" ", "").replace("　", "")
            if "労働日数" in val: ROW_MAPPING['work_days'] = row
            if "労働時間" in val: ROW_MAPPING['work_hours'] = row
            if "時間外労働" in val: ROW_MAPPING['overtime_hours'] = row
            if "基本給" in val: ROW_MAPPING['base_salary'] = row
            if "総支給" in val: ROW_MAPPING['gross_salary'] = row
            if "雇用保険" in val: ROW_MAPPING['employment_insurance'] = row
            if "所得税" in val: ROW_MAPPING['income_tax'] = row
            if "住民税" in val: ROW_MAPPING['resident_tax'] = row
            if "差引" in val: ROW_MAPPING['net_salary'] = row

        # =================================================================================
        # 1. FIND MONTH COLUMNS
        # =================================================================================
        month_cols = {} # {1: col_idx, 2: col_idx ...}
        
        # Scan wide area for month headers (Row 3, 4, 5)
        for row in range(2, 7): 
            for col in range(1, 20):
                val = str(ws.cell(row=row, column=col).value or "").strip()
                if "月" in val:
                    # Try to extract month number "1月" -> 1, "04月" -> 4
                    try:
                        m_str = "".join(filter(str.isdigit, val))
                        if m_str:
                            month = int(m_str)
                            if 1 <= month <= 12:
                                month_cols[month] = col
                    except: pass
        
        # Fallback: Image shows 1月 at Col C (3)?
        # Let's check: A(1), B(2), C(3)...
        # If dynamic fails, assume C=1, D=2...
        if not month_cols:
            print("[WARN] No month headers found. Using fallback columns.")
            for m in range(1, 13):
                month_cols[m] = m + 2

        # =================================================================================
        # 2. FILL DATA
        # =================================================================================
        for record in records:
            try:
                p_str = record.get('period', '')
                if '年' in p_str and '月' in p_str:
                     m_part = p_str.split('年')[1].replace('月', '')
                     month = int(m_part)
                else:
                    continue
            except:
                continue

            if month in month_cols:
                col = month_cols[month]
                for field, row_idx in ROW_MAPPING.items():
                    val = record.get(field, 0)
                    if val is not None:
                         # Write value. If 0, maybe write "-" or 0.
                         ws.cell(row=row_idx, column=col).value = val

    def _fill_header_info(self, ws, employee: Dict, year: int):
        """Smart fill header info - Writing to cell BELOW the header"""
        targets = {
            "氏名": employee.get("name", ""),
            "部門": employee.get("dispatch_company", ""),
            "所属": employee.get("dispatch_company", ""),
            "No": employee.get("employee_id", ""),
            "番号": employee.get("employee_id", ""),
            "入社": employee.get("hire_date", ""),
            "年度": f"{year}" # Just the number often, or string
        }
        
        # Also fill big title Year if found "●●年"
        # Scan area
        for row in range(1, 6):
            for col in range(1, 20):
                cell = ws.cell(row=row, column=col)
                val = str(cell.value or "").replace(" ", "").replace("　", "")
                
                # Title Year Logic
                if "年" in val and "賃金台帳" in val:
                    # Replace ●● or just set the year
                    current = str(cell.value)
                    if "●●" in current:
                        cell.value = current.replace("●●", str(year))
                    elif "  " in current: # If existing has space "  年"
                        cell.value = f"{year}年 賃金台帳"
                    else:
                        # Just Prefix
                        cell.value = f"{year}年 賃金台帳"
                
                # Header Logic
                for label, data_val in targets.items():
                    if label in val:
                        # TARGET: Cell BELOW the header (Row + 1)
                        target_cell = ws.cell(row=row+1, column=col)
                        
                        # Safety: Don't overwrite if it looks like a header (unlikely in row+1)
                        # Just write it.
                        target_cell.value = data_val
                        # Also break to avoid matching "氏名" twice for same cell
                        # But maybe "所属" and "氏名" are on same row.


    
    def create_batch_zip(self, export_requests: List[Dict], year: int) -> str:
        """
        Generates multiple Excel files and zips them.
        export_requests: List of {employee: Dict, records: List[Dict], template: str}
        """
        temp_files = []
        try:
            # Create a temp ZIP file
            with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as tmp_zip:
                zip_path = tmp_zip.name
            
            with zipfile.ZipFile(zip_path, 'w') as zf:
                for req in export_requests:
                    emp = req['employee']
                    try:
                         xl_path = self.generate_ledger(emp, req['records'], req['template'], year)
                         temp_files.append(xl_path)
                         
                         filename = f"{emp['employee_id']}_{emp['name']}_{year}_台帳.xlsx"
                         zf.write(xl_path, arcname=filename)
                    except Exception as e:
                        print(f"[ERROR] Failed to generate for {emp['employee_id']}: {e}")
                        # Continue with others
            
            return zip_path

        finally:
            # Cleanup individual excel files
            for f in temp_files:
                if os.path.exists(f):
                    os.unlink(f)

