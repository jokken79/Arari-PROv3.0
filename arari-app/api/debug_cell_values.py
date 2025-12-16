"""Debug script to see actual cell values being read"""

from pathlib import Path

import openpyxl

# Open first Excel file
excel_path = Path(r"D:\給料明細\給与明細(派遣社員)2025.1(0217支給).xlsm")
wb = openpyxl.load_workbook(
    excel_path, data_only=False
)  # Read formulas, not calculated values

# Get first data sheet (skip summary)
sheet = wb[wb.sheetnames[1]]  # First data sheet after 集計
print(f"Sheet: {sheet.title}")
print("=" * 80)

# Check first employee
# Employee ID should be at row 6, col 10
base_col = 10
print(f"\nBase column: {base_col}")
print(f"Employee ID (row 6, col {base_col}): {sheet.cell(6, base_col).value}")
print(f"Period (row 4, col {base_col + 2}): {sheet.cell(4, base_col + 2).value}")
print()

# Check some value cells
# According to COLUMN_OFFSETS, values should be at base_col + 3
value_col = base_col + 3
print(f"Value column: {value_col}")
print()

# Check rows 10-50 for any values
print("Checking rows 10-50 in value column:")
for row in range(10, 51):
    cell_val = sheet.cell(row, value_col).value
    cell_label = sheet.cell(row, base_col + 1).value  # Label column
    if cell_val is not None and cell_val != 0:
        print(f"Row {row}: Label='{cell_label}' Value={cell_val}")

print("\n" + "=" * 80)
print("Checking first 5 rows for first employee:")
for row in range(1, 6):
    print(f"Row {row}:")
    for col in range(base_col, base_col + 6):
        val = sheet.cell(row, col).value
        print(f"  Col {col}: {val}")

wb.close()
