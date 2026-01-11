import subprocess
import threading
import time

"""
粗利 PRO v2.0 - Backend API
FastAPI + SQLite backend for profit management system
"""

# Load environment variables from .env file
import json
import logging  # Import logging
import os
import sqlite3
import sys
import tempfile
import webbrowser
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import uvicorn
from dotenv import load_dotenv
from fastapi import Depends, FastAPI, File, Header, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response, StreamingResponse
from pydantic import BaseModel
from starlette.concurrency import run_in_threadpool

from additional_costs import COST_TYPES as ADDITIONAL_COST_TYPES
from additional_costs import AdditionalCostsService
from agent_commissions import AgentCommissionService
from alerts import AlertService
from audit import AuditService
from auth import AuthService, validate_token
from auth_dependencies import (
    clear_rate_limit,
    log_action,
    rate_limit_login,
    require_admin,
    require_auth,
)
from backup import BackupService
from budget import BudgetService
from database import get_db, init_db
from employee_parser import DBGenzaiXParser
from models import Employee, EmployeeCreate, PayrollRecord, PayrollRecordCreate
from notifications import NotificationService
from reports import ReportService
from roi import ROIService
from salary_parser import SalaryStatementParser
from search import SearchService
from services import ExcelParser, PayrollService
from template_manager import TemplateManager, create_template_from_excel
from validation import ValidationService

# Import modular routers
from routers import (
    employees_router,
    payroll_router,
    statistics_router,
    settings_router,
    additional_costs_router,
    companies_router,
    auth_router,
    alerts_router,
    reports_router,
    audit_router,
    budget_router,
    notifications_router,
    search_router,
    validation_router,
    backup_router,
    roi_router,
    cache_router,
)

load_dotenv()

frontend_process = None

def start_frontend_in_thread(app_dir: Path, port: int):
    global frontend_process

    frontend_path = app_dir # The bundled arari-app directory
    node_exe = None
    npm_exe = None

    if sys.platform == "win32":
        # PyInstaller on Windows might bundle node.exe and npm.cmd
        # We need to find them within the bundle
        if getattr(sys, 'frozen', False):
            # When running as a PyInstaller bundle, sys._MEIPASS is the temp folder
            # where bundled files are extracted.
            # Assuming node.exe is in the root of the bundle, or a specific path
            # within the bundle depending on how they are added.
            # It's usually in the base of the bundle or a 'node' subfolder

            # Search for node.exe and npm.cmd in common PyInstaller locations
            possible_node_paths = [
                Path(sys._MEIPASS) / "node" / "node.exe",
                Path(sys._MEIPASS) / "node.exe",
            ]
            possible_npm_paths = [
                Path(sys._MEIPASS) / "node" / "npm.cmd",
                Path(sys._MEIPASS) / "npm.cmd",
            ]

            for p in possible_node_paths:
                if p.exists():
                    node_exe = p
                    break
            for p in possible_npm_paths:
                if p.exists():
                    npm_exe = p
                    break

            if not node_exe or not npm_exe:
                logging.error(f"Node.js or npm not found in bundle at {sys._MEIPASS}. Node found: {node_exe}, NPM found: {npm_exe}")
                logging.error("Trying system default npm/node. Ensure Node.js is installed on the target system.")
                node_exe = None # Fallback to system PATH
                npm_exe = None # Fallback to system PATH
            else:
                logging.info(f"Found bundled Node.js: {node_exe} and NPM: {npm_exe}")


    # Command to start Next.js development server
    npm_command = [str(npm_exe)] if npm_exe else ["npm"]
    command = npm_command + ["run", "dev"] # We want 'next dev' behavior

    env = os.environ.copy()
    # Need to set NEXT_PUBLIC_API_URL here for the frontend at runtime
    # When bundled, the backend is on 127.0.0.1:8000
    env["NEXT_PUBLIC_API_URL"] = f"http://127.0.0.1:{port}"
    env["BROWSER"] = "none" # Prevent Next.js from opening browser

    # Run in the 'arari-app' directory
    logging.info(f"Starting Next.js frontend with command: {' '.join(command)} in {frontend_path}")
    logging.info(f"Frontend API URL set to: {env['NEXT_PUBLIC_API_URL']}")

    try:
        # Use shell=True for windows if npm.cmd is found as it's a batch file,
        # or if falling back to system npm which might be a global shell command.
        # But for direct executable, it's usually better to be explicit.
        # Let's try without shell=True first and add if needed.
        frontend_process = subprocess.Popen(
            command,
            cwd=frontend_path,
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1, # Line-buffered output
            creationflags=subprocess.CREATE_NEW_PROCESS_GROUP if sys.platform == "win32" else 0 # For Windows to allow graceful termination
        )

        # Read stdout/stderr in separate threads to avoid deadlocks
        def log_stream(stream, log_func):
            for line in iter(stream.readline, ''):
                log_func(f"[Frontend] {line.strip()}")
            stream.close()

        threading.Thread(target=log_stream, args=(frontend_process.stdout, logging.info)).start()
        threading.Thread(target=log_stream, args=(frontend_process.stderr, logging.error)).start()

        logging.info(f"Frontend process started with PID: {frontend_process.pid}")
        # Give frontend some time to start up
        time.sleep(5)

    except FileNotFoundError:
        logging.error("Failed to start frontend: npm/node command not found. Please ensure Node.js and npm are installed and in PATH, or correctly bundled.")
    except Exception as e:
        logging.error(f"Unhandled error starting frontend: {e}")


@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup: Initialize database
    init_db()
    logging.info("Database initialized successfully")

    # Configure logging
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(log_dir / "arari_pro.log"),
            logging.StreamHandler()
        ]
    )
    logging.info("Application startup complete.")

    # Start frontend if running as bundled app
    if getattr(sys, 'frozen', False):
        logging.info("Running as PyInstaller bundled app. Starting Next.js frontend...")
        # Path to the bundled 'arari-app' directory
        # sys._MEIPASS is where PyInstaller extracts bundled files
        # The arari-app folder is added as data: --add-data "arari-app;arari-app"
        bundled_app_path = Path(sys._MEIPASS) / "arari-app"

        # Ensure node_modules exist, if not, attempt npm install
        # This is a fallback in case node_modules were not bundled or corrupted
        if not (bundled_app_path / "node_modules").exists():
            logging.warning(f"node_modules not found in bundled arari-app at {bundled_app_path}. Attempting npm install...")
            try:
                npm_exe = None
                if sys.platform == "win32":
                    possible_npm_paths = [
                        Path(sys._MEIPASS) / "node" / "npm.cmd",
                        Path(sys._MEIPASS) / "npm.cmd",
                    ]
                    for p in possible_npm_paths:
                        if p.exists():
                            npm_exe = p
                            break

                npm_command = [str(npm_exe)] if npm_exe else ["npm"]

                logging.info(f"Running npm install with command: {' '.join(npm_command + ['install', '--force'])} in {bundled_app_path}")
                subprocess.run(
                    npm_command + ["install", "--force"],
                    cwd=bundled_app_path,
                    check=True,
                    capture_output=True,
                    text=True,
                    creationflags=subprocess.CREATE_NEW_PROCESS_GROUP if sys.platform == "win32" else 0
                )
                logging.info("npm install completed in bundled app.")
            except subprocess.CalledProcessError as e:
                logging.error(f"npm install failed in bundled app: {e.stdout}\n{e.stderr}")
                raise RuntimeError("Failed to install frontend dependencies in bundled app.")
            except Exception as e:
                logging.error(f"Error during npm install fallback: {e}")
                raise RuntimeError(f"Failed to install frontend dependencies: {e}")


        frontend_port = 3000 # Default Next.js port
        # Start frontend in a separate thread
        threading.Thread(target=start_frontend_in_thread, args=(bundled_app_path, frontend_port)).start()

        logging.info(f"Frontend started (or attempting to start) on port {frontend_port}")

        # Give frontend some time to fully initialize
        time.sleep(10)
        webbrowser.open(f"http://127.0.0.1:{frontend_port}")
    else:
        logging.info("Running in development/unfrozen environment. Frontend should be started separately.")

    yield
    # Shutdown
    if frontend_process:
        logging.info(f"Terminating frontend process (PID: {frontend_process.pid})...")
        if sys.platform == "win32":
            # For Windows, use taskkill for a more robust termination of process group
            try:
                subprocess.run(["taskkill", "/F", "/T", "/PID", str(frontend_process.pid)], check=True)
            except subprocess.CalledProcessError as e:
                logging.error(f"Error during taskkill: {e}")
        else:
            frontend_process.terminate()
        frontend_process.wait(timeout=10) # Give it some time to terminate
        if frontend_process.poll() is None:
            logging.warning("Frontend process did not terminate gracefully, killing it.")
            frontend_process.kill()
        logging.info("Frontend process terminated.")

    logging.info("Application shutdown complete")

app = FastAPI(
    title="粗利 PRO API",
    description="利益管理システム - Backend API",
    version="2.0.0",
    lifespan=lifespan
)

# CORS middleware for React frontend
# Allow dynamic localhost ports for development + production FRONTEND_URL
FRONTEND_URL = os.environ.get("FRONTEND_URL", "")
cors_origins = []

# Add production frontend URL if configured
if FRONTEND_URL:
    cors_origins.append(FRONTEND_URL)
    # Also allow without trailing slash
    if FRONTEND_URL.endswith("/"):
        cors_origins.append(FRONTEND_URL.rstrip("/"))
    else:
        cors_origins.append(FRONTEND_URL + "/")
    logging.info(f"[CORS] Production frontend URL: {FRONTEND_URL}")

# For development and Vercel preview deployments
# Regex allows: localhost, LAN IPs, and only arari-* Vercel deployments (security fix)
app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins if cors_origins else [],
    allow_origin_regex=r"(http://(localhost|127\.0\.0\.1|192\.168\.\d+\.\d+)(:\d+)?|https://arari[a-z0-9-]*\.vercel\.app)",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],  # Allow client to read response headers
    max_age=3600,  # Cache preflight requests for 1 hour
)

# ============== Include Modular Routers ==============
app.include_router(employees_router)
app.include_router(payroll_router)
app.include_router(statistics_router)
app.include_router(settings_router)
app.include_router(additional_costs_router)
app.include_router(companies_router)
app.include_router(auth_router)
app.include_router(alerts_router)
app.include_router(reports_router)
app.include_router(audit_router)
app.include_router(budget_router)
app.include_router(notifications_router)
app.include_router(search_router)
app.include_router(validation_router)
app.include_router(backup_router)
app.include_router(roi_router)
app.include_router(cache_router)

# ============== Health Check ==============

@app.get("/api/health")
async def health_check():
    return {"status": "healthy", "version": "2.0.0"}

# ============== Employees, Payroll, Statistics ==============
# MOVED TO: routers/employees.py, routers/payroll.py, routers/statistics.py

# ============== Sync Employees ==============

@app.post("/api/sync-employees")
async def sync_employees(
    db: sqlite3.Connection = Depends(get_db),
    current_user: Dict[str, Any] = Depends(require_auth)
):
    """Sync/update employees from ChinginGenerator database (requires authentication)"""
    try:
        from migrate_employees import migrate_employees_sync
        success, stats = migrate_employees_sync(db)

        if success:
            return {
                "success": True,
                "message": f"Sincronización completada: {stats['total_added']} empleados importados/actualizados",
                "stats": stats
            }
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Error en sincronización: {stats.get('error', 'Error desconocido')}"
            )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

# ============== Import Employees Endpoint ==============

@app.post("/api/import-employees")
async def import_employees(
    file: UploadFile = File(...),
    db: sqlite3.Connection = Depends(get_db),
    current_user: Dict[str, Any] = Depends(require_admin)
):
    """
    Dedicated endpoint for importing employees from Excel (DBGenzaiX format).
    Used by EmployeeUploader.tsx (requires admin)
    """
    allowed_extensions = ['.xlsx', '.xlsm', '.xls']
    file_ext = '.' + file.filename.split('.')[-1].lower() if '.' in file.filename else ''

    if file_ext not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type. Allowed: {', '.join(allowed_extensions)}"
        )

    try:
        content = await file.read()

        # Save temp file for parser
        import os
        import tempfile

        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        try:
            parser = DBGenzaiXParser()
            employees, stats = parser.parse_employees(tmp_path)
            logging.info(f"/api/import-employees: Parsed {len(employees)} employees. Stats: {stats}")

            service = PayrollService(db)
            added_count = 0
            updated_count = 0

            for emp in employees:
                emp_data = EmployeeCreate(
                    employee_id=emp.employee_id,
                    name=emp.name,
                    name_kana=emp.name_kana,
                    dispatch_company=emp.dispatch_company if emp.dispatch_company else "Unknown",
                    department=emp.department,
                    hourly_rate=emp.hourly_rate,
                    billing_rate=emp.billing_rate,
                    status=emp.status,
                    hire_date=emp.hire_date,
                    employee_type=emp.employee_type,
                    gender=emp.gender,
                    birth_date=emp.birth_date,
                    termination_date=emp.termination_date,
                )

                existing = service.get_employee(emp.employee_id)
                if existing:
                    service.update_employee(emp.employee_id, emp_data)
                    updated_count += 1
                else:
                    service.create_employee(emp_data)
                    added_count += 1

            return {
                "status": "success",
                "message": f"Successfully imported {len(employees)} employees",
                "employees_added": added_count,
                "employees_updated": updated_count,
                "employees_skipped": stats['rows_skipped'],
                "total_employees": len(employees),
                "errors": parser.errors
            }

        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    except Exception as e:
        logging.error(f"Import failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/upload")
async def upload_payroll_file(
    file: UploadFile = File(...),
    db: sqlite3.Connection = Depends(get_db),
    current_user: Dict[str, Any] = Depends(require_admin)
):
    """Upload and parse a payroll file (Excel or CSV) with Streaming Log (requires admin)"""
    import asyncio
    import concurrent.futures


    async def log_generator():
        try:
            yield json.dumps({"type": "info", "message": f"Upload started: {file.filename}"}) + "\n"

            # Validate file type
            allowed_extensions = ['.xlsx', '.xlsm', '.xls', '.csv']
            file_ext = '.' + file.filename.split('.')[-1].lower() if '.' in file.filename else ''

            if file_ext not in allowed_extensions:
                yield json.dumps({
                    "type": "error",
                    "message": f"Invalid extension. Allowed: {', '.join(allowed_extensions)}"
                }) + "\n"
                return

            # Validate file size (max 50MB)
            MAX_FILE_SIZE = 50 * 1024 * 1024
            content = await file.read()

            if len(content) > MAX_FILE_SIZE:
                 yield json.dumps({
                    "type": "error",
                    "message": f"File too large (>50MB). Size: {len(content) / 1024 / 1024:.2f}MB"
                }) + "\n"
                 return

            yield json.dumps({"type": "info", "message": "File read successfully. Detecting type..."}) + "\n"

            # Shared resources for keepalive mechanism (prevents timeout on large files)
            loop = asyncio.get_event_loop()
            executor = concurrent.futures.ThreadPoolExecutor(max_workers=1)
            keepalive_interval = 5  # seconds

            # ---------------------------------------------------------
            # CASE A: Payroll File (給与明細)
            # ---------------------------------------------------------
            if file_ext in ['.xlsm', '.xlsx'] and ('給与' in file.filename or '給料' in file.filename or '明細' in file.filename):
                yield json.dumps({"type": "info", "message": "Detected: Payroll Statement (給与明細)"}) + "\n"
                yield json.dumps({"type": "progress", "message": "Parsing Excel file (this may take a moment)..."}) + "\n"

                parser = SalaryStatementParser(use_intelligent_mode=True)

                # Run parser in thread with keepalive messages to prevent timeout
                future = loop.run_in_executor(executor, parser.parse, content)

                elapsed = 0
                while not future.done():
                    try:
                        # Wait up to keepalive_interval seconds for result
                        records = await asyncio.wait_for(asyncio.shield(future), timeout=keepalive_interval)
                        break
                    except asyncio.TimeoutError:
                        # Not done yet, yield keepalive message immediately
                        elapsed += keepalive_interval
                        yield json.dumps({
                            "type": "progress",
                            "message": f"Processing Excel... ({elapsed}s elapsed)"
                        }) + "\n"
                else:
                    # Future completed between checks
                    records = future.result()

                if records is None:
                    records = []

                yield json.dumps({"type": "info", "message": f"Parsed {len(records)} records. Saving to database..."}) + "\n"

                service = PayrollService(db)
                saved_count = 0
                error_count = 0

                # Batch save or single transaction? Using single transaction for speed/consistency
                cursor = db.cursor()
                cursor.execute("BEGIN")  # Standard SQL (works in SQLite and PostgreSQL)
                try:
                    total = len(records)
                    for i, record_data in enumerate(records):
                        try:
                            service.create_payroll_record(record_data)
                            saved_count += 1
                        except Exception:
                            error_count += 1

                        # Report progress every 50 records
                        if (i+1) % 50 == 0:
                             yield json.dumps({
                                 "type": "progress",
                                 "message": f"Saving records [{i+1}/{total}]...",
                                 "current": i+1,
                                 "total": total
                             }) + "\n"

                    db.commit()
                    yield json.dumps({
                        "type": "success",
                        "message": f"Successfully saved {saved_count} records.",
                        "stats": {"total": total, "saved": saved_count, "errors": error_count}
                    }) + "\n"

                except Exception as e:
                    db.rollback()
                    yield json.dumps({"type": "error", "message": f"Database Error: {str(e)}"}) + "\n"
                    raise e

            # ---------------------------------------------------------
            # CASE B: Employee Master File (社員台帳)
            # ---------------------------------------------------------
            elif file_ext in ['.xlsx', '.xlsm', '.xls'] and ('社員' in file.filename or 'Employee' in file.filename or '台帳' in file.filename):
                yield json.dumps({"type": "info", "message": "Detected: Employee Master (社員台帳)"}) + "\n"

                with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
                    tmp.write(content)
                    tmp_path = tmp.name

                try:
                    parser = DBGenzaiXParser()
                    yield json.dumps({"type": "progress", "message": "Parsing Employees..."}) + "\n"

                    # Run parser in thread with keepalive messages to prevent timeout
                    emp_future = loop.run_in_executor(executor, parser.parse_employees, tmp_path)
                    elapsed = 0
                    employees = None
                    stats = None
                    while not emp_future.done():
                        try:
                            employees, stats = await asyncio.wait_for(asyncio.shield(emp_future), timeout=keepalive_interval)
                            break
                        except asyncio.TimeoutError:
                            elapsed += keepalive_interval
                            yield json.dumps({
                                "type": "progress",
                                "message": f"Parsing employees... ({elapsed}s elapsed)"
                            }) + "\n"
                    else:
                        employees, stats = emp_future.result()
                    yield json.dumps({"type": "info", "message": f"Found {len(employees)} employees."}) + "\n"

                    service = PayrollService(db)
                    imported_count = 0
                    total = len(employees)

                    for i, emp in enumerate(employees):
                        # Convert/Upsert logic...
                        emp_data = EmployeeCreate(
                            employee_id=emp.employee_id,
                            name=emp.name,
                            name_kana=emp.name_kana,
                            dispatch_company=emp.dispatch_company if emp.dispatch_company else "Unknown",
                            department=emp.department,
                            hourly_rate=emp.hourly_rate,
                            billing_rate=emp.billing_rate,
                            status=emp.status,
                            hire_date=emp.hire_date,
                            employee_type=emp.employee_type,
                            gender=emp.gender,
                            birth_date=emp.birth_date,
                            termination_date=emp.termination_date,
                        )
                        existing = service.get_employee(emp.employee_id)
                        if existing:
                            service.update_employee(emp.employee_id, emp_data)
                        else:
                            service.create_employee(emp_data)

                        imported_count += 1

                        if (i+1) % 50 == 0:
                             yield json.dumps({
                                 "type": "progress",
                                 "message": f"Syncing employees [{i+1}/{total}]...",
                                 "current": i+1,
                                 "total": total
                             }) + "\n"

                    yield json.dumps({
                        "type": "success",
                        "message": f"Imported {imported_count} employees.",
                        "stats": {"total": total, "imported": imported_count}
                    }) + "\n"

                finally:
                    if os.path.exists(tmp_path):
                        os.unlink(tmp_path)

            else:
                # Fallback (Generic)
                yield json.dumps({"type": "info", "message": "Detected: Generic Excel/CSV"}) + "\n"
                parser = ExcelParser()

                # Run parser in thread with keepalive messages to prevent timeout
                gen_future = loop.run_in_executor(executor, parser.parse, content)
                elapsed = 0
                records = None
                while not gen_future.done():
                    try:
                        records = await asyncio.wait_for(asyncio.shield(gen_future), timeout=keepalive_interval)
                        break
                    except asyncio.TimeoutError:
                        elapsed += keepalive_interval
                        yield json.dumps({
                            "type": "progress",
                            "message": f"Processing... ({elapsed}s elapsed)"
                        }) + "\n"
                else:
                    records = gen_future.result()

                if records is None:
                    records = []

                service = PayrollService(db)
                saved_count = 0
                # Simple save loop
                for i, r in enumerate(records):
                    service.create_payroll_record(r)
                    saved_count += 1
                    if (i+1) % 50 == 0:
                        yield json.dumps({"type": "progress", "message": f"Saving {i+1}..."}) + "\n"

                yield json.dumps({
                    "type": "success",
                    "message": f"Saved {saved_count} records.",
                    "stats": {"saved": saved_count}
                }) + "\n"

            # Final Completion Signal
            yield json.dumps({"type": "complete", "message": "Processing Finished."}) + "\n"

            # Clean up executor
            executor.shutdown(wait=False)

        except Exception as e:
            yield json.dumps({"type": "error", "message": f"Critical Error: {str(e)}"}) + "\n"

    # Add explicit headers for CORS and streaming compatibility
    headers = {
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",  # Disable nginx buffering
        "Connection": "keep-alive",
    }
    return StreamingResponse(log_generator(), media_type="application/x-ndjson", headers=headers)

# ============== Export ==============

@app.get("/api/export/employees")
async def export_employees(db: sqlite3.Connection = Depends(get_db)):
    """Export all employees as JSON"""
    service = PayrollService(db)
    return service.get_employees()

@app.get("/api/export/payroll")
async def export_payroll(
    db: sqlite3.Connection = Depends(get_db),
    period: Optional[str] = None
):
    """Export payroll records as JSON"""
    service = PayrollService(db)
    return service.get_payroll_records(period=period)

@app.get("/api/export/all")
async def export_all_data(
    format: str = "excel",
    db: sqlite3.Connection = Depends(get_db)
):
    """Export all data (employees + payroll) as Excel"""
    from datetime import datetime
    service = PayrollService(db)
    employees = service.get_employees()
    payroll = service.get_payroll_records()

    if format == "excel":
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()

        # Employees sheet
        ws_emp = wb.active
        ws_emp.title = "従業員一覧"
        if employees:
            # Headers
            headers = list(employees[0].keys())
            for col, header in enumerate(headers, 1):
                ws_emp.cell(row=1, column=col, value=header)
            # Data
            for row, emp in enumerate(employees, 2):
                for col, key in enumerate(headers, 1):
                    ws_emp.cell(row=row, column=col, value=emp.get(key))

        # Payroll sheet
        ws_pay = wb.create_sheet("給与明細")
        if payroll:
            headers = list(payroll[0].keys())
            for col, header in enumerate(headers, 1):
                ws_pay.cell(row=1, column=col, value=header)
            for row, record in enumerate(payroll, 2):
                for col, key in enumerate(headers, 1):
                    ws_pay.cell(row=row, column=col, value=record.get(key))

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"arari_pro_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    return {"employees": employees, "payroll": payroll}


# ---------------------------------------------------------
# Wage Ledger Export (賃金台帳)
# ---------------------------------------------------------
class WageLedgerRequest(BaseModel):
    template_name: str  # 'format_b' or 'format_c'
    year: int
    target: str         # 'single' or 'all_in_company'
    employee_id: Optional[str] = None
    company_name: Optional[str] = None

@app.post("/api/export/wage-ledger")
async def export_wage_ledger(
    request: WageLedgerRequest,
    db: sqlite3.Connection = Depends(get_db)
):
    from api.export_service import WageLedgerExportService
    from starlette.background import BackgroundTask

    # Path to templates
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    TEMPLATE_DIR = os.path.join(BASE_DIR, "templates")

    export_service = WageLedgerExportService(TEMPLATE_DIR)
    payroll_service = PayrollService(db)

    try:
        if request.target == 'single':
            if not request.employee_id:
                raise HTTPException(status_code=400, detail="Employee ID required for single export")

            # Fetch employee as dict (handling object or dict return)
            emp_obj = payroll_service.get_employee(request.employee_id)
            if not emp_obj:
                 raise HTTPException(status_code=404, detail="Employee not found")

            # Convert to dict for export service
            employee = emp_obj.__dict__ if hasattr(emp_obj, '__dict__') else emp_obj

            records = payroll_service.get_payroll_by_employee_year(request.employee_id, request.year)

            file_path = await run_in_threadpool(
                export_service.generate_ledger,
                employee, records, request.template_name, request.year
            )

            filename = f"賃金台帳_{employee['name']}_{request.year}.xlsx"
            return FileResponse(
                path=file_path,
                filename=filename,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                background=BackgroundTask(lambda: os.unlink(file_path))
            )

        elif request.target == 'all_in_company':
            if not request.company_name:
                raise HTTPException(status_code=400, detail="Company Name required for batch export")

            # 1. Get all employees in company
            all_employees = payroll_service.get_employees(company=request.company_name)

            export_requests = []

            for emp in all_employees:
                emp_dict = emp.__dict__ if hasattr(emp, '__dict__') else emp
                records = payroll_service.get_payroll_by_employee_year(emp_dict['employee_id'], request.year)

                export_requests.append({
                    "employee": emp_dict,
                    "records": records,
                    "template": request.template_name
                })

            if not export_requests:
                raise HTTPException(status_code=404, detail=f"No employees found for company: {request.company_name}")

            zip_path = await run_in_threadpool(
                export_service.create_batch_zip,
                export_requests, request.year
            )

            filename = f"賃金台帳_{request.company_name}_{request.year}.zip"
            return FileResponse(
                path=zip_path,
                filename=filename,
                media_type='application/zip',
                background=BackgroundTask(lambda: os.unlink(zip_path))
            )

        else:
             raise HTTPException(status_code=400, detail="Invalid target specified")

    except Exception as e:
        logging.exception(f"Error in wage ledger export: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/sync-from-folder")
async def sync_from_folder(
    payload: dict,
    db: sqlite3.Connection = Depends(get_db),
    current_user: Dict[str, Any] = Depends(require_admin)
):
    """Sync all .xlsm files from a folder path (Streaming Log) (requires admin)"""
    import json
    from pathlib import Path

    from fastapi.responses import StreamingResponse

    folder_path = payload.get("folder_path", "").strip()

    if not folder_path:
        raise HTTPException(status_code=400, detail="folder_path is required")

    # Normalize path for Windows
    folder_path = folder_path.replace("/", "\\")
    path = Path(folder_path)

    if not path.exists():
        raise HTTPException(status_code=400, detail=f"Folder not found: {folder_path}")

    if not path.is_dir():
        raise HTTPException(status_code=400, detail=f"Path is not a directory: {folder_path}")

    async def log_generator():
        # Setup initial state
        yield json.dumps({"type": "info", "message": f"Scanning folder: {folder_path}..."}) + "\n"

        xlsm_files = list(path.glob("*.xlsm"))

        if not xlsm_files:
             yield json.dumps({
                 "type": "error",
                 "message": f"No .xlsm files found in {folder_path}",
                 "files_found": 0
             }) + "\n"
             return

        yield json.dumps({
            "type": "info",
            "message": f"Found {len(xlsm_files)} .xlsm files. Starting process...",
            "files_found": len(xlsm_files)
        }) + "\n"

        service = PayrollService(db)
        total_saved = 0
        total_errors = 0
        files_processed = 0

        for i, file_path in enumerate(xlsm_files):
            filename = file_path.name
            try:
                yield json.dumps({
                    "type": "progress",
                    "message": f"[{i+1}/{len(xlsm_files)}] Processing: {filename}...",
                    "current": i+1,
                    "total": len(xlsm_files),
                    "filename": filename
                }) + "\n"

                # Read file content
                with open(file_path, "rb") as f:
                    file_content = f.read()

                # Detect file type and parse
                lower_name = filename.lower()
                if "給" in lower_name or "給与" in lower_name or "給料" in lower_name:
                    parser = SalaryStatementParser(use_intelligent_mode=True)
                    payroll_records = parser.parse(file_content)
                else:
                    parser = ExcelParser()
                    payroll_records = parser.parse(file_content)

                # Insert records
                cursor = db.cursor()
                cursor.execute("BEGIN")  # Standard SQL (works in SQLite and PostgreSQL)
                file_saved_count = 0

                try:
                    for record_data in payroll_records:
                        try:
                            service.create_payroll_record(record_data)
                            file_saved_count += 1
                            total_saved += 1
                        except (ValueError, TypeError, sqlite3.Error) as record_err:
                            # Skip bad records but keep others from same file
                            logging.warning(f"Skipping invalid record: {record_err}")

                    db.commit()
                    files_processed += 1

                    yield json.dumps({
                        "type": "success",
                        "message": f"  -> Success: Saved {file_saved_count} records.",
                        "file_saved": file_saved_count
                    }) + "\n"

                except Exception as db_err:
                    db.rollback()
                    raise db_err

            except Exception as e:
                total_errors += 1
                yield json.dumps({
                    "type": "error",
                    "message": f"  -> Error processing {filename}: {str(e)}"
                }) + "\n"

        # Summary
        yield json.dumps({
            "type": "complete",
            "message": f"\n--- SYNC COMPLETED ---\nFiles Processed: {files_processed}\nTotal Records Saved: {total_saved}\nErrors: {total_errors}",
            "stats": {
                "files_processed": files_processed,
                "total_saved": total_saved,
                "total_errors": total_errors
            }
        }) + "\n"

    return StreamingResponse(log_generator(), media_type="application/x-ndjson")


# NOTE: Duplicate endpoint removed - using /api/import-employees defined at line ~216

# ============== SETTINGS, COMPANIES, ADDITIONAL COSTS ==============
# MOVED TO: routers/settings.py, routers/companies.py, routers/additional_costs.py

# ============== AGENT COMMISSIONS (仲介手数料) ==============

@app.get("/api/agent-commissions/agents")
async def get_available_agents(db: sqlite3.Connection = Depends(get_db)):
    """Get list of available agents with their configurations"""
    service = AgentCommissionService(db)
    return service.get_available_agents()


@app.get("/api/agent-commissions/calculate/{agent_id}")
async def calculate_agent_commission(
    agent_id: str,
    period: str,
    company: Optional[str] = None,
    db: sqlite3.Connection = Depends(get_db)
):
    """Calculate commission for an agent for a given period"""
    service = AgentCommissionService(db)
    result = service.calculate_commission(
        agent_id=agent_id,
        period=period,
        company_filter=company
    )
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@app.post("/api/agent-commissions/register")
async def register_commission_to_costs(
    payload: dict,
    db: sqlite3.Connection = Depends(get_db),
    current_user: Dict[str, Any] = Depends(require_auth)
):
    """Register calculated commission to additional costs"""
    service = AgentCommissionService(db)

    agent_id = payload.get("agent_id")
    period = payload.get("period")
    company = payload.get("company")
    amount = payload.get("amount")
    notes = payload.get("notes")

    if not all([agent_id, period, company, amount]):
        raise HTTPException(status_code=400, detail="Missing required fields")

    # Registration now handles duplicate check atomically within a transaction
    # to prevent TOCTOU race conditions
    result = service.register_to_additional_costs(
        agent_id=agent_id,
        period=period,
        company=company,
        amount=amount,
        notes=notes
    )

    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])

    log_action(db, current_user, "create", "agent_commission",
               f"{agent_id}_{period}_{company}",
               f"Registered commission: ¥{amount}")

    return result


@app.get("/api/agent-commissions/history")
async def get_commission_history(
    agent_id: Optional[str] = None,
    period: Optional[str] = None,
    db: sqlite3.Connection = Depends(get_db)
):
    """Get history of registered commissions"""
    service = AgentCommissionService(db)
    return service.get_commission_history(agent_id=agent_id, period=period)


@app.get("/api/agent-commissions/check/{agent_id}")
async def check_commission_registered(
    agent_id: str,
    period: str,
    company: str,
    db: sqlite3.Connection = Depends(get_db)
):
    """Check if commission is already registered"""
    service = AgentCommissionService(db)
    is_registered = service.is_already_registered(agent_id, period, company)
    return {"registered": is_registered}


# ============== TEMPLATES ==============

@app.get("/api/templates")
async def list_templates(include_inactive: bool = False):
    """List all factory templates"""
    template_manager = TemplateManager()
    templates = template_manager.list_templates(include_inactive=include_inactive)
    stats = template_manager.get_template_stats()
    return {
        "templates": templates,
        "stats": stats
    }

@app.get("/api/templates/{factory_id}")
async def get_template(factory_id: str):
    """Get a specific template by factory identifier"""
    template_manager = TemplateManager()
    template = template_manager.load_template(factory_id)
    if not template:
        raise HTTPException(status_code=404, detail=f"Template not found: {factory_id}")
    return template

@app.put("/api/templates/{factory_id}")
async def update_template(
    factory_id: str,
    payload: dict,
    current_user: Dict[str, Any] = Depends(require_admin)
):
    """Update a template's field positions or settings (requires admin)"""
    template_manager = TemplateManager()

    # Load existing template
    existing = template_manager.load_template(factory_id)
    if not existing:
        raise HTTPException(status_code=404, detail=f"Template not found: {factory_id}")

    # Merge updates
    field_positions = payload.get('field_positions', existing.get('field_positions', {}))
    column_offsets = payload.get('column_offsets', existing.get('column_offsets', {}))
    detected_allowances = payload.get('detected_allowances', existing.get('detected_allowances', {}))
    non_billable_allowances = payload.get('non_billable_allowances', existing.get('non_billable_allowances', []))
    template_name = payload.get('template_name', existing.get('template_name'))
    notes = payload.get('notes', existing.get('notes'))

    success = template_manager.save_template(
        factory_identifier=factory_id,
        field_positions=field_positions,
        column_offsets=column_offsets,
        detected_allowances=detected_allowances,
        non_billable_allowances=non_billable_allowances,
        template_name=template_name,
        notes=notes,
        detection_confidence=existing.get('detection_confidence', 0.0),
        sample_employee_id=existing.get('sample_employee_id'),
        sample_period=existing.get('sample_period'),
    )

    if not success:
        raise HTTPException(status_code=500, detail="Failed to update template")

    return {"status": "success", "message": f"Template '{factory_id}' updated"}

@app.delete("/api/templates/{factory_id}")
async def delete_template(
    factory_id: str,
    hard_delete: bool = False,
    current_user: Dict[str, Any] = Depends(require_admin)
):
    """Delete (or deactivate) a template (requires admin)"""
    template_manager = TemplateManager()
    success = template_manager.delete_template(factory_id, hard_delete=hard_delete)

    if not success:
        raise HTTPException(status_code=404, detail=f"Template not found: {factory_id}")

    action = "deleted" if hard_delete else "deactivated"
    return {"status": "success", "message": f"Template '{factory_id}' {action}"}

@app.post("/api/templates/analyze")
async def analyze_excel_for_templates(file: UploadFile = File(...)):
    """
    Analyze an Excel file and generate templates for all sheets.
    Does NOT import payroll data - only creates templates.
    """
    # Validate file type
    allowed_extensions = ['.xlsx', '.xlsm', '.xls']
    file_ext = '.' + file.filename.split('.')[-1].lower() if '.' in file.filename else ''

    if file_ext not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type. Allowed: {', '.join(allowed_extensions)}"
        )

    try:
        content = await file.read()
        template_manager = TemplateManager()

        results = create_template_from_excel(content, template_manager)

        return {
            "status": "success",
            "filename": file.filename,
            "templates_created": results['templates_created'],
            "templates_failed": results['templates_failed'],
            "errors": results['errors'][:10] if results['errors'] else None
        }

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error analyzing file: {str(e)}")

@app.post("/api/templates/create")
async def create_template_manually(
    payload: dict,
    current_user: Dict[str, Any] = Depends(require_admin)
):
    """Create a new template manually (for custom factory formats) (requires admin)"""
    template_manager = TemplateManager()

    factory_identifier = payload.get('factory_identifier')
    if not factory_identifier:
        raise HTTPException(status_code=400, detail="'factory_identifier' is required")

    field_positions = payload.get('field_positions', {})
    if not field_positions:
        raise HTTPException(status_code=400, detail="'field_positions' is required")

    success = template_manager.save_template(
        factory_identifier=factory_identifier,
        field_positions=field_positions,
        column_offsets=payload.get('column_offsets', {
            'label': 1,
            'value': 3,
            'days': 5,
            'period': 8,
            'employee_id': 9,
        }),
        detected_allowances=payload.get('detected_allowances', {}),
        non_billable_allowances=payload.get('non_billable_allowances', []),
        employee_column_width=payload.get('employee_column_width', 14),
        detection_confidence=1.0,  # Manual = full confidence
        template_name=payload.get('template_name', factory_identifier),
        notes=payload.get('notes', 'Manually created template'),
    )

    if not success:
        raise HTTPException(status_code=500, detail="Failed to create template")

    return {"status": "success", "message": f"Template '{factory_identifier}' created"}

# ============== AUTH ENDPOINTS ==============
# MOVED TO: routers/auth.py

# ============== ALERTS, AUDIT, REPORTS, BUDGET ENDPOINTS ==============
# MOVED TO: routers/alerts.py, routers/audit.py, routers/reports.py, routers/budget.py

# ============== NOTIFICATIONS, SEARCH, VALIDATION ENDPOINTS ==============
# MOVED TO: routers/notifications.py, routers/search.py, routers/validation.py

# ============== BACKUP, ROI, CACHE ENDPOINTS ==============
# MOVED TO: routers/backup.py, routers/roi.py, routers/cache.py

# ============== RESET DATA ==============

@app.delete("/api/reset-db")
async def reset_database(
    db: sqlite3.Connection = Depends(get_db),
    target: str = "all",
    current_user: Dict[str, Any] = Depends(require_admin)
):
    """
    Delete data from the database based on the target (ADMIN ONLY).
    - `payroll`: Deletes only payroll records.
    - `employees`: Deletes employees and their associated payroll records.
    - `all`: Deletes all data (default).
    """
    if target not in ["payroll", "employees", "all"]:
        raise HTTPException(status_code=400, detail="El parámetro 'target' no es válido. Los valores permitidos son 'payroll', 'employees', 'all'.")

    try:
        cursor = db.cursor()

        if target == "payroll":
            cursor.execute("DELETE FROM payroll_records")
            message = "Todos los registros de nómina han sido eliminados."
        elif target == "employees":
            cursor.execute("DELETE FROM payroll_records")
            cursor.execute("DELETE FROM employees")
            message = "Todos los empleados y sus registros de nómina han sido eliminados."
        elif target == "all":
            cursor.execute("DELETE FROM payroll_records")
            cursor.execute("DELETE FROM employees")
            message = "Todos los datos han sido eliminados."

        db.commit()

        log_action(db, current_user, "delete", "database", target, f"Reset database: {target}")
        logging.warning(f"Database reset by {current_user.get('username')}: target={target}")

        return {"status": "success", "message": message}
    except sqlite3.Error as e:
        db.rollback()
        logging.error(f"Database reset failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to delete data: {str(e)}")

if __name__ == "__main__":
    if getattr(sys, 'frozen', False):
        # Running as PyInstaller bundled app
        uvicorn.run(
            "main:app",
            host="127.0.0.1",
            port=8000,
            reload=False,  # IMPORTANT: Set reload to False for production
            lifespan="on"  # Ensure lifespan events are triggered
        )
    else:
        # Running in development environment
        uvicorn.run(
            "main:app",
            host="0.0.0.0",
            port=8000,
            reload=True
        )
