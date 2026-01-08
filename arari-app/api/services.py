"""
Business logic services for 粗利 PRO
Supports both SQLite (local) and PostgreSQL (Railway production)
"""

import csv
import io
import sqlite3
from typing import Any, Dict, List, Optional

from database import USE_POSTGRES
from models import EmployeeCreate, PayrollRecordCreate


def _q(query: str) -> str:
    """Convert SQLite query to PostgreSQL if needed (? -> %s)"""
    if USE_POSTGRES:
        return query.replace("?", "%s")
    return query


def _get_count(row) -> int:
    """Extract count value from a row (handles dict for PostgreSQL, tuple for SQLite)"""
    if row is None:
        return 0
    if isinstance(row, dict):
        return row.get("count", row.get("COUNT(*)", 0))
    return row[0] if row[0] is not None else 0


def _get_first_col(row):
    """Extract first column value from a row (handles dict for PostgreSQL, tuple for SQLite)"""
    if row is None:
        return None
    if isinstance(row, dict):
        # Get first value from dict
        return next(iter(row.values()), None) if row else None
    return row[0]


def _get_ignored_companies(cursor) -> List[str]:
    """
    Get list of ignored companies from settings.
    Works with both SQLite and PostgreSQL.
    Returns empty list if no ignored companies setting exists.
    """
    try:
        cursor.execute(_q("SELECT value FROM settings WHERE key = ?"), ("ignored_companies",))
        result = cursor.fetchone()
        if result:
            import json
            value = result['value'] if isinstance(result, dict) else result[0]
            if value:
                return json.loads(value)
    except Exception:
        pass
    return []


def _build_company_filter(ignored_companies: List[str], column: str = "e.dispatch_company") -> tuple:
    """
    Build SQL filter clause for excluding ignored companies.
    Returns (sql_clause, params) tuple.

    If no ignored companies, returns empty clause that doesn't filter anything.
    """
    if not ignored_companies:
        return ("", tuple())

    if USE_POSTGRES:
        placeholders = ", ".join(["%s"] * len(ignored_companies))
    else:
        placeholders = ", ".join(["?"] * len(ignored_companies))

    return (f"AND {column} NOT IN ({placeholders})", tuple(ignored_companies))


class PayrollService:
    """Service class for payroll and employee operations"""

    def __init__(self, db):
        self.db = db
        # Set row factory for SQLite (PostgreSQL uses RealDictCursor)
        if not USE_POSTGRES and hasattr(db, 'row_factory'):
            self.db.row_factory = sqlite3.Row

    # ============== Settings Operations ==============

    def get_setting(self, key: str, default: str = None) -> Optional[str]:
        """Get a single setting value by key"""
        cursor = self.db.cursor()
        cursor.execute(_q("SELECT value FROM settings WHERE key = ?"), (key,))
        row = cursor.fetchone()
        return row["value"] if row else default

    def get_all_settings(self) -> List[Dict]:
        """Get all settings"""
        cursor = self.db.cursor()
        cursor.execute("SELECT * FROM settings ORDER BY key")
        return [dict(row) for row in cursor.fetchall()]

    def update_setting(self, key: str, value: str, description: str = None) -> bool:
        """Update or create a setting"""
        cursor = self.db.cursor()
        if description:
            cursor.execute(
                _q("""
                INSERT INTO settings (key, value, description, updated_at)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(key) DO UPDATE SET
                    value = excluded.value,
                    description = excluded.description,
                    updated_at = CURRENT_TIMESTAMP
            """),
                (key, value, description),
            )
        else:
            cursor.execute(
                _q("""
                INSERT INTO settings (key, value, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(key) DO UPDATE SET
                    value = excluded.value,
                    updated_at = CURRENT_TIMESTAMP
            """),
                (key, value),
            )
        self.db.commit()
        return cursor.rowcount > 0

    def get_insurance_rates(self) -> Dict[str, float]:
        """Get current insurance rates from settings"""
        return {
            "employment_insurance_rate": float(
                self.get_setting("employment_insurance_rate", "0.0090")
            ),
            "workers_comp_rate": float(self.get_setting("workers_comp_rate", "0.003")),
        }

    def get_ignored_companies(self) -> List[str]:
        """Get list of ignored/deactivated companies"""
        import json

        val = self.get_setting("ignored_companies", "[]")
        try:
            return json.loads(val)
        except (json.JSONDecodeError, TypeError):
            return []

    def set_company_active(self, company_name: str, active: bool) -> bool:
        """Set company active/inactive status"""
        import json

        ignored = self.get_ignored_companies()

        if active:
            if company_name in ignored:
                ignored.remove(company_name)
        else:
            if company_name not in ignored:
                ignored.append(company_name)

        return self.update_setting(
            "ignored_companies", json.dumps(ignored), "Ignored companies list"
        )

    # ============== Employee Operations ==============

    def get_employees(
        self,
        search: Optional[str] = None,
        company: Optional[str] = None,
        employee_type: Optional[str] = None,
    ) -> List[Dict]:
        """Get all employees with optional filtering by search, company, and employee_type"""
        cursor = self.db.cursor()

        # Use placeholder based on database type
        ph = "%s" if USE_POSTGRES else "?"

        query = """
            SELECT e.*,
                   (e.billing_rate - e.hourly_rate) as profit_per_hour,
                   CASE WHEN e.billing_rate > 0
                        THEN ((e.billing_rate - e.hourly_rate) / e.billing_rate * 100)
                        ELSE 0 END as margin_rate
            FROM employees e
            WHERE 1=1
        """
        params = []

        if search:
            query += f" AND (e.employee_id LIKE {ph} OR e.name LIKE {ph} OR e.name_kana LIKE {ph})"
            search_param = f"%{search}%"
            params.extend([search_param, search_param, search_param])

        if company:
            query += f" AND e.dispatch_company = {ph}"
            params.append(company)

        if employee_type:
            query += f" AND e.employee_type = {ph}"
            params.append(employee_type)

        query += " ORDER BY e.employee_id"

        cursor.execute(query, params)
        return [dict(row) for row in cursor.fetchall()]

    def get_employee(self, employee_id: str) -> Optional[Dict]:
        """Get a single employee by ID"""
        cursor = self.db.cursor()
        cursor.execute(
            _q("""
            SELECT e.*,
                   (e.billing_rate - e.hourly_rate) as profit_per_hour,
                   CASE WHEN e.billing_rate > 0
                        THEN ((e.billing_rate - e.hourly_rate) / e.billing_rate * 100)
                        ELSE 0 END as margin_rate
            FROM employees e
            WHERE e.employee_id = ?
        """),
            (employee_id,),
        )
        row = cursor.fetchone()
        return dict(row) if row else None

    def create_employee(self, employee: EmployeeCreate) -> Dict:
        """Create a new employee"""
        cursor = self.db.cursor()
        cursor.execute(
            _q("""
            INSERT INTO employees (employee_id, name, name_kana, dispatch_company, department,
                                   hourly_rate, billing_rate, status, hire_date,
                                   employee_type, gender, birth_date, termination_date, nationality)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """),
            (
                employee.employee_id,
                employee.name,
                employee.name_kana,
                employee.dispatch_company,
                employee.department,
                employee.hourly_rate,
                employee.billing_rate,
                employee.status,
                employee.hire_date,
                getattr(employee, "employee_type", "haken"),
                getattr(employee, "gender", None),
                getattr(employee, "birth_date", None),
                getattr(employee, "termination_date", None),
                getattr(employee, "nationality", None),
            ),
        )
        self.db.commit()
        return self.get_employee(employee.employee_id)

    def update_employee(
        self, employee_id: str, employee: EmployeeCreate
    ) -> Optional[Dict]:
        """Update an existing employee"""
        cursor = self.db.cursor()
        cursor.execute(
            _q("""
            UPDATE employees
            SET name = ?, name_kana = ?, dispatch_company = ?, department = ?,
                hourly_rate = ?, billing_rate = ?, status = ?, hire_date = ?,
                employee_type = ?, gender = ?, birth_date = ?, termination_date = ?,
                nationality = ?, updated_at = CURRENT_TIMESTAMP
            WHERE employee_id = ?
        """),
            (
                employee.name,
                employee.name_kana,
                employee.dispatch_company,
                employee.department,
                employee.hourly_rate,
                employee.billing_rate,
                employee.status,
                employee.hire_date,
                getattr(employee, "employee_type", "haken"),
                getattr(employee, "gender", None),
                getattr(employee, "birth_date", None),
                getattr(employee, "termination_date", None),
                getattr(employee, "nationality", None),
                employee_id,
            ),
        )
        self.db.commit()
        return self.get_employee(employee_id) if cursor.rowcount > 0 else None

    def delete_employee(self, employee_id: str) -> bool:
        """Delete an employee"""
        cursor = self.db.cursor()
        cursor.execute(_q("DELETE FROM employees WHERE employee_id = ?"), (employee_id,))
        self.db.commit()
        return cursor.rowcount > 0

    # ============== Payroll Operations ==============

    def get_payroll_records(
        self, period: Optional[str] = None, employee_id: Optional[str] = None
    ) -> List[Dict]:
        """Get payroll records with optional filtering"""
        cursor = self.db.cursor()

        # Use placeholder based on database type
        ph = "%s" if USE_POSTGRES else "?"

        query = """
            SELECT p.*, e.name as employee_name, e.dispatch_company
            FROM payroll_records p
            LEFT JOIN employees e ON p.employee_id = e.employee_id
            WHERE 1=1
        """
        params = []

        if period:
            query += f" AND p.period = {ph}"
            params.append(period)

        if employee_id:
            query += f" AND p.employee_id = {ph}"
            params.append(employee_id)

        query += " ORDER BY p.period DESC, p.employee_id"

        cursor.execute(query, params)
        return [dict(row) for row in cursor.fetchall()]

    def get_available_periods(self) -> List[str]:
        """Get list of available periods"""
        cursor = self.db.cursor()
        cursor.execute(
            """
            SELECT DISTINCT period FROM payroll_records ORDER BY period DESC
        """
        )
        return [row["period"] for row in cursor.fetchall()]

    def get_payroll_by_employee_year(
        self, employee_id: str, year: int
    ) -> List[Dict]:
        """Get payroll records for a specific employee and year (for wage ledger export)"""
        cursor = self.db.cursor()

        # Filter periods that match the year (e.g., "2025年1月", "2025年2月", etc.)
        year_pattern = f"{year}年%"

        cursor.execute(
            _q("""
            SELECT p.*, e.name as employee_name, e.dispatch_company
            FROM payroll_records p
            LEFT JOIN employees e ON p.employee_id = e.employee_id
            WHERE p.employee_id = ? AND p.period LIKE ?
            ORDER BY p.period ASC
        """),
            (employee_id, year_pattern),
        )
        return [dict(row) for row in cursor.fetchall()]

    # Insurance rates (2025年度) - configurable constants
    EMPLOYMENT_INSURANCE_RATE = 0.0090  # 雇用保険（会社負担）0.90% ← 2025年度
    WORKERS_COMP_RATE = 0.003  # 労災保険 0.3% (派遣業の場合、業種により0.25%~0.88%)

    # Billing multipliers (for factory billing)
    # These are what the factory pays us, NOT what we pay the employee
    BILLING_MULTIPLIERS = {
        "overtime_normal": 1.25,  # 残業 ≤60h: ×1.25
        "overtime_over_60h": 1.5,  # 残業 >60h: ×1.5
        "night": 0.25,  # 深夜: +0.25 (extra on top of regular or overtime)
        "holiday": 1.35,  # 休日: ×1.35
    }

    def calculate_billing_amount(
        self, record: PayrollRecordCreate, employee: Dict
    ) -> float:
        """
        Calculate billing amount based on hours and employee's 単価 (billing_rate)

        Formula:
        - 基本時間: 単価 × work_hours
        - 残業 (≤60h): 単価 × overtime_hours × 1.25
        - 残業 (>60h): 単価 × overtime_over_60h × 1.5
        - 深夜: 単価 × night_hours × 0.25 (extra on top)
        - 休日: 単価 × holiday_hours × 1.35

        Args:
            record: PayrollRecordCreate with hours data
            employee: Employee dict with billing_rate

        Returns:
            Calculated billing amount
        """
        tanka = employee.get("billing_rate", 0)
        if tanka <= 0:
            return 0.0

        # Get hours from record
        work_hours = getattr(record, "work_hours", 0) or 0
        overtime_hours = getattr(record, "overtime_hours", 0) or 0
        overtime_over_60h = getattr(record, "overtime_over_60h", 0) or 0
        night_hours = getattr(record, "night_hours", 0) or 0
        holiday_hours = getattr(record, "holiday_hours", 0) or 0

        # Calculate billing components
        # 基本時間 (normal hours)
        base_billing = work_hours * tanka

        # 残業 ≤60h: ×1.25
        overtime_billing = (
            overtime_hours * tanka * self.BILLING_MULTIPLIERS["overtime_normal"]
        )

        # 残業 >60h: ×1.5 (60H過残業)
        overtime_over_60h_billing = (
            overtime_over_60h * tanka * self.BILLING_MULTIPLIERS["overtime_over_60h"]
        )

        # 深夜: +0.25 extra (factory pays extra for night work)
        night_billing = night_hours * tanka * self.BILLING_MULTIPLIERS["night"]

        # 休日: ×1.35
        holiday_billing = holiday_hours * tanka * self.BILLING_MULTIPLIERS["holiday"]

        # Other Allowances (皆勤手当, 深夜残業, etc.) are passed through to billing
        # EXCLUDING: Transport and Non-Billable (業務手当) which are in separate fields
        other_allowances_billing = getattr(record, "other_allowances", 0) or 0

        total_billing = (
            base_billing
            + overtime_billing
            + overtime_over_60h_billing
            + night_billing
            + holiday_billing
            + other_allowances_billing
        )

        return round(total_billing)

    def create_payroll_record(self, record: PayrollRecordCreate) -> Dict:
        """Create a new payroll record with calculated fields"""
        # Get employee info for calculations
        employee = self.get_employee(record.employee_id)
        if not employee:
            raise ValueError(f"Employee {record.employee_id} not found")

        hourly_rate = employee["hourly_rate"]
        billing_rate = employee["billing_rate"]

        # Get new hour fields with defaults
        night_hours = getattr(record, "night_hours", 0) or 0
        holiday_hours = getattr(record, "holiday_hours", 0) or 0
        overtime_over_60h = getattr(record, "overtime_over_60h", 0) or 0

        # Get new pay fields with defaults
        night_pay = getattr(record, "night_pay", 0) or 0
        holiday_pay = getattr(record, "holiday_pay", 0) or 0
        overtime_over_60h_pay = getattr(record, "overtime_over_60h_pay", 0) or 0

        # Calculate billing_amount if not provided or is 0
        billing_amount = record.billing_amount
        if billing_amount <= 0 and billing_rate > 0:
            billing_amount = self.calculate_billing_amount(record, employee)

        # Calculate company costs
        # Get insurance rates from settings (dynamic)
        rates = self.get_insurance_rates()

        # 社会保険（会社負担）= 本人負担と同額 (労使折半)
        # NOTE: 社会保険 = 健康保険 + 厚生年金 (both employer and employee pay equal amounts)
        welfare_pension = getattr(record, "welfare_pension", 0) or 0
        company_social_insurance = record.company_social_insurance or (
            record.social_insurance + welfare_pension
        )

        # 雇用保険（会社負担）- Rate from settings (2025年度: 0.90%)
        company_employment_insurance = record.company_employment_insurance or round(
            record.gross_salary * rates["employment_insurance_rate"]
        )

        # 労災保険（会社負担100%）- Rate from settings (製造業: 0.3%)
        company_workers_comp = getattr(record, "company_workers_comp", None) or round(
            record.gross_salary * rates["workers_comp_rate"]
        )

        # ================================================================
        # 有給コスト (Paid Leave Cost) Calculation
        # ================================================================
        # IMPORTANT: 有給支給 (paid_leave_amount) is ALREADY INCLUDED in gross_salary
        # when it appears in the Excel 支給の部 section.
        #
        # We store paid_leave_amount for display purposes, but do NOT add it
        # again to total_company_cost (that would be double-counting).
        #
        # The formula is:
        #   会社総コスト = 総支給額 + 法定福利費(会社負担)
        #
        # Where 法定福利費(会社負担) = 健康保険(会社) + 厚生年金(会社) + 雇用保険 + 労災保険
        # ================================================================
        paid_leave_amount = getattr(record, "paid_leave_amount", 0) or 0

        # NOTE: paid_leave_amount is already in gross_salary - DO NOT add again
        # NOTE: Smart check for Transport Allowance exclusion
        # In some Excel formats, Transport Allowance (14,002) is NOT included in Gross Pay (325,587),
        # even though the UI might say so. We need to check if we need to add it to Company Cost.

        # Calculate sum of known components
        base_salary = getattr(record, "base_salary", 0) or 0
        overtime_pay = getattr(record, "overtime_pay", 0) or 0
        night_pay = getattr(record, "night_pay", 0) or 0
        holiday_pay = getattr(record, "holiday_pay", 0) or 0
        overtime_over_60h_pay = getattr(record, "overtime_over_60h_pay", 0) or 0
        paid_leave_amount = getattr(record, "paid_leave_amount", 0) or 0
        other_allowances = getattr(record, "other_allowances", 0) or 0
        transport_allowance = getattr(record, "transport_allowance", 0) or 0
        non_billable_allowances = getattr(record, "non_billable_allowances", 0) or 0

        # Sum WITHOUT Transport
        sum_without_transport = (
            base_salary
            + overtime_pay
            + night_pay
            + holiday_pay
            + overtime_over_60h_pay
            + paid_leave_amount
            + other_allowances
            + non_billable_allowances
        )

        # Check if Gross matches Sum WITHOUT Transport (tolerance for rounding)
        # If Gross is close to Sum Without Transport, then Transport is NOT in Gross.
        # We must ADD it to Company Cost.
        is_transport_excluded = abs(record.gross_salary - sum_without_transport) <= 100

        transport_cost_adder = (
            transport_allowance
            if (is_transport_excluded and transport_allowance > 0)
            else 0
        )

        total_company_cost = record.total_company_cost or round(
            record.gross_salary
            + company_social_insurance
            + company_employment_insurance
            + company_workers_comp
            + transport_cost_adder
        )

        # Calculate profit
        gross_profit = record.gross_profit or round(billing_amount - total_company_cost)
        profit_margin = record.profit_margin or (
            round((gross_profit / billing_amount * 100), 1) if billing_amount > 0 else 0
        )

        cursor = self.db.cursor()

        # Get non_billable_allowances from record
        non_billable_allowances = getattr(record, "non_billable_allowances", 0) or 0

        # Build the values tuple for the INSERT
        values = (
            record.employee_id,
            record.period,
            record.work_days,
            record.work_hours,
            record.overtime_hours,
            night_hours,
            holiday_hours,
            overtime_over_60h,
            record.paid_leave_hours,
            record.paid_leave_days,
            paid_leave_amount,
            record.base_salary,
            record.overtime_pay,
            night_pay,
            holiday_pay,
            overtime_over_60h_pay,
            record.transport_allowance,
            record.other_allowances,
            non_billable_allowances,
            record.gross_salary,
            record.social_insurance,
            welfare_pension,
            record.employment_insurance,
            record.income_tax,
            record.resident_tax,
            record.rent_deduction,
            record.utilities_deduction,
            record.meal_deduction,
            record.advance_payment,
            record.year_end_adjustment,
            record.other_deductions,
            record.net_salary,
            billing_amount,
            company_social_insurance,
            company_employment_insurance,
            company_workers_comp,
            total_company_cost,
            gross_profit,
            profit_margin,
        )

        # Use UPSERT pattern - compatible with both SQLite and PostgreSQL
        if USE_POSTGRES:
            # PostgreSQL: INSERT ... ON CONFLICT DO UPDATE
            cursor.execute(
                """
                INSERT INTO payroll_records (
                    employee_id, period, work_days, work_hours, overtime_hours,
                    night_hours, holiday_hours, overtime_over_60h,
                    paid_leave_hours, paid_leave_days, paid_leave_amount,
                    base_salary, overtime_pay, night_pay, holiday_pay, overtime_over_60h_pay,
                    transport_allowance, other_allowances, non_billable_allowances, gross_salary,
                    social_insurance, welfare_pension, employment_insurance, income_tax, resident_tax,
                    rent_deduction, utilities_deduction, meal_deduction, advance_payment, year_end_adjustment,
                    other_deductions, net_salary, billing_amount, company_social_insurance,
                    company_employment_insurance, company_workers_comp, total_company_cost, gross_profit, profit_margin
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (employee_id, period) DO UPDATE SET
                    work_days = EXCLUDED.work_days,
                    work_hours = EXCLUDED.work_hours,
                    overtime_hours = EXCLUDED.overtime_hours,
                    night_hours = EXCLUDED.night_hours,
                    holiday_hours = EXCLUDED.holiday_hours,
                    overtime_over_60h = EXCLUDED.overtime_over_60h,
                    paid_leave_hours = EXCLUDED.paid_leave_hours,
                    paid_leave_days = EXCLUDED.paid_leave_days,
                    paid_leave_amount = EXCLUDED.paid_leave_amount,
                    base_salary = EXCLUDED.base_salary,
                    overtime_pay = EXCLUDED.overtime_pay,
                    night_pay = EXCLUDED.night_pay,
                    holiday_pay = EXCLUDED.holiday_pay,
                    overtime_over_60h_pay = EXCLUDED.overtime_over_60h_pay,
                    transport_allowance = EXCLUDED.transport_allowance,
                    other_allowances = EXCLUDED.other_allowances,
                    non_billable_allowances = EXCLUDED.non_billable_allowances,
                    gross_salary = EXCLUDED.gross_salary,
                    social_insurance = EXCLUDED.social_insurance,
                    welfare_pension = EXCLUDED.welfare_pension,
                    employment_insurance = EXCLUDED.employment_insurance,
                    income_tax = EXCLUDED.income_tax,
                    resident_tax = EXCLUDED.resident_tax,
                    rent_deduction = EXCLUDED.rent_deduction,
                    utilities_deduction = EXCLUDED.utilities_deduction,
                    meal_deduction = EXCLUDED.meal_deduction,
                    advance_payment = EXCLUDED.advance_payment,
                    year_end_adjustment = EXCLUDED.year_end_adjustment,
                    other_deductions = EXCLUDED.other_deductions,
                    net_salary = EXCLUDED.net_salary,
                    billing_amount = EXCLUDED.billing_amount,
                    company_social_insurance = EXCLUDED.company_social_insurance,
                    company_employment_insurance = EXCLUDED.company_employment_insurance,
                    company_workers_comp = EXCLUDED.company_workers_comp,
                    total_company_cost = EXCLUDED.total_company_cost,
                    gross_profit = EXCLUDED.gross_profit,
                    profit_margin = EXCLUDED.profit_margin
            """,
                values,
            )
        else:
            # SQLite: INSERT OR REPLACE
            cursor.execute(
                """
                INSERT OR REPLACE INTO payroll_records (
                    employee_id, period, work_days, work_hours, overtime_hours,
                    night_hours, holiday_hours, overtime_over_60h,
                    paid_leave_hours, paid_leave_days, paid_leave_amount,
                    base_salary, overtime_pay, night_pay, holiday_pay, overtime_over_60h_pay,
                    transport_allowance, other_allowances, non_billable_allowances, gross_salary,
                    social_insurance, welfare_pension, employment_insurance, income_tax, resident_tax,
                    rent_deduction, utilities_deduction, meal_deduction, advance_payment, year_end_adjustment,
                    other_deductions, net_salary, billing_amount, company_social_insurance,
                    company_employment_insurance, company_workers_comp, total_company_cost, gross_profit, profit_margin
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
                values,
            )

        # NOTE: Commit is handled by the calling endpoint to allow transactions
        # self.db.commit()  # Removed - caller must commit

        # Return the created record as a dictionary
        cursor.execute(
            _q("""
            SELECT p.*, e.name as employee_name, e.dispatch_company
            FROM payroll_records p
            LEFT JOIN employees e ON p.employee_id = e.employee_id
            WHERE p.employee_id = ? AND p.period = ?
        """),
            (record.employee_id, record.period),
        )

        row = cursor.fetchone()

        # Helper function to convert Row to dict safely
        if row:
            result = dict(row)
            # Ensure we return floats for numeric values to match test expectations
            numeric_fields = [
                "company_social_insurance",
                "company_employment_insurance",
                "total_company_cost",
                "billing_amount",
            ]
            for field in numeric_fields:
                if field in result and result[field] is not None:
                    result[field] = float(result[field])
            return result

        return {}

    # ============== Statistics ==============

    def get_statistics(self, period: Optional[str] = None) -> Dict:
        """Get dashboard statistics"""
        cursor = self.db.cursor()

        # Get ignored companies list (works with both SQLite and PostgreSQL)
        ignored_companies = _get_ignored_companies(cursor)
        company_filter, company_params = _build_company_filter(ignored_companies)

        # Get latest period if not specified
        if not period:
            cursor.execute("SELECT MAX(period) FROM payroll_records")
            result = cursor.fetchone()
            period = _get_first_col(result)

        if not period:
            return self._empty_statistics()

        # Basic counts
        cursor.execute("SELECT COUNT(*) FROM employees")
        total_employees = _get_count(cursor.fetchone())

        cursor.execute("SELECT COUNT(*) FROM employees WHERE status = 'active'")
        active_employees = _get_count(cursor.fetchone())

        # Count companies (excluding ignored)
        if ignored_companies:
            company_filter_direct, company_params_direct = _build_company_filter(
                ignored_companies, "dispatch_company"
            )
            cursor.execute(
                _q(f"SELECT COUNT(DISTINCT dispatch_company) FROM employees WHERE 1=1 {company_filter_direct}"),
                company_params_direct
            )
        else:
            cursor.execute("SELECT COUNT(DISTINCT dispatch_company) FROM employees")
        total_companies = _get_count(cursor.fetchone())

        # Period statistics
        cursor.execute(
            _q(f"""
            SELECT
                AVG(gross_profit) as average_profit,
                AVG(profit_margin) as average_margin,
                SUM(billing_amount) as total_revenue,
                SUM(total_company_cost) as total_cost,
                SUM(gross_profit) as total_profit
            FROM payroll_records p
            LEFT JOIN employees e ON p.employee_id = e.employee_id
            WHERE p.period = ?
            {company_filter}
        """),
            (period,) + company_params,
        )
        stats = cursor.fetchone()

        # Profit trend (last 6 periods) - Sort properly handling Japanese period format
        if USE_POSTGRES:
            # PostgreSQL version with compatible string functions
            order_clause = """
                ORDER BY
                    CAST(SUBSTRING(period FROM 1 FOR 4) AS INTEGER) DESC,
                    CAST(REGEXP_REPLACE(SUBSTRING(period FROM 6), '[^0-9]', '', 'g') AS INTEGER) DESC
            """
        else:
            # SQLite version
            order_clause = """
                ORDER BY
                    CAST(SUBSTR(period, 1, 4) AS INTEGER) DESC,
                    CAST(REPLACE(REPLACE(SUBSTR(period, 6), '月', ''), '年', '') AS INTEGER) DESC
            """

        if company_filter:
            cursor.execute(
                _q(f"""
                SELECT
                    period,
                    SUM(billing_amount) as revenue,
                    SUM(total_company_cost) as cost,
                    SUM(gross_profit) as profit,
                    AVG(profit_margin) as margin
                FROM payroll_records p
                LEFT JOIN employees e ON p.employee_id = e.employee_id
                WHERE 1=1 {company_filter}
                GROUP BY period
                {order_clause}
                LIMIT 6
            """),
                company_params
            )
        else:
            cursor.execute(f"""
                SELECT
                    period,
                    SUM(billing_amount) as revenue,
                    SUM(total_company_cost) as cost,
                    SUM(gross_profit) as profit,
                    AVG(profit_margin) as margin
                FROM payroll_records p
                LEFT JOIN employees e ON p.employee_id = e.employee_id
                GROUP BY period
                {order_clause}
                LIMIT 6
            """)
        profit_trend = [dict(row) for row in cursor.fetchall()][
            ::-1
        ]  # Reverse for chronological order

        # Profit distribution
        profit_distribution = self._calculate_profit_distribution(period)

        # Top companies
        cursor.execute(
            _q(f"""
            SELECT
                e.dispatch_company as company_name,
                COUNT(DISTINCT e.employee_id) as employee_count,
                AVG(e.hourly_rate) as average_hourly_rate,
                AVG(e.billing_rate) as average_billing_rate,
                AVG(e.billing_rate - e.hourly_rate) as average_profit,
                AVG((e.billing_rate - e.hourly_rate) / NULLIF(e.billing_rate, 0) * 100) as average_margin,
                SUM(p.gross_profit) as total_monthly_profit
            FROM employees e
            LEFT JOIN payroll_records p ON e.employee_id = p.employee_id AND p.period = ?
            WHERE 1=1 {company_filter}
            GROUP BY e.dispatch_company
            ORDER BY total_monthly_profit DESC NULLS LAST
            LIMIT 5
        """),
            (period,) + company_params,
        )
        top_companies = [dict(row) for row in cursor.fetchall()]

        # Recent payrolls
        cursor.execute(
            _q(f"""
            SELECT p.*, e.name as employee_name, e.dispatch_company
            FROM payroll_records p
            LEFT JOIN employees e ON p.employee_id = e.employee_id
            WHERE p.period = ?
            {company_filter}
            ORDER BY p.gross_profit DESC
            LIMIT 10
        """),
            (period,) + company_params,
        )
        recent_payrolls = [dict(row) for row in cursor.fetchall()]

        return {
            "total_employees": total_employees,
            "active_employees": active_employees,
            "total_companies": total_companies,
            "average_profit": stats["average_profit"] or 0 if stats else 0,
            "average_margin": stats["average_margin"] or 0 if stats else 0,
            "total_monthly_revenue": stats["total_revenue"] or 0 if stats else 0,
            "total_monthly_cost": stats["total_cost"] or 0 if stats else 0,
            "total_monthly_profit": stats["total_profit"] or 0 if stats else 0,
            "profit_trend": profit_trend,
            "profit_distribution": profit_distribution,
            "top_companies": top_companies,
            "recent_payrolls": recent_payrolls,
            "current_period": period,
        }

    def _calculate_profit_distribution(self, period: str) -> List[Dict]:
        """Calculate profit distribution for a period

        Ranges are based on 製造派遣 target margin of 12%:
        - <7%: Critical (赤字リスク)
        - 7-10%: Below target (要改善)
        - 10-12%: Close to target (良好)
        - ≥12%: Target achieved (目標達成)
        """
        cursor = self.db.cursor()

        # Get ignored companies list (works with both SQLite and PostgreSQL)
        ignored_companies = _get_ignored_companies(cursor)
        company_filter, company_params = _build_company_filter(ignored_companies)

        ranges = [
            ("<7%", -999999999, 7),   # Critical
            ("7-10%", 7, 10),         # Needs improvement
            ("10-12%", 10, 12),       # Good - close to target
            ("≥12%", 12, 999999999),  # Target achieved
        ]

        cursor.execute(
            _q(f"""
            SELECT COUNT(*)
            FROM payroll_records p
            LEFT JOIN employees e ON p.employee_id = e.employee_id
            WHERE p.period = ?
            {company_filter}
        """),
            (period,) + company_params,
        )
        total = _get_count(cursor.fetchone())

        distribution = []
        for range_name, min_val, max_val in ranges:
            cursor.execute(
                _q(f"""
                SELECT COUNT(*)
                FROM payroll_records p
                LEFT JOIN employees e ON p.employee_id = e.employee_id
                WHERE p.period = ?
                AND p.profit_margin >= ?
                AND p.profit_margin < ?
                {company_filter}
            """),
                (period, min_val, max_val) + company_params,
            )
            count = _get_count(cursor.fetchone())
            percentage = (count / total * 100) if total > 0 else 0
            distribution.append(
                {
                    "range": range_name,
                    "count": count,
                    "percentage": round(percentage, 1),
                }
            )

        return distribution

    def _empty_statistics(self) -> Dict:
        """Return empty statistics"""
        return {
            "total_employees": 0,
            "active_employees": 0,
            "total_companies": 0,
            "average_profit": 0,
            "average_margin": 0,
            "total_monthly_revenue": 0,
            "total_monthly_cost": 0,
            "total_monthly_profit": 0,
            "profit_trend": [],
            "profit_distribution": [],
            "top_companies": [],
            "recent_payrolls": [],
            "current_period": None,
        }

    def get_monthly_statistics(
        self, year: Optional[int] = None, month: Optional[int] = None
    ) -> List[Dict]:
        """Get monthly statistics"""
        cursor = self.db.cursor()

        query = """
            SELECT
                period,
                COUNT(DISTINCT employee_id) as total_employees,
                SUM(billing_amount) as total_revenue,
                SUM(total_company_cost) as total_cost,
                SUM(gross_profit) as total_profit,
                AVG(profit_margin) as average_margin,
                SUM(company_social_insurance) as total_social_insurance,
                SUM(paid_leave_hours * (SELECT hourly_rate FROM employees WHERE employees.employee_id = payroll_records.employee_id)) as total_paid_leave_cost
            FROM payroll_records
        """

        params = []
        if year and month:
            period = f"{year}年{month}月"
            query += " WHERE period = ?"
            params.append(period)

        query += " GROUP BY period ORDER BY period DESC"

        cursor.execute(query, params)
        return [dict(row) for row in cursor.fetchall()]

    def get_company_statistics(self, period: str = None) -> List[Dict]:
        """Get statistics by company, including additional costs"""
        cursor = self.db.cursor()

        # Get latest period if not specified
        if not period:
            cursor.execute("SELECT MAX(period) FROM payroll_records")
            period = _get_first_col(cursor.fetchone())

        cursor.execute(
            _q("""
            SELECT
                e.dispatch_company as company_name,
                COUNT(DISTINCT e.employee_id) as employee_count,
                AVG(e.hourly_rate) as average_hourly_rate,
                AVG(e.billing_rate) as average_billing_rate,
                AVG(e.billing_rate - e.hourly_rate) as average_profit,
                AVG((e.billing_rate - e.hourly_rate) / NULLIF(e.billing_rate, 0) * 100) as average_margin,
                COALESCE(SUM(p.gross_profit), 0) as total_monthly_profit,
                COALESCE(SUM(p.billing_amount), 0) as total_monthly_revenue
            FROM employees e
            LEFT JOIN payroll_records p ON e.employee_id = p.employee_id AND p.period = ?
            GROUP BY e.dispatch_company
            ORDER BY total_monthly_profit DESC
        """),
            (period,),
        )

        result = [dict(row) for row in cursor.fetchall()]

        # Fetch all additional costs for the period in ONE query (fix N+1 pattern)
        cursor.execute(
            _q("""
            SELECT dispatch_company, COALESCE(SUM(amount), 0) as additional_costs
            FROM company_additional_costs
            WHERE period = ?
            GROUP BY dispatch_company
        """),
            (period,),
        )
        additional_costs_map = {
            row["dispatch_company"]: float(row["additional_costs"])
            for row in cursor.fetchall()
        }

        # Add additional costs and adjusted profit for each company (no more N+1)
        for c in result:
            company_name = c["company_name"]
            additional_costs = additional_costs_map.get(company_name, 0.0)

            c["additional_costs"] = additional_costs
            c["adjusted_profit"] = float(c["total_monthly_profit"]) - additional_costs

            # Calculate adjusted margin
            revenue = float(c["total_monthly_revenue"] or 0)
            if revenue > 0:
                c["adjusted_margin"] = (c["adjusted_profit"] / revenue) * 100
            else:
                c["adjusted_margin"] = 0.0

        # Add 'is_active' flag
        ignored = self.get_ignored_companies()
        for c in result:
            c["is_active"] = c["company_name"] not in ignored

        # Re-sort by adjusted profit
        result.sort(key=lambda x: x["adjusted_profit"], reverse=True)

        return result

    def get_profit_trend(self, months: int = 6) -> List[Dict]:
        """Get profit trend for last N months"""
        cursor = self.db.cursor()
        cursor.execute(
            """
            SELECT
                period,
                SUM(billing_amount) as revenue,
                SUM(total_company_cost) as cost,
                SUM(gross_profit) as profit,
                AVG(profit_margin) as margin
            FROM payroll_records
            GROUP BY period
            ORDER BY period DESC
            LIMIT ?
        """,
            (months,),
        )
        return [dict(row) for row in cursor.fetchall()][::-1]


class ExcelParser:
    """Parser for Excel and CSV payroll files"""

    # Column mappings (Japanese -> internal)
    COLUMN_MAPPINGS = {
        "社員番号": "employee_id",
        "従業員番号": "employee_id",
        "ID": "employee_id",
        "対象期間": "period",
        "期間": "period",
        "月": "period",
        "出勤日数": "work_days",
        "労働時間": "work_hours",
        "勤務時間": "work_hours",
        "残業時間": "overtime_hours",
        "深夜時間": "night_hours",
        "深夜": "night_hours",
        "休日時間": "holiday_hours",
        "休日": "holiday_hours",
        "60H過残業": "overtime_over_60h",
        "60時間超": "overtime_over_60h",
        "有給時間": "paid_leave_hours",
        "有給日数": "paid_leave_days",
        "有給金額": "paid_leave_amount",
        "有給支給額": "paid_leave_amount",
        "有給手当": "paid_leave_amount",
        "有休金額": "paid_leave_amount",
        "有休支給": "paid_leave_amount",
        "基本給": "base_salary",
        "残業代": "overtime_pay",
        "残業手当": "overtime_pay",
        "深夜手当": "night_pay",
        "休日手当": "holiday_pay",
        "60H過手当": "overtime_over_60h_pay",
        "通勤費": "transport_allowance",
        "交通費": "transport_allowance",
        "その他手当": "other_allowances",
        "総支給額": "gross_salary",
        "社会保険料": "social_insurance",
        "社会保険": "social_insurance",
        "雇用保険料": "employment_insurance",
        "雇用保険": "employment_insurance",
        "所得税": "income_tax",
        "住民税": "resident_tax",
        "その他控除": "other_deductions",
        "差引支給額": "net_salary",
        "手取り": "net_salary",
        "請求金額": "billing_amount",
        "請求額": "billing_amount",
        "売上": "billing_amount",
    }

    def parse(self, content: bytes, file_ext: str) -> List[PayrollRecordCreate]:
        """Parse file content and return list of PayrollRecordCreate objects"""
        if file_ext == ".csv":
            return self._parse_csv(content)
        elif file_ext in [".xlsx", ".xlsm", ".xls"]:
            return self._parse_excel(content)
        else:
            raise ValueError(f"Unsupported file type: {file_ext}")

    def _parse_csv(self, content: bytes) -> List[PayrollRecordCreate]:
        """Parse CSV content"""
        records = []

        # Try different encodings
        for encoding in ["utf-8", "shift_jis", "cp932"]:
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("Could not decode file with supported encodings")

        reader = csv.DictReader(io.StringIO(text))

        for row in reader:
            record = self._map_row_to_record(row)
            if record:
                records.append(record)

        return records

    def _parse_excel(self, content: bytes) -> List[PayrollRecordCreate]:
        """Parse Excel content"""
        try:
            from io import BytesIO

            import openpyxl

            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
            ws = wb.active

            # Get header row
            headers = [cell.value for cell in ws[1]]
            records = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                record = self._map_row_to_record(row_dict)
                if record:
                    records.append(record)

            return records

        except ImportError:
            raise ValueError(
                "openpyxl is required to parse Excel files. Install with: pip install openpyxl"
            )

    def _map_row_to_record(self, row: dict) -> Optional[PayrollRecordCreate]:
        """Map a row dictionary to PayrollRecordCreate"""
        mapped = {}

        for jp_col, value in row.items():
            if jp_col and value is not None:
                internal_col = self.COLUMN_MAPPINGS.get(str(jp_col).strip())
                if internal_col:
                    mapped[internal_col] = self._convert_value(value, internal_col)

        # Validate required fields
        if not mapped.get("employee_id") or not mapped.get("period"):
            return None

        # Filter out invalid employee IDs (0, 000000)
        if (
            hasattr(mapped.get("employee_id"), "isdigit")
            and mapped["employee_id"].isdigit()
            and int(mapped["employee_id"]) == 0
        ):
            return None

        # Set defaults for missing fields
        defaults = {
            "work_days": 0,
            "work_hours": 0,
            "overtime_hours": 0,
            "night_hours": 0,
            "holiday_hours": 0,
            "overtime_over_60h": 0,
            "paid_leave_hours": 0,
            "paid_leave_days": 0,
            "paid_leave_amount": 0,
            "base_salary": 0,
            "overtime_pay": 0,
            "night_pay": 0,
            "holiday_pay": 0,
            "overtime_over_60h_pay": 0,
            "transport_allowance": 0,
            "other_allowances": 0,
            "gross_salary": 0,
            "social_insurance": 0,
            "employment_insurance": 0,
            "income_tax": 0,
            "resident_tax": 0,
            "rent_deduction": 0,
            "utilities_deduction": 0,
            "meal_deduction": 0,
            "advance_payment": 0,
            "year_end_adjustment": 0,
            "other_deductions": 0,
            "net_salary": 0,
            "billing_amount": 0,
        }

        for key, default_value in defaults.items():
            if key not in mapped:
                mapped[key] = default_value

        return PayrollRecordCreate(**mapped)

    def _convert_value(self, value: Any, field: str) -> Any:
        """Convert value to appropriate type"""
        if value is None:
            return None

        if field in ["employee_id", "period"]:
            return str(value).strip()

        # Numeric fields
        try:
            if isinstance(value, str):
                # Remove currency symbols and commas
                value = value.replace("¥", "").replace(",", "").replace(" ", "").strip()
            return float(value)
        except (ValueError, TypeError):
            return 0
