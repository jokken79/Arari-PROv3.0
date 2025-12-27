"""
ReportAgent - Report Generation System
Generates PDF and Excel reports for 粗利 PRO
"""

import json
import sqlite3
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List

from database import USE_POSTGRES

# Note: For PDF generation, you'll need to install: pip install reportlab
# For Excel: openpyxl is already installed
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# PDF generation with ReportLab
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.graphics.shapes import Drawing, String, Rect
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    # Register Japanese font
    pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
    REPORTLAB_AVAILABLE = True
    JAPANESE_FONT = 'HeiseiKakuGo-W5'
except ImportError:
    REPORTLAB_AVAILABLE = False
    JAPANESE_FONT = 'Helvetica'

# PDF Color scheme
PDF_COLORS = {
    'primary': colors.HexColor('#4472C4'),       # Blue (headers)
    'secondary': colors.HexColor('#6B7280'),     # Gray
    'success': colors.HexColor('#10B981'),       # Green (margin >= 15%)
    'warning': colors.HexColor('#F59E0B'),       # Yellow (margin 10-15%)
    'danger': colors.HexColor('#EF4444'),        # Red (margin < 10%)
    'light': colors.HexColor('#F3F4F6'),         # Light gray background
    'white': colors.white,
    'black': colors.black,
}


def _q(query: str) -> str:
    """Convert SQLite query to PostgreSQL if needed (? -> %s)"""
    if USE_POSTGRES:
        return query.replace("?", "%s")
    return query


def _get_row_value(row, key: str, index: int, default=None):
    """Extract value from row - handles both dict (PostgreSQL) and tuple (SQLite)"""
    if row is None:
        return default
    if isinstance(row, dict):
        return row.get(key, default)
    return row[index] if len(row) > index else default


def init_reports_tables(conn):
    """Initialize reports tables"""
    cursor = conn.cursor()

    # SQL type mappings for cross-database compatibility
    PK_TYPE = "SERIAL PRIMARY KEY" if USE_POSTGRES else "INTEGER PRIMARY KEY AUTOINCREMENT"

    # Generated reports table (for tracking/caching)
    cursor.execute(
        f"""
        CREATE TABLE IF NOT EXISTS generated_reports (
            id {PK_TYPE},
            report_type TEXT NOT NULL,
            report_name TEXT NOT NULL,
            period TEXT,
            entity_type TEXT,
            entity_id TEXT,
            format TEXT NOT NULL,
            file_path TEXT,
            file_size INTEGER,
            generated_by TEXT,
            parameters TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """
    )

    # Report templates table
    cursor.execute(
        f"""
        CREATE TABLE IF NOT EXISTS report_templates (
            id {PK_TYPE},
            template_name TEXT UNIQUE NOT NULL,
            report_type TEXT NOT NULL,
            description TEXT,
            default_params TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """
    )

    conn.commit()


class ReportService:
    """Service for generating reports"""

    def __init__(self, conn: sqlite3.Connection):
        self.conn = conn
        self.cursor = conn.cursor()

    def get_monthly_report_data(self, period: str) -> Dict[str, Any]:
        """Get data for monthly profit report"""

        # Summary statistics
        self.cursor.execute(
            _q("""
            SELECT
                COUNT(*) as employee_count,
                SUM(billing_amount) as total_revenue,
                SUM(total_company_cost) as total_cost,
                SUM(gross_profit) as total_profit,
                AVG(profit_margin) as avg_margin,
                SUM(work_hours) as total_work_hours,
                SUM(overtime_hours) as total_overtime,
                SUM(overtime_over_60h) as total_overtime_60h,
                SUM(night_hours) as total_night_hours,
                SUM(holiday_hours) as total_holiday_hours
            FROM payroll_records
            WHERE period = ?
        """),
            (period,),
        )

        row = self.cursor.fetchone()
        summary = {
            "employee_count": _get_row_value(row, "employee_count", 0, 0),
            "total_revenue": _get_row_value(row, "total_revenue", 1, 0),
            "total_cost": _get_row_value(row, "total_cost", 2, 0),
            "total_profit": _get_row_value(row, "total_profit", 3, 0),
            "avg_margin": _get_row_value(row, "avg_margin", 4, 0),
            "total_work_hours": _get_row_value(row, "total_work_hours", 5, 0),
            "total_overtime": _get_row_value(row, "total_overtime", 6, 0),
            "total_overtime_60h": _get_row_value(row, "total_overtime_60h", 7, 0),
            "total_night_hours": _get_row_value(row, "total_night_hours", 8, 0),
            "total_holiday_hours": _get_row_value(row, "total_holiday_hours", 9, 0),
        }

        # By company breakdown
        self.cursor.execute(
            _q("""
            SELECT
                e.dispatch_company,
                COUNT(*) as employee_count,
                SUM(p.billing_amount) as revenue,
                SUM(p.total_company_cost) as cost,
                SUM(p.gross_profit) as profit,
                AVG(p.profit_margin) as avg_margin
            FROM payroll_records p
            JOIN employees e ON p.employee_id = e.employee_id
            WHERE p.period = ?
            GROUP BY e.dispatch_company
            ORDER BY profit DESC
        """),
            (period,),
        )

        by_company = []
        for row in self.cursor.fetchall():
            by_company.append(
                {
                    "company": _get_row_value(row, "dispatch_company", 0),
                    "employee_count": _get_row_value(row, "employee_count", 1, 0),
                    "revenue": _get_row_value(row, "revenue", 2, 0),
                    "cost": _get_row_value(row, "cost", 3, 0),
                    "profit": _get_row_value(row, "profit", 4, 0),
                    "margin": _get_row_value(row, "avg_margin", 5, 0),
                }
            )

        # Top/Bottom performers
        self.cursor.execute(
            _q("""
            SELECT
                p.employee_id, e.name, e.dispatch_company,
                p.billing_amount, p.total_company_cost, p.gross_profit, p.profit_margin
            FROM payroll_records p
            JOIN employees e ON p.employee_id = e.employee_id
            WHERE p.period = ?
            ORDER BY p.profit_margin DESC
            LIMIT 5
        """),
            (period,),
        )

        top_performers = []
        for row in self.cursor.fetchall():
            top_performers.append(
                {
                    "employee_id": _get_row_value(row, "employee_id", 0),
                    "name": _get_row_value(row, "name", 1),
                    "company": _get_row_value(row, "dispatch_company", 2),
                    "revenue": _get_row_value(row, "billing_amount", 3, 0),
                    "cost": _get_row_value(row, "total_company_cost", 4, 0),
                    "profit": _get_row_value(row, "gross_profit", 5, 0),
                    "margin": _get_row_value(row, "profit_margin", 6, 0),
                }
            )

        self.cursor.execute(
            _q("""
            SELECT
                p.employee_id, e.name, e.dispatch_company,
                p.billing_amount, p.total_company_cost, p.gross_profit, p.profit_margin
            FROM payroll_records p
            JOIN employees e ON p.employee_id = e.employee_id
            WHERE p.period = ?
            ORDER BY p.profit_margin ASC
            LIMIT 5
        """),
            (period,),
        )

        bottom_performers = []
        for row in self.cursor.fetchall():
            bottom_performers.append(
                {
                    "employee_id": _get_row_value(row, "employee_id", 0),
                    "name": _get_row_value(row, "name", 1),
                    "company": _get_row_value(row, "dispatch_company", 2),
                    "revenue": _get_row_value(row, "billing_amount", 3, 0),
                    "cost": _get_row_value(row, "total_company_cost", 4, 0),
                    "profit": _get_row_value(row, "gross_profit", 5, 0),
                    "margin": _get_row_value(row, "profit_margin", 6, 0),
                }
            )

        return {
            "period": period,
            "generated_at": datetime.now().isoformat(),
            "summary": summary,
            "by_company": by_company,
            "top_performers": top_performers,
            "bottom_performers": bottom_performers,
        }

    def get_employee_report_data(
        self, employee_id: str, months: int = 6
    ) -> Dict[str, Any]:
        """Get data for employee detail report"""

        # Employee info
        self.cursor.execute(
            _q("""
            SELECT employee_id, name, name_kana, dispatch_company,
                   hourly_rate, billing_rate, status, hire_date
            FROM employees WHERE employee_id = ?
        """),
            (employee_id,),
        )

        emp_row = self.cursor.fetchone()
        if not emp_row:
            return {"error": "Employee not found"}

        employee = {
            "employee_id": _get_row_value(emp_row, "employee_id", 0),
            "name": _get_row_value(emp_row, "name", 1),
            "name_kana": _get_row_value(emp_row, "name_kana", 2),
            "company": _get_row_value(emp_row, "dispatch_company", 3),
            "hourly_rate": _get_row_value(emp_row, "hourly_rate", 4),
            "billing_rate": _get_row_value(emp_row, "billing_rate", 5),
            "status": _get_row_value(emp_row, "status", 6),
            "hire_date": _get_row_value(emp_row, "hire_date", 7),
        }

        # Historical payroll data
        self.cursor.execute(
            _q("""
            SELECT period, work_hours, overtime_hours, overtime_over_60h,
                   night_hours, holiday_hours, gross_salary, billing_amount,
                   total_company_cost, gross_profit, profit_margin,
                   paid_leave_days, paid_leave_amount
            FROM payroll_records
            WHERE employee_id = ?
            ORDER BY period DESC
            LIMIT ?
        """),
            (employee_id, months),
        )

        history = []
        for row in self.cursor.fetchall():
            history.append(
                {
                    "period": _get_row_value(row, "period", 0),
                    "work_hours": _get_row_value(row, "work_hours", 1, 0),
                    "overtime_hours": _get_row_value(row, "overtime_hours", 2, 0),
                    "overtime_over_60h": _get_row_value(row, "overtime_over_60h", 3, 0),
                    "night_hours": _get_row_value(row, "night_hours", 4, 0),
                    "holiday_hours": _get_row_value(row, "holiday_hours", 5, 0),
                    "gross_salary": _get_row_value(row, "gross_salary", 6, 0),
                    "billing_amount": _get_row_value(row, "billing_amount", 7, 0),
                    "cost": _get_row_value(row, "total_company_cost", 8, 0),
                    "profit": _get_row_value(row, "gross_profit", 9, 0),
                    "margin": _get_row_value(row, "profit_margin", 10, 0),
                    "paid_leave_days": _get_row_value(row, "paid_leave_days", 11, 0),
                    "paid_leave_amount": _get_row_value(row, "paid_leave_amount", 12, 0),
                }
            )

        # Calculate averages
        if history:
            avg_margin = sum(h["margin"] or 0 for h in history) / len(history)
            avg_profit = sum(h["profit"] or 0 for h in history) / len(history)
            total_profit = sum(h["profit"] or 0 for h in history)
        else:
            avg_margin = 0
            avg_profit = 0
            total_profit = 0

        return {
            "employee": employee,
            "history": history,
            "summary": {
                "periods_count": len(history),
                "avg_margin": avg_margin,
                "avg_profit": avg_profit,
                "total_profit": total_profit,
            },
            "generated_at": datetime.now().isoformat(),
        }

    def get_company_report_data(self, company: str) -> Dict[str, Any]:
        """Get data for company/client report"""

        # Employees in this company
        self.cursor.execute(
            _q("""
            SELECT employee_id, name, hourly_rate, billing_rate, status
            FROM employees
            WHERE dispatch_company = ?
        """),
            (company,),
        )

        employees = []
        for row in self.cursor.fetchall():
            employees.append(
                {
                    "employee_id": _get_row_value(row, "employee_id", 0),
                    "name": _get_row_value(row, "name", 1),
                    "hourly_rate": _get_row_value(row, "hourly_rate", 2, 0),
                    "billing_rate": _get_row_value(row, "billing_rate", 3, 0),
                    "status": _get_row_value(row, "status", 4),
                }
            )

        # Monthly aggregates
        self.cursor.execute(
            _q("""
            SELECT
                p.period,
                COUNT(*) as employee_count,
                SUM(p.billing_amount) as revenue,
                SUM(p.total_company_cost) as cost,
                SUM(p.gross_profit) as profit,
                AVG(p.profit_margin) as avg_margin
            FROM payroll_records p
            JOIN employees e ON p.employee_id = e.employee_id
            WHERE e.dispatch_company = ?
            GROUP BY p.period
            ORDER BY p.period DESC
            LIMIT 12
        """),
            (company,),
        )

        monthly_data = []
        for row in self.cursor.fetchall():
            monthly_data.append(
                {
                    "period": _get_row_value(row, "period", 0),
                    "employee_count": _get_row_value(row, "employee_count", 1, 0),
                    "revenue": _get_row_value(row, "revenue", 2, 0),
                    "cost": _get_row_value(row, "cost", 3, 0),
                    "profit": _get_row_value(row, "profit", 4, 0),
                    "margin": _get_row_value(row, "avg_margin", 5, 0),
                }
            )

        # Overall summary
        self.cursor.execute(
            _q("""
            SELECT
                SUM(p.billing_amount) as total_revenue,
                SUM(p.total_company_cost) as total_cost,
                SUM(p.gross_profit) as total_profit,
                AVG(p.profit_margin) as avg_margin
            FROM payroll_records p
            JOIN employees e ON p.employee_id = e.employee_id
            WHERE e.dispatch_company = ?
        """),
            (company,),
        )

        totals = self.cursor.fetchone()

        return {
            "company": company,
            "employees": employees,
            "employee_count": len(employees),
            "monthly_data": monthly_data,
            "totals": {
                "revenue": _get_row_value(totals, "total_revenue", 0, 0),
                "cost": _get_row_value(totals, "total_cost", 1, 0),
                "profit": _get_row_value(totals, "total_profit", 2, 0),
                "avg_margin": _get_row_value(totals, "avg_margin", 3, 0),
            },
            "generated_at": datetime.now().isoformat(),
        }

    def get_all_employees_report_data(self, period: str) -> Dict[str, Any]:
        """Get report data for ALL employees in a period"""
        self.cursor.execute(
            _q("""
            SELECT
                p.employee_id, e.name, e.name_kana, e.dispatch_company,
                e.hourly_rate, e.billing_rate,
                p.work_hours, p.overtime_hours, p.overtime_over_60h,
                p.night_hours, p.holiday_hours,
                p.gross_salary, p.billing_amount, p.total_company_cost,
                p.gross_profit, p.profit_margin
            FROM payroll_records p
            JOIN employees e ON p.employee_id = e.employee_id
            WHERE p.period = ?
            ORDER BY e.dispatch_company, e.name
        """),
            (period,),
        )

        employees = []
        for row in self.cursor.fetchall():
            employees.append({
                "employee_id": _get_row_value(row, "employee_id", 0),
                "name": _get_row_value(row, "name", 1),
                "name_kana": _get_row_value(row, "name_kana", 2),
                "company": _get_row_value(row, "dispatch_company", 3),
                "hourly_rate": _get_row_value(row, "hourly_rate", 4, 0),
                "billing_rate": _get_row_value(row, "billing_rate", 5, 0),
                "work_hours": _get_row_value(row, "work_hours", 6, 0),
                "overtime_hours": _get_row_value(row, "overtime_hours", 7, 0),
                "overtime_over_60h": _get_row_value(row, "overtime_over_60h", 8, 0),
                "night_hours": _get_row_value(row, "night_hours", 9, 0),
                "holiday_hours": _get_row_value(row, "holiday_hours", 10, 0),
                "gross_salary": _get_row_value(row, "gross_salary", 11, 0),
                "billing_amount": _get_row_value(row, "billing_amount", 12, 0),
                "total_cost": _get_row_value(row, "total_company_cost", 13, 0),
                "gross_profit": _get_row_value(row, "gross_profit", 14, 0),
                "profit_margin": _get_row_value(row, "profit_margin", 15, 0),
            })

        # Summary totals
        self.cursor.execute(
            _q("""
            SELECT
                COUNT(*) as employee_count,
                SUM(billing_amount) as total_revenue,
                SUM(total_company_cost) as total_cost,
                SUM(gross_profit) as total_profit,
                AVG(profit_margin) as avg_margin
            FROM payroll_records
            WHERE period = ?
        """),
            (period,),
        )
        totals_row = self.cursor.fetchone()

        return {
            "period": period,
            "employees": employees,
            "totals": {
                "employee_count": _get_row_value(totals_row, "employee_count", 0, 0),
                "total_revenue": _get_row_value(totals_row, "total_revenue", 1, 0),
                "total_cost": _get_row_value(totals_row, "total_cost", 2, 0),
                "total_profit": _get_row_value(totals_row, "total_profit", 3, 0),
                "avg_margin": _get_row_value(totals_row, "avg_margin", 4, 0),
            },
            "generated_at": datetime.now().isoformat(),
        }

    def get_all_companies_report_data(self, period: str) -> Dict[str, Any]:
        """Get report data for ALL companies in a period"""
        self.cursor.execute(
            _q("""
            SELECT
                e.dispatch_company,
                COUNT(*) as employee_count,
                SUM(p.work_hours) as total_work_hours,
                SUM(p.billing_amount) as revenue,
                SUM(p.total_company_cost) as cost,
                SUM(p.gross_profit) as profit,
                AVG(p.profit_margin) as avg_margin
            FROM payroll_records p
            JOIN employees e ON p.employee_id = e.employee_id
            WHERE p.period = ?
            GROUP BY e.dispatch_company
            ORDER BY profit DESC
        """),
            (period,),
        )

        companies = []
        for row in self.cursor.fetchall():
            companies.append({
                "company": _get_row_value(row, "dispatch_company", 0),
                "employee_count": _get_row_value(row, "employee_count", 1, 0),
                "work_hours": _get_row_value(row, "total_work_hours", 2, 0),
                "revenue": _get_row_value(row, "revenue", 3, 0),
                "cost": _get_row_value(row, "cost", 4, 0),
                "profit": _get_row_value(row, "profit", 5, 0),
                "margin": _get_row_value(row, "avg_margin", 6, 0),
            })

        # Summary totals
        self.cursor.execute(
            _q("""
            SELECT
                COUNT(DISTINCT e.dispatch_company) as company_count,
                COUNT(*) as employee_count,
                SUM(p.billing_amount) as total_revenue,
                SUM(p.total_company_cost) as total_cost,
                SUM(p.gross_profit) as total_profit,
                AVG(p.profit_margin) as avg_margin
            FROM payroll_records p
            JOIN employees e ON p.employee_id = e.employee_id
            WHERE p.period = ?
        """),
            (period,),
        )
        totals_row = self.cursor.fetchone()

        return {
            "period": period,
            "companies": companies,
            "totals": {
                "company_count": _get_row_value(totals_row, "company_count", 0, 0),
                "employee_count": _get_row_value(totals_row, "employee_count", 1, 0),
                "total_revenue": _get_row_value(totals_row, "total_revenue", 2, 0),
                "total_cost": _get_row_value(totals_row, "total_cost", 3, 0),
                "total_profit": _get_row_value(totals_row, "total_profit", 4, 0),
                "avg_margin": _get_row_value(totals_row, "avg_margin", 5, 0),
            },
            "generated_at": datetime.now().isoformat(),
        }

    def get_cost_breakdown_report_data(self, period: str) -> Dict[str, Any]:
        """Get cost breakdown report data for a period"""
        self.cursor.execute(
            _q("""
            SELECT
                p.employee_id, e.name, e.dispatch_company,
                p.gross_salary,
                p.social_insurance, p.welfare_pension, p.employment_insurance,
                p.company_social_insurance,
                p.company_employment_insurance, p.company_workers_comp,
                p.transport_allowance, p.paid_leave_amount,
                p.total_company_cost, p.billing_amount, p.gross_profit
            FROM payroll_records p
            JOIN employees e ON p.employee_id = e.employee_id
            WHERE p.period = ?
            ORDER BY e.dispatch_company, e.name
        """),
            (period,),
        )

        employees = []
        for row in self.cursor.fetchall():
            employees.append({
                "employee_id": _get_row_value(row, "employee_id", 0),
                "name": _get_row_value(row, "name", 1),
                "company": _get_row_value(row, "dispatch_company", 2),
                "gross_salary": _get_row_value(row, "gross_salary", 3, 0),
                "social_insurance": _get_row_value(row, "social_insurance", 4, 0),
                "welfare_pension": _get_row_value(row, "welfare_pension", 5, 0),
                "employment_insurance": _get_row_value(row, "employment_insurance", 6, 0),
                "company_social_insurance": _get_row_value(row, "company_social_insurance", 7, 0),
                "company_employment_insurance": _get_row_value(row, "company_employment_insurance", 8, 0),
                "company_workers_comp": _get_row_value(row, "company_workers_comp", 9, 0),
                "commuting_allowance": _get_row_value(row, "transport_allowance", 10, 0),
                "paid_leave_amount": _get_row_value(row, "paid_leave_amount", 11, 0),
                "total_cost": _get_row_value(row, "total_company_cost", 12, 0),
                "billing_amount": _get_row_value(row, "billing_amount", 13, 0),
                "gross_profit": _get_row_value(row, "gross_profit", 14, 0),
            })

        # Summary totals
        self.cursor.execute(
            _q("""
            SELECT
                SUM(gross_salary) as total_salary,
                SUM(social_insurance + welfare_pension) as total_employee_insurance,
                SUM(company_social_insurance) as total_company_insurance,
                SUM(company_employment_insurance) as total_employment_ins,
                SUM(company_workers_comp) as total_workers_comp,
                SUM(transport_allowance) as total_commuting,
                SUM(total_company_cost) as total_cost,
                SUM(billing_amount) as total_revenue,
                SUM(gross_profit) as total_profit
            FROM payroll_records
            WHERE period = ?
        """),
            (period,),
        )
        totals_row = self.cursor.fetchone()

        return {
            "period": period,
            "employees": employees,
            "totals": {
                "total_salary": _get_row_value(totals_row, "total_salary", 0, 0),
                "total_employee_insurance": _get_row_value(totals_row, "total_employee_insurance", 1, 0),
                "total_company_insurance": _get_row_value(totals_row, "total_company_insurance", 2, 0),
                "total_employment_ins": _get_row_value(totals_row, "total_employment_ins", 3, 0),
                "total_workers_comp": _get_row_value(totals_row, "total_workers_comp", 4, 0),
                "total_commuting": _get_row_value(totals_row, "total_commuting", 5, 0),
                "total_cost": _get_row_value(totals_row, "total_cost", 6, 0),
                "total_revenue": _get_row_value(totals_row, "total_revenue", 7, 0),
                "total_profit": _get_row_value(totals_row, "total_profit", 8, 0),
            },
            "generated_at": datetime.now().isoformat(),
        }

    def get_summary_report_data(self, period: str) -> Dict[str, Any]:
        """Get executive summary report data"""
        # Get monthly data
        monthly_data = self.get_monthly_report_data(period)

        # Get top 5 companies by profit
        self.cursor.execute(
            _q("""
            SELECT
                e.dispatch_company,
                COUNT(*) as employee_count,
                SUM(p.gross_profit) as profit,
                AVG(p.profit_margin) as avg_margin
            FROM payroll_records p
            JOIN employees e ON p.employee_id = e.employee_id
            WHERE p.period = ?
            GROUP BY e.dispatch_company
            ORDER BY profit DESC
            LIMIT 5
        """),
            (period,),
        )

        top_companies = []
        for row in self.cursor.fetchall():
            top_companies.append({
                "company": _get_row_value(row, "dispatch_company", 0),
                "employee_count": _get_row_value(row, "employee_count", 1, 0),
                "profit": _get_row_value(row, "profit", 2, 0),
                "margin": _get_row_value(row, "avg_margin", 3, 0),
            })

        return {
            "period": period,
            "summary": monthly_data.get("summary", {}),
            "by_company": monthly_data.get("by_company", [])[:10],
            "top_companies": top_companies,
            "top_performers": monthly_data.get("top_performers", []),
            "bottom_performers": monthly_data.get("bottom_performers", []),
            "generated_at": datetime.now().isoformat(),
        }

    def _safe_sheet_title(self, title: str) -> str:
        """Create safe worksheet title (ASCII-only, max 31 chars)"""
        # Replace Japanese year/month markers
        safe = title.replace('年', '_').replace('月', '')
        # Remove or replace any remaining non-ASCII characters
        safe = ''.join(c if ord(c) < 128 else '_' for c in safe)
        # Remove invalid Excel sheet name characters
        for char in ['/', '\\', '?', '*', '[', ']', ':']:
            safe = safe.replace(char, '_')
        # Limit to 31 characters (Excel limit)
        return safe[:31]

    def generate_excel_report(self, report_type: str, data: Dict[str, Any]) -> bytes:
        """Generate Excel report from data"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation")

        wb = Workbook()
        ws = wb.active

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        if report_type == "monthly":
            period = data.get('period', '')
            ws.title = self._safe_sheet_title(f"Monthly_{period}")
            self._write_monthly_excel(ws, data, header_font, header_fill, border)

        elif report_type == "employee":
            emp = data.get("employee", {})
            ws.title = self._safe_sheet_title(f"Employee_{emp.get('employee_id', '')}")
            self._write_employee_excel(ws, data, header_font, header_fill, border)

        elif report_type == "company":
            company = data.get('company', '')
            ws.title = self._safe_sheet_title(f"Company_{company}")
            self._write_company_excel(ws, data, header_font, header_fill, border)

        elif report_type == "all-employees":
            period = data.get('period', '')
            ws.title = self._safe_sheet_title(f"AllEmployees_{period}")
            self._write_all_employees_excel(ws, data, header_font, header_fill, border)

        elif report_type == "all-companies":
            period = data.get('period', '')
            ws.title = self._safe_sheet_title(f"AllCompanies_{period}")
            self._write_all_companies_excel(ws, data, header_font, header_fill, border)

        elif report_type == "cost-breakdown":
            period = data.get('period', '')
            ws.title = self._safe_sheet_title(f"CostBreakdown_{period}")
            self._write_cost_breakdown_excel(ws, data, header_font, header_fill, border)

        elif report_type == "summary":
            period = data.get('period', '')
            ws.title = self._safe_sheet_title(f"Summary_{period}")
            self._write_summary_excel(ws, data, header_font, header_fill, border)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _write_monthly_excel(self, ws, data, header_font, header_fill, border):
        """Write monthly report to Excel worksheet"""
        summary = data.get("summary", {})

        # Title
        ws["A1"] = f"月次粗利レポート - {data.get('period', '')}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")

        # Summary section
        ws["A3"] = "サマリー"
        ws["A3"].font = Font(bold=True)

        ws["A4"] = "従業員数"
        ws["B4"] = summary.get("employee_count") or 0
        ws["A5"] = "売上合計"
        ws["B5"] = f"¥{(summary.get('total_revenue') or 0):,.0f}"
        ws["A6"] = "コスト合計"
        ws["B6"] = f"¥{(summary.get('total_cost') or 0):,.0f}"
        ws["A7"] = "粗利合計"
        ws["B7"] = f"¥{(summary.get('total_profit') or 0):,.0f}"
        ws["A8"] = "平均マージン"
        ws["B8"] = f"{(summary.get('avg_margin') or 0):.1f}%"

        # By company section
        ws["A10"] = "派遣先別"
        ws["A10"].font = Font(bold=True)

        headers = ["派遣先", "従業員数", "売上", "コスト", "粗利", "マージン"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, company in enumerate(data.get("by_company", []), 12):
            ws.cell(row=row_idx, column=1, value=company.get("company") or "").border = (
                border
            )
            ws.cell(
                row=row_idx, column=2, value=company.get("employee_count") or 0
            ).border = border
            ws.cell(
                row=row_idx, column=3, value=f"¥{(company.get('revenue') or 0):,.0f}"
            ).border = border
            ws.cell(
                row=row_idx, column=4, value=f"¥{(company.get('cost') or 0):,.0f}"
            ).border = border
            ws.cell(
                row=row_idx, column=5, value=f"¥{(company.get('profit') or 0):,.0f}"
            ).border = border
            ws.cell(
                row=row_idx, column=6, value=f"{(company.get('margin') or 0):.1f}%"
            ).border = border

        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _write_employee_excel(self, ws, data, header_font, header_fill, border):
        """Write employee report to Excel worksheet"""
        emp = data.get("employee", {})

        # Title
        ws["A1"] = f"従業員レポート - {emp.get('name', '')}"
        ws["A1"].font = Font(bold=True, size=14)

        # Employee info
        ws["A3"] = "従業員情報"
        ws["A3"].font = Font(bold=True)

        ws["A4"] = "ID"
        ws["B4"] = emp.get("employee_id") or ""
        ws["A5"] = "名前"
        ws["B5"] = emp.get("name") or ""
        ws["A6"] = "派遣先"
        ws["B6"] = emp.get("company") or ""
        ws["A7"] = "時給"
        ws["B7"] = f"¥{(emp.get('hourly_rate') or 0):,.0f}"
        ws["A8"] = "単価"
        ws["B8"] = f"¥{(emp.get('billing_rate') or 0):,.0f}"

        # History section
        ws["A10"] = "履歴"
        ws["A10"].font = Font(bold=True)

        headers = [
            "期間",
            "労働時間",
            "残業",
            "深夜",
            "売上",
            "コスト",
            "粗利",
            "マージン",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, h in enumerate(data.get("history", []), 12):
            ws.cell(row=row_idx, column=1, value=h.get("period") or "").border = border
            ws.cell(
                row=row_idx, column=2, value=f"{(h.get('work_hours') or 0):.1f}h"
            ).border = border
            ws.cell(
                row=row_idx, column=3, value=f"{(h.get('overtime_hours') or 0):.1f}h"
            ).border = border
            ws.cell(
                row=row_idx, column=4, value=f"{(h.get('night_hours') or 0):.1f}h"
            ).border = border
            ws.cell(
                row=row_idx, column=5, value=f"¥{(h.get('billing_amount') or 0):,.0f}"
            ).border = border
            ws.cell(row=row_idx, column=6, value=f"¥{(h.get('cost') or 0):,.0f}").border = (
                border
            )
            ws.cell(
                row=row_idx, column=7, value=f"¥{(h.get('profit') or 0):,.0f}"
            ).border = border
            ws.cell(
                row=row_idx, column=8, value=f"{(h.get('margin') or 0):.1f}%"
            ).border = border

        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _write_company_excel(self, ws, data, header_font, header_fill, border):
        """Write company report to Excel worksheet"""
        # Title
        ws["A1"] = f"派遣先レポート - {data.get('company', '')}"
        ws["A1"].font = Font(bold=True, size=14)

        totals = data.get("totals", {})

        # Summary
        ws["A3"] = "サマリー"
        ws["A3"].font = Font(bold=True)

        ws["A4"] = "従業員数"
        ws["B4"] = data.get("employee_count") or 0
        ws["A5"] = "売上合計"
        ws["B5"] = f"¥{(totals.get('revenue') or 0):,.0f}"
        ws["A6"] = "粗利合計"
        ws["B6"] = f"¥{(totals.get('profit') or 0):,.0f}"
        ws["A7"] = "平均マージン"
        ws["B7"] = f"{(totals.get('avg_margin') or 0):.1f}%"

        # Monthly data
        ws["A9"] = "月次推移"
        ws["A9"].font = Font(bold=True)

        headers = ["期間", "従業員数", "売上", "コスト", "粗利", "マージン"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, m in enumerate(data.get("monthly_data", []), 11):
            ws.cell(row=row_idx, column=1, value=m.get("period") or "").border = border
            ws.cell(row=row_idx, column=2, value=m.get("employee_count") or 0).border = (
                border
            )
            ws.cell(
                row=row_idx, column=3, value=f"¥{(m.get('revenue') or 0):,.0f}"
            ).border = border
            ws.cell(row=row_idx, column=4, value=f"¥{(m.get('cost') or 0):,.0f}").border = (
                border
            )
            ws.cell(
                row=row_idx, column=5, value=f"¥{(m.get('profit') or 0):,.0f}"
            ).border = border
            ws.cell(
                row=row_idx, column=6, value=f"{(m.get('margin') or 0):.1f}%"
            ).border = border

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _write_all_employees_excel(self, ws, data, header_font, header_fill, border):
        """Write all employees report to Excel worksheet"""
        period = data.get('period', '')
        ws["A1"] = f"従業員別詳細レポート - {period}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:P1")

        totals = data.get("totals", {})

        # Summary section
        ws["A3"] = "サマリー"
        ws["A3"].font = Font(bold=True)
        ws["A4"] = "従業員数"
        ws["B4"] = totals.get("employee_count") or 0
        ws["C4"] = "売上合計"
        ws["D4"] = f"¥{(totals.get('total_revenue') or 0):,.0f}"
        ws["E4"] = "粗利合計"
        ws["F4"] = f"¥{(totals.get('total_profit') or 0):,.0f}"
        ws["G4"] = "平均マージン"
        ws["H4"] = f"{(totals.get('avg_margin') or 0):.1f}%"

        # Employee details
        ws["A6"] = "従業員別詳細"
        ws["A6"].font = Font(bold=True)

        headers = ["ID", "氏名", "派遣先", "時給", "単価", "労働時間", "残業", "60h超",
                   "深夜", "休日", "総支給", "請求額", "コスト", "粗利", "マージン"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, emp in enumerate(data.get("employees", []), 8):
            ws.cell(row=row_idx, column=1, value=emp.get("employee_id") or "").border = border
            ws.cell(row=row_idx, column=2, value=emp.get("name") or "").border = border
            ws.cell(row=row_idx, column=3, value=emp.get("company") or "").border = border
            ws.cell(row=row_idx, column=4, value=f"¥{(emp.get('hourly_rate') or 0):,.0f}").border = border
            ws.cell(row=row_idx, column=5, value=f"¥{(emp.get('billing_rate') or 0):,.0f}").border = border
            ws.cell(row=row_idx, column=6, value=f"{(emp.get('work_hours') or 0):.1f}").border = border
            ws.cell(row=row_idx, column=7, value=f"{(emp.get('overtime_hours') or 0):.1f}").border = border
            ws.cell(row=row_idx, column=8, value=f"{(emp.get('overtime_over_60h') or 0):.1f}").border = border
            ws.cell(row=row_idx, column=9, value=f"{(emp.get('night_hours') or 0):.1f}").border = border
            ws.cell(row=row_idx, column=10, value=f"{(emp.get('holiday_hours') or 0):.1f}").border = border
            ws.cell(row=row_idx, column=11, value=f"¥{(emp.get('gross_salary') or 0):,.0f}").border = border
            ws.cell(row=row_idx, column=12, value=f"¥{(emp.get('billing_amount') or 0):,.0f}").border = border
            ws.cell(row=row_idx, column=13, value=f"¥{(emp.get('total_cost') or 0):,.0f}").border = border
            ws.cell(row=row_idx, column=14, value=f"¥{(emp.get('gross_profit') or 0):,.0f}").border = border
            ws.cell(row=row_idx, column=15, value=f"{(emp.get('profit_margin') or 0):.1f}%").border = border

        for col in range(1, 16):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _write_all_companies_excel(self, ws, data, header_font, header_fill, border):
        """Write all companies report to Excel worksheet"""
        period = data.get('period', '')
        ws["A1"] = f"派遣先別分析レポート - {period}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:G1")

        totals = data.get("totals", {})

        # Summary section
        ws["A3"] = "サマリー"
        ws["A3"].font = Font(bold=True)
        ws["A4"] = "派遣先数"
        ws["B4"] = totals.get("company_count") or 0
        ws["C4"] = "従業員数"
        ws["D4"] = totals.get("employee_count") or 0
        ws["E4"] = "売上合計"
        ws["F4"] = f"¥{(totals.get('total_revenue') or 0):,.0f}"

        ws["A5"] = "コスト合計"
        ws["B5"] = f"¥{(totals.get('total_cost') or 0):,.0f}"
        ws["C5"] = "粗利合計"
        ws["D5"] = f"¥{(totals.get('total_profit') or 0):,.0f}"
        ws["E5"] = "平均マージン"
        ws["F5"] = f"{(totals.get('avg_margin') or 0):.1f}%"

        # Company details
        ws["A7"] = "派遣先別詳細"
        ws["A7"].font = Font(bold=True)

        headers = ["派遣先", "従業員数", "労働時間", "売上", "コスト", "粗利", "マージン"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, comp in enumerate(data.get("companies", []), 9):
            ws.cell(row=row_idx, column=1, value=comp.get("company") or "").border = border
            ws.cell(row=row_idx, column=2, value=comp.get("employee_count") or 0).border = border
            ws.cell(row=row_idx, column=3, value=f"{(comp.get('work_hours') or 0):,.1f}").border = border
            ws.cell(row=row_idx, column=4, value=f"¥{(comp.get('revenue') or 0):,.0f}").border = border
            ws.cell(row=row_idx, column=5, value=f"¥{(comp.get('cost') or 0):,.0f}").border = border
            ws.cell(row=row_idx, column=6, value=f"¥{(comp.get('profit') or 0):,.0f}").border = border
            ws.cell(row=row_idx, column=7, value=f"{(comp.get('margin') or 0):.1f}%").border = border

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _write_cost_breakdown_excel(self, ws, data, header_font, header_fill, border):
        """Write cost breakdown report to Excel worksheet"""
        period = data.get('period', '')
        ws["A1"] = f"コスト内訳レポート - {period}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:P1")

        totals = data.get("totals", {})

        # Summary section
        ws["A3"] = "コスト内訳サマリー"
        ws["A3"].font = Font(bold=True)
        ws["A4"] = "総支給額"
        ws["B4"] = f"¥{(totals.get('total_salary') or 0):,.0f}"
        ws["C4"] = "社保(本人)"
        ws["D4"] = f"¥{(totals.get('total_employee_insurance') or 0):,.0f}"
        ws["E4"] = "社保(会社)"
        ws["F4"] = f"¥{(totals.get('total_company_insurance') or 0):,.0f}"

        ws["A5"] = "雇用保険"
        ws["B5"] = f"¥{(totals.get('total_employment_ins') or 0):,.0f}"
        ws["C5"] = "労災保険"
        ws["D5"] = f"¥{(totals.get('total_workers_comp') or 0):,.0f}"
        ws["E5"] = "通勤費"
        ws["F5"] = f"¥{(totals.get('total_commuting') or 0):,.0f}"

        ws["A6"] = "会社総コスト"
        ws["B6"] = f"¥{(totals.get('total_cost') or 0):,.0f}"
        ws["C6"] = "売上合計"
        ws["D6"] = f"¥{(totals.get('total_revenue') or 0):,.0f}"
        ws["E6"] = "粗利合計"
        ws["F6"] = f"¥{(totals.get('total_profit') or 0):,.0f}"

        # Employee cost details
        ws["A8"] = "従業員別コスト詳細"
        ws["A8"].font = Font(bold=True)

        headers = ["ID", "氏名", "派遣先", "総支給", "健保", "厚年", "雇保",
                   "会社社保", "会社雇保", "労災", "通勤費", "有給", "総コスト", "請求", "粗利"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, emp in enumerate(data.get("employees", []), 10):
            ws.cell(row=row_idx, column=1, value=emp.get("employee_id") or "").border = border
            ws.cell(row=row_idx, column=2, value=emp.get("name") or "").border = border
            ws.cell(row=row_idx, column=3, value=emp.get("company") or "").border = border
            ws.cell(row=row_idx, column=4, value=(emp.get('gross_salary') or 0)).border = border
            ws.cell(row=row_idx, column=5, value=(emp.get('social_insurance') or 0)).border = border
            ws.cell(row=row_idx, column=6, value=(emp.get('welfare_pension') or 0)).border = border
            ws.cell(row=row_idx, column=7, value=(emp.get('employment_insurance') or 0)).border = border
            ws.cell(row=row_idx, column=8, value=(emp.get('company_social_insurance') or 0)).border = border
            ws.cell(row=row_idx, column=9, value=(emp.get('company_employment_insurance') or 0)).border = border
            ws.cell(row=row_idx, column=10, value=(emp.get('company_workers_comp') or 0)).border = border
            ws.cell(row=row_idx, column=11, value=(emp.get('commuting_allowance') or 0)).border = border
            ws.cell(row=row_idx, column=12, value=(emp.get('paid_leave_amount') or 0)).border = border
            ws.cell(row=row_idx, column=13, value=(emp.get('total_cost') or 0)).border = border
            ws.cell(row=row_idx, column=14, value=(emp.get('billing_amount') or 0)).border = border
            ws.cell(row=row_idx, column=15, value=(emp.get('gross_profit') or 0)).border = border

        for col in range(1, 16):
            ws.column_dimensions[get_column_letter(col)].width = 10

    def _write_summary_excel(self, ws, data, header_font, header_fill, border):
        """Write executive summary report to Excel worksheet"""
        period = data.get('period', '')
        ws["A1"] = f"経営サマリーレポート - {period}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")

        summary = data.get("summary", {})

        # Key metrics
        ws["A3"] = "主要指標"
        ws["A3"].font = Font(bold=True)

        ws["A4"] = "従業員数"
        ws["B4"] = summary.get("employee_count") or 0
        ws["A5"] = "売上合計"
        ws["B5"] = f"¥{(summary.get('total_revenue') or 0):,.0f}"
        ws["A6"] = "コスト合計"
        ws["B6"] = f"¥{(summary.get('total_cost') or 0):,.0f}"
        ws["A7"] = "粗利合計"
        ws["B7"] = f"¥{(summary.get('total_profit') or 0):,.0f}"
        ws["A8"] = "平均マージン"
        ws["B8"] = f"{(summary.get('avg_margin') or 0):.1f}%"

        # Top companies
        ws["A10"] = "利益上位派遣先"
        ws["A10"].font = Font(bold=True)

        headers = ["派遣先", "従業員数", "粗利", "マージン"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, comp in enumerate(data.get("top_companies", []), 12):
            ws.cell(row=row_idx, column=1, value=comp.get("company") or "").border = border
            ws.cell(row=row_idx, column=2, value=comp.get("employee_count") or 0).border = border
            ws.cell(row=row_idx, column=3, value=f"¥{(comp.get('profit') or 0):,.0f}").border = border
            ws.cell(row=row_idx, column=4, value=f"{(comp.get('margin') or 0):.1f}%").border = border

        # Top performers
        start_row = 12 + len(data.get("top_companies", [])) + 2
        ws.cell(row=start_row, column=1, value="利益率上位従業員").font = Font(bold=True)

        headers = ["氏名", "派遣先", "売上", "粗利", "マージン"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, emp in enumerate(data.get("top_performers", []), start_row + 2):
            ws.cell(row=row_idx, column=1, value=emp.get("name") or "").border = border
            ws.cell(row=row_idx, column=2, value=emp.get("company") or "").border = border
            ws.cell(row=row_idx, column=3, value=f"¥{(emp.get('revenue') or 0):,.0f}").border = border
            ws.cell(row=row_idx, column=4, value=f"¥{(emp.get('profit') or 0):,.0f}").border = border
            ws.cell(row=row_idx, column=5, value=f"{(emp.get('margin') or 0):.1f}%").border = border

        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15

    # ============== PDF GENERATION METHODS ==============

    def _get_margin_color(self, margin: float) -> colors.Color:
        """Get color based on profit margin"""
        if margin >= 15:
            return PDF_COLORS['success']
        elif margin >= 10:
            return PDF_COLORS['warning']
        else:
            return PDF_COLORS['danger']

    def _format_yen(self, value: float) -> str:
        """Format number as Japanese Yen"""
        if value is None:
            return "¥0"
        if abs(value) >= 1000000:
            return f"¥{value/1000000:.1f}M"
        elif abs(value) >= 1000:
            return f"¥{value/1000:.0f}K"
        return f"¥{value:,.0f}"

    def _create_bar_chart(self, data: List[Dict], width: float = 400, height: float = 200) -> Drawing:
        """Create a bar chart for company profits"""
        drawing = Drawing(width, height)

        if not data:
            return drawing

        chart = VerticalBarChart()
        chart.x = 50
        chart.y = 30
        chart.width = width - 80
        chart.height = height - 60

        # Prepare data
        profits = [d.get('profit', 0) / 1000000 for d in data[:5]]  # Convert to millions
        labels = [d.get('company', '')[:8] for d in data[:5]]  # Truncate labels

        chart.data = [profits]
        chart.categoryAxis.categoryNames = labels
        chart.categoryAxis.labels.fontName = JAPANESE_FONT
        chart.categoryAxis.labels.fontSize = 8
        chart.categoryAxis.labels.angle = 0

        chart.valueAxis.valueMin = 0
        chart.valueAxis.valueMax = max(profits) * 1.2 if profits else 100
        chart.valueAxis.labelTextFormat = '%.1fM'
        chart.valueAxis.labels.fontName = JAPANESE_FONT
        chart.valueAxis.labels.fontSize = 8

        chart.bars[0].fillColor = PDF_COLORS['primary']

        drawing.add(chart)
        return drawing

    def _create_kpi_table(self, metrics: List[Dict]) -> Table:
        """Create KPI cards as a table"""
        data = []
        row1 = []  # Labels
        row2 = []  # Values

        for m in metrics:
            row1.append(m.get('label', ''))
            row2.append(m.get('value', ''))

        data = [row1, row2]

        table = Table(data, colWidths=[120] * len(metrics))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), PDF_COLORS['primary']),
            ('TEXTCOLOR', (0, 0), (-1, 0), PDF_COLORS['white']),
            ('BACKGROUND', (0, 1), (-1, 1), PDF_COLORS['light']),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), JAPANESE_FONT),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, 1), 14),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BOX', (0, 0), (-1, -1), 1, PDF_COLORS['primary']),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, PDF_COLORS['light']),
        ]))
        return table

    def generate_pdf_report(self, report_type: str, data: Dict[str, Any]) -> bytes:
        """Generate professional PDF report"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF generation. Install with: pip install reportlab")

        output = BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )

        styles = getSampleStyleSheet()

        # Custom styles with Japanese font support
        styles.add(ParagraphStyle(
            'Title_Custom',
            parent=styles['Title'],
            fontName=JAPANESE_FONT,
            fontSize=18,
            spaceAfter=12,
            textColor=PDF_COLORS['primary']
        ))
        styles.add(ParagraphStyle(
            'Subtitle_Custom',
            parent=styles['Normal'],
            fontName=JAPANESE_FONT,
            fontSize=12,
            spaceAfter=20,
            textColor=PDF_COLORS['secondary']
        ))
        styles.add(ParagraphStyle(
            'Section_Header',
            parent=styles['Heading2'],
            fontName=JAPANESE_FONT,
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
            textColor=PDF_COLORS['primary']
        ))
        # Override Normal style for Japanese
        styles['Normal'].fontName = JAPANESE_FONT

        story = []

        if report_type == "summary":
            self._write_summary_pdf(story, data, styles)
        elif report_type == "monthly":
            self._write_monthly_pdf(story, data, styles)
        elif report_type == "all-companies":
            self._write_companies_pdf(story, data, styles)
        else:
            # Default: basic PDF with data dump
            story.append(Paragraph(f"Report: {report_type}", styles['Title_Custom']))
            story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Subtitle_Custom']))

        doc.build(story)
        output.seek(0)
        return output.getvalue()

    def _write_summary_pdf(self, story: List, data: Dict, styles) -> None:
        """Write executive summary PDF"""
        period = data.get('period', '')
        summary = data.get('summary', {})

        # Title
        story.append(Paragraph("Arari PRO - Executive Summary", styles['Title_Custom']))
        story.append(Paragraph(f"Period: {period} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Subtitle_Custom']))

        # KPI Cards
        kpis = [
            {'label': 'Employees', 'value': str(summary.get('employee_count', 0))},
            {'label': 'Revenue', 'value': self._format_yen(summary.get('total_revenue', 0))},
            {'label': 'Cost', 'value': self._format_yen(summary.get('total_cost', 0))},
            {'label': 'Profit', 'value': self._format_yen(summary.get('total_profit', 0))},
        ]
        story.append(self._create_kpi_table(kpis))
        story.append(Spacer(1, 10*mm))

        # Margin indicator
        margin = summary.get('avg_margin', 0) or 0
        margin_color = self._get_margin_color(margin)
        margin_text = f"Average Margin: {margin:.1f}%"
        story.append(Paragraph(f"<font color='{margin_color.hexval()}'><b>{margin_text}</b></font>", styles['Normal']))
        story.append(Spacer(1, 10*mm))

        # Top Companies Chart
        story.append(Paragraph("Top Companies by Profit", styles['Section_Header']))
        by_company = data.get('by_company', [])[:5]
        if by_company:
            chart = self._create_bar_chart(by_company)
            story.append(chart)
            story.append(Spacer(1, 5*mm))

        # Top Companies Table
        if by_company:
            table_data = [['Company', 'Employees', 'Revenue', 'Profit', 'Margin']]
            for comp in by_company:
                margin_val = comp.get('margin', 0) or 0
                table_data.append([
                    comp.get('company', '')[:15],
                    str(comp.get('employee_count', 0)),
                    self._format_yen(comp.get('revenue', 0)),
                    self._format_yen(comp.get('profit', 0)),
                    f"{margin_val:.1f}%"
                ])

            table = Table(table_data, colWidths=[100, 60, 80, 80, 60])
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), PDF_COLORS['primary']),
                ('TEXTCOLOR', (0, 0), (-1, 0), PDF_COLORS['white']),
                ('FONTNAME', (0, 0), (-1, 0), JAPANESE_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, PDF_COLORS['secondary']),
            ]
            # Color code margins
            for i, comp in enumerate(by_company, 1):
                margin_val = comp.get('margin', 0) or 0
                color = self._get_margin_color(margin_val)
                table_style.append(('TEXTCOLOR', (4, i), (4, i), color))

            table.setStyle(TableStyle(table_style))
            story.append(table)
            story.append(Spacer(1, 10*mm))

        # Top/Bottom Performers
        story.append(Paragraph("Top Performers", styles['Section_Header']))
        top_performers = data.get('top_performers', [])[:5]
        if top_performers:
            perf_data = [['Name', 'Company', 'Profit', 'Margin']]
            for emp in top_performers:
                margin_val = emp.get('margin', 0) or 0
                perf_data.append([
                    emp.get('name', '')[:12],
                    emp.get('company', '')[:10],
                    self._format_yen(emp.get('profit', 0)),
                    f"{margin_val:.1f}%"
                ])

            perf_table = Table(perf_data, colWidths=[100, 100, 80, 60])
            perf_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), PDF_COLORS['success']),
                ('TEXTCOLOR', (0, 0), (-1, 0), PDF_COLORS['white']),
                ('FONTNAME', (0, 0), (-1, 0), JAPANESE_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, PDF_COLORS['secondary']),
            ]))
            story.append(perf_table)

        # Bottom Performers (attention needed)
        bottom_performers = data.get('bottom_performers', [])[:5]
        if bottom_performers:
            story.append(Spacer(1, 5*mm))
            story.append(Paragraph("Attention Needed (Low Margin)", styles['Section_Header']))

            bottom_data = [['Name', 'Company', 'Profit', 'Margin']]
            for emp in bottom_performers:
                margin_val = emp.get('margin', 0) or 0
                bottom_data.append([
                    emp.get('name', '')[:12],
                    emp.get('company', '')[:10],
                    self._format_yen(emp.get('profit', 0)),
                    f"{margin_val:.1f}%"
                ])

            bottom_table = Table(bottom_data, colWidths=[100, 100, 80, 60])
            bottom_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), PDF_COLORS['danger']),
                ('TEXTCOLOR', (0, 0), (-1, 0), PDF_COLORS['white']),
                ('FONTNAME', (0, 0), (-1, 0), JAPANESE_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, PDF_COLORS['secondary']),
            ]))
            story.append(bottom_table)

    def _write_monthly_pdf(self, story: List, data: Dict, styles) -> None:
        """Write monthly profit report PDF"""
        period = data.get('period', '')
        summary = data.get('summary', {})

        # Title
        story.append(Paragraph("Monthly Profit Report", styles['Title_Custom']))
        story.append(Paragraph(f"Period: {period} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Subtitle_Custom']))

        # KPI Cards
        kpis = [
            {'label': 'Employees', 'value': str(summary.get('employee_count', 0))},
            {'label': 'Revenue', 'value': self._format_yen(summary.get('total_revenue', 0))},
            {'label': 'Cost', 'value': self._format_yen(summary.get('total_cost', 0))},
            {'label': 'Profit', 'value': self._format_yen(summary.get('total_profit', 0))},
        ]
        story.append(self._create_kpi_table(kpis))
        story.append(Spacer(1, 5*mm))

        # Hours breakdown
        hours_kpis = [
            {'label': 'Work Hours', 'value': f"{summary.get('total_work_hours', 0) or 0:,.0f}h"},
            {'label': 'Overtime', 'value': f"{summary.get('total_overtime', 0) or 0:,.0f}h"},
            {'label': 'Night', 'value': f"{summary.get('total_night_hours', 0) or 0:,.0f}h"},
            {'label': 'Holiday', 'value': f"{summary.get('total_holiday_hours', 0) or 0:,.0f}h"},
        ]
        story.append(self._create_kpi_table(hours_kpis))
        story.append(Spacer(1, 10*mm))

        # By Company breakdown
        story.append(Paragraph("Breakdown by Company", styles['Section_Header']))
        by_company = data.get('by_company', [])

        if by_company:
            # Chart
            chart = self._create_bar_chart(by_company[:8])
            story.append(chart)
            story.append(Spacer(1, 5*mm))

            # Table
            table_data = [['Company', 'Employees', 'Revenue', 'Cost', 'Profit', 'Margin']]
            for comp in by_company:
                margin_val = comp.get('margin', 0) or 0
                table_data.append([
                    comp.get('company', '')[:15],
                    str(comp.get('employee_count', 0)),
                    self._format_yen(comp.get('revenue', 0)),
                    self._format_yen(comp.get('cost', 0)),
                    self._format_yen(comp.get('profit', 0)),
                    f"{margin_val:.1f}%"
                ])

            table = Table(table_data, colWidths=[100, 55, 70, 70, 70, 50])
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), PDF_COLORS['primary']),
                ('TEXTCOLOR', (0, 0), (-1, 0), PDF_COLORS['white']),
                ('FONTNAME', (0, 0), (-1, 0), JAPANESE_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, PDF_COLORS['secondary']),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [PDF_COLORS['white'], PDF_COLORS['light']]),
            ]
            # Color code margins
            for i, comp in enumerate(by_company, 1):
                margin_val = comp.get('margin', 0) or 0
                color = self._get_margin_color(margin_val)
                table_style.append(('TEXTCOLOR', (5, i), (5, i), color))

            table.setStyle(TableStyle(table_style))
            story.append(table)

    def _write_companies_pdf(self, story: List, data: Dict, styles) -> None:
        """Write all companies analysis PDF"""
        period = data.get('period', '')
        totals = data.get('totals', {})

        # Title
        story.append(Paragraph("Company Analysis Report", styles['Title_Custom']))
        story.append(Paragraph(f"Period: {period} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Subtitle_Custom']))

        # KPI Cards
        kpis = [
            {'label': 'Companies', 'value': str(totals.get('company_count', 0))},
            {'label': 'Employees', 'value': str(totals.get('employee_count', 0))},
            {'label': 'Revenue', 'value': self._format_yen(totals.get('total_revenue', 0))},
            {'label': 'Profit', 'value': self._format_yen(totals.get('total_profit', 0))},
        ]
        story.append(self._create_kpi_table(kpis))
        story.append(Spacer(1, 10*mm))

        # Chart
        story.append(Paragraph("Profit by Company", styles['Section_Header']))
        companies = data.get('companies', [])
        if companies:
            chart = self._create_bar_chart(companies[:8])
            story.append(chart)
            story.append(Spacer(1, 5*mm))

            # Full table
            table_data = [['Company', 'Employees', 'Hours', 'Revenue', 'Cost', 'Profit', 'Margin']]
            for comp in companies:
                margin_val = comp.get('margin', 0) or 0
                table_data.append([
                    comp.get('company', '')[:15],
                    str(comp.get('employee_count', 0)),
                    f"{comp.get('work_hours', 0) or 0:,.0f}",
                    self._format_yen(comp.get('revenue', 0)),
                    self._format_yen(comp.get('cost', 0)),
                    self._format_yen(comp.get('profit', 0)),
                    f"{margin_val:.1f}%"
                ])

            table = Table(table_data, colWidths=[90, 50, 55, 65, 65, 65, 45])
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), PDF_COLORS['primary']),
                ('TEXTCOLOR', (0, 0), (-1, 0), PDF_COLORS['white']),
                ('FONTNAME', (0, 0), (-1, 0), JAPANESE_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, PDF_COLORS['secondary']),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [PDF_COLORS['white'], PDF_COLORS['light']]),
            ]
            # Color margins
            for i, comp in enumerate(companies, 1):
                margin_val = comp.get('margin', 0) or 0
                color = self._get_margin_color(margin_val)
                table_style.append(('TEXTCOLOR', (6, i), (6, i), color))

            table.setStyle(TableStyle(table_style))
            story.append(table)

    def log_report_generation(
        self,
        report_type: str,
        format: str,
        period: str = None,
        entity_type: str = None,
        entity_id: str = None,
        generated_by: str = None,
        file_size: int = None,
        params: Dict = None,
    ) -> int:
        """Log a generated report"""
        report_name = f"{report_type}_{period or 'all'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        self.cursor.execute(
            _q("""
            INSERT INTO generated_reports (
                report_type, report_name, period, entity_type, entity_id,
                format, generated_by, file_size, parameters
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """),
            (
                report_type,
                report_name,
                period,
                entity_type,
                entity_id,
                format,
                generated_by,
                file_size,
                json.dumps(params) if params else None,
            ),
        )
        self.conn.commit()

        return self.cursor.lastrowid

    def get_report_history(self, limit: int = 50) -> List[Dict[str, Any]]:
        """Get history of generated reports"""
        try:
            self.cursor.execute(
                _q("""
                SELECT id, report_type, report_name, period, format,
                       generated_by, file_size, created_at
                FROM generated_reports
                ORDER BY created_at DESC
                LIMIT ?
            """),
                (limit,),
            )
        except Exception:
            # Table might not exist yet, return empty list
            return []

        return [
            {
                "id": _get_row_value(row, "id", 0),
                "type": _get_row_value(row, "report_type", 1),
                "name": _get_row_value(row, "report_name", 2),
                "period": _get_row_value(row, "period", 3),
                "format": _get_row_value(row, "format", 4),
                "generated_by": _get_row_value(row, "generated_by", 5),
                "file_size": _get_row_value(row, "file_size", 6),
                "created_at": _get_row_value(row, "created_at", 7),
            }
            for row in self.cursor.fetchall()
        ]
